from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Criar workbook
wb = Workbook()
ws = wb.active
ws.title = "Processo 1"

# Estilos
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
label_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
white_font = Font(color="FFFFFF", bold=True, size=11)
bold_font = Font(bold=True, size=11)
normal_font = Font(size=10)
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# ==================== LINHA 1 - BOTÃO VOLTAR ====================
ws['A1'] = "Voltar para página inicial"
ws['A1'].font = Font(color="0563C1", underline="single", size=11)
ws['A1'].alignment = Alignment(horizontal='left', vertical='center')

# ==================== LINHA 2 - INFO PROJETO (A2:G2) ====================
ws['A2'] = "Projeto/Demanda:"
ws['A2'].fill = label_fill
ws['A2'].font = bold_font
ws['B2'] = "Mapeamento Processos GGOV"
ws['C2'] = "Descrição:"
ws['C2'].fill = label_fill
ws['C2'].font = bold_font
ws['D2'] = "Realizar o mapeamento completo dos processos administrativos e operacionais do Gabinete de Governança (GGOV)"
ws['E2'] = "Unidade Demandante:"
ws['E2'].fill = label_fill
ws['E2'].font = bold_font
ws['F2'] = "Gabinete de Governança"

# ==================== LINHA 3 - INFO PROJETO (A3:H3) ====================
ws['A3'] = "Número SEI:"
ws['A3'].fill = label_fill
ws['A3'].font = bold_font
ws['B3'] = "244466666"
ws['C3'] = "Categoria:"
ws['C3'].fill = label_fill
ws['C3'].font = bold_font
ws['D3'] = "Remodelagem de processos"
ws['E3'] = "Prioridade:"
ws['E3'].fill = label_fill
ws['E3'].font = bold_font
ws['F3'] = "Alta"
ws['G3'] = "Responsável Demanda:"
ws['G3'].fill = label_fill
ws['G3'].font = bold_font
ws['H3'] = "Luma e Suerlei"

# ==================== LINHA 4 - INFO PROJETO (A4:G4) ====================
ws['A4'] = "Data Início:"
ws['A4'].fill = label_fill
ws['A4'].font = bold_font
ws['B4'] = "1/12/2025"
ws['C4'] = "Data Término:"
ws['C4'].fill = label_fill
ws['C4'].font = bold_font
ws['D4'] = "11/12/2025"
ws['E4'] = "Duração (dias):"
ws['E4'].fill = label_fill
ws['E4'].font = bold_font
ws['F4'] = 10
ws['G4'] = "Dias Restantes:"
ws['G4'].fill = label_fill
ws['G4'].font = bold_font
ws['H4'] = 0

# ==================== LINHA 5 - INFO PROJETO (A5:C5) ====================
ws['A5'] = "Progresso Geral:"
ws['A5'].fill = label_fill
ws['A5'].font = bold_font
ws['B5'] = "16,67%"
ws['C5'] = "Status Atual:"
ws['C5'].fill = label_fill
ws['C5'].font = bold_font
ws['D5'] = "Paralisado"

# ==================== LINHA 7 - CABEÇALHO ETAPAS ====================
headers = [
    "Etapas", "Status", "Responsável", "Dt. Início", "Dt. Término", 
    "Produtos", "Dependências", "Progresso%", "Horas Est.", "Horas Reais", 
    "Peso%", "Situação", "Tarefas", "Observações"
]

for col_num, header in enumerate(headers, 1):
    cell = ws.cell(row=7, column=col_num)
    cell.value = header
    cell.fill = header_fill
    cell.font = white_font
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = border

# ==================== LINHA 8+ - DADOS DAS ETAPAS (EXEMPLOS) ====================
etapas_exemplo = [
    ["Levantamento de Informações", "Concluído", "Luma Damon", "8/12/2025", "11/12/2025", 
     "Plano do projeto", "-", 100, 40, 38, 20, "No prazo", "1. Entrevistas\n2. Coleta docs", ""],
    ["Mapeamento de Processos", "Paralisado", "Suerlei Gondim", "10/12/2025", "31/01/2026", 
     "Relatório Levantamento", "-", 20, 80, 15, 30, "Bloqueado", "1. Análise de fluxos\n2. Documentação", "Aguardando a Nilda validar o Processo SEI"],
    ["Análise de Processos", "Não iniciado", "", "", "", 
     "", "-", 0, 60, 0, 25, "", "", ""],
    ["Documentação e Relatório Final", "Não iniciado", "", "", "", 
     "", "-", 0, 40, 0, 15, "", "", ""],
    ["Validação e Aprovação", "Não iniciado", "", "", "", 
     "", "-", 0, 20, 0, 5, "", "", ""],
    ["Entrega e Implementação", "Não iniciado", "", "", "", 
     "", "-", 0, 30, 0, 5, "", "", ""]
]

for row_num, etapa in enumerate(etapas_exemplo, 8):
    for col_num, value in enumerate(etapa, 1):
        cell = ws.cell(row=row_num, column=col_num)
        cell.value = value
        cell.font = normal_font
        cell.border = border
        if col_num in [8, 9, 10, 11]:  # Colunas numéricas
            cell.alignment = Alignment(horizontal='center', vertical='center')
        elif col_num in [13]:  # Tarefas (permite quebra de linha)
            cell.alignment = Alignment(wrap_text=True, vertical='top')

# ==================== LINHA 17 - CABEÇALHO TAREFAS ====================
ws['A16'] = ""  # Linha em branco

tarefas_headers = [
    "Etapa", "Nome da Tarefa", "Status", "Responsável", 
    "Prioridade", "Prazo", "Progresso%", "Horas", "Observações"
]

for col_num, header in enumerate(tarefas_headers, 1):
    cell = ws.cell(row=17, column=col_num)
    cell.value = header
    cell.fill = header_fill
    cell.font = white_font
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = border

# ==================== LINHA 18+ - DADOS DAS TAREFAS (EXEMPLOS) ====================
tarefas_exemplo = [
    ["Levantamento de Informações", "Realizar entrevistas com gestores", "Concluído", "Luma Damon", 
     "Alta", "10/12/2025", 100, 20, ""],
    ["Levantamento de Informações", "Coletar documentos dos processos atuais", "Concluído", "Luma Damon", 
     "Alta", "11/12/2025", 100, 18, ""],
    ["Mapeamento de Processos", "Mapear fluxo atual de trabalho", "Em execução", "Suerlei Gondim", 
     "Alta", "15/12/2025", 30, 15, ""],
    ["Mapeamento de Processos", "Identificar gargalos e oportunidades", "Não iniciado", "", 
     "Média", "20/12/2025", 0, 0, ""],
]

for row_num, tarefa in enumerate(tarefas_exemplo, 18):
    for col_num, value in enumerate(tarefa, 1):
        cell = ws.cell(row=row_num, column=col_num)
        cell.value = value
        cell.font = normal_font
        cell.border = border
        if col_num in [7, 8]:  # Colunas numéricas
            cell.alignment = Alignment(horizontal='center', vertical='center')

# ==================== AJUSTAR LARGURAS DAS COLUNAS ====================
column_widths = {
    'A': 35, 'B': 15, 'C': 18, 'D': 12, 'E': 12,
    'F': 25, 'G': 15, 'H': 12, 'I': 12, 'J': 12,
    'K': 10, 'L': 15, 'M': 30, 'N': 40
}

for col, width in column_widths.items():
    ws.column_dimensions[col].width = width

# Ajustar altura das linhas
ws.row_dimensions[7].height = 30  # Cabeçalho etapas
ws.row_dimensions[17].height = 30  # Cabeçalho tarefas

# Salvar arquivo
wb.save("Modelo_Planilha_Hibrida.xlsx")
print("✅ Arquivo 'Modelo_Planilha_Hibrida.xlsx' criado com sucesso!")
