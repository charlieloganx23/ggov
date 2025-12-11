"""
Sistema de Gerenciamento de Processos LUMA - VERSÃO LIGHT ⚡
Baseado no modelo_luma.xlsx com recursos avançados
Recursos: Gráficos Avançados + KPIs Inteligentes + Alertas Visuais
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, PieChart, DoughnutChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule
from datetime import datetime, timedelta

wb = Workbook()

border_thin = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

border_thick = Border(
    left=Side(style='medium'), right=Side(style='medium'),
    top=Side(style='medium'), bottom=Side(style='medium')
)

# ==================== ABA 1: PÁGINA INICIAL (Resumo Inteligente) ====================
ws_inicial = wb.active
ws_inicial.title = "Página inicial"

# Título
ws_inicial.merge_cells('A1:L1')
ws_inicial['A1'] = "🎯 GABINETE DE GOVERNANÇA - GERENCIAMENTO DE PROCESSOS (VERSÃO LIGHT)"
ws_inicial['A1'].font = Font(size=14, bold=True, color="FFFFFF")
ws_inicial['A1'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
ws_inicial['A1'].alignment = Alignment(horizontal="center", vertical="center")
ws_inicial.row_dimensions[1].height = 25

# Cabeçalhos
headers_inicial = ["Processo", "Status", "Responsável", "Data de Início", 
                   "Data de Término", "% Concluído", "Tempo Estimado (dias)", 
                   "Tempo Real (dias)", "Dias Restantes", "Risco", "Saúde", "Link"]

header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True, size=11)

for col_num, header in enumerate(headers_inicial, 1):
    cell = ws_inicial.cell(row=2, column=col_num)
    cell.value = header
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border_thin

ws_inicial.row_dimensions[2].height = 30

# Dados de exemplo dos processos
processos_exemplo = [
    ["Mapeamento Processos GGOV", "='Processo 1'!H2", "Luma Damon / Suerlei Gondim", 
     datetime(2025, 12, 10), datetime(2026, 1, 31), "='Processo 1'!K2", 
     "=G3-F3", "=HOJE()-F3", "=G3-HOJE()",
     "=SE(I3<0,\"🔴 Alto\",SE(I3<5,\"🟡 Médio\",\"🟢 Baixo\"))",
     "=SE(F3>=0.8,\"🟢\",SE(F3>=0.5,\"🟡\",\"🔴\"))", "Processo 1"],
    
    ["Atualização Normativas Internas", "Em execução", "Maria Santos", 
     datetime(2025, 12, 5), datetime(2025, 12, 20), 0.65, 
     "=G4-F4", "=HOJE()-F4", "=G4-HOJE()",
     "=SE(I4<0,\"🔴 Alto\",SE(I4<5,\"🟡 Médio\",\"🟢 Baixo\"))",
     "=SE(F4>=0.8,\"🟢\",SE(F4>=0.5,\"🟡\",\"🔴\"))", ""],
    
    ["Implantação Sistema GED", "Planejada", "Carlos Pereira", 
     datetime(2025, 12, 15), datetime(2026, 2, 15), 0.15, 
     "=G5-F5", "=HOJE()-F5", "=G5-HOJE()",
     "=SE(I5<0,\"🔴 Alto\",SE(I5<5,\"🟡 Médio\",\"🟢 Baixo\"))",
     "=SE(F5>=0.8,\"🟢\",SE(F5>=0.5,\"🟡\",\"🔴\"))", ""],
    
    ["Auditoria Processos Internos", "Em execução", "Ana Oliveira", 
     datetime(2025, 12, 1), datetime(2025, 12, 18), 0.80, 
     "=G6-F6", "=HOJE()-F6", "=G6-HOJE()",
     "=SE(I6<0,\"🔴 Alto\",SE(I6<5,\"🟡 Médio\",\"🟢 Baixo\"))",
     "=SE(F6>=0.8,\"🟢\",SE(F6>=0.5,\"🟡\",\"🔴\"))", ""],
    
    ["Capacitação Equipe Governança", "Concluída", "Luís Almeida", 
     datetime(2025, 11, 20), datetime(2025, 12, 5), 1.00, 
     "=G7-F7", "=HOJE()-F7", "=G7-HOJE()",
     "=SE(I7<0,\"🔴 Alto\",SE(I7<5,\"🟡 Médio\",\"🟢 Baixo\"))",
     "=SE(F7>=0.8,\"🟢\",SE(F7>=0.5,\"🟡\",\"🔴\"))", ""],
]

for row_num, row_data in enumerate(processos_exemplo, 3):
    for col_num, value in enumerate(row_data, 1):
        cell = ws_inicial.cell(row=row_num, column=col_num)
        cell.value = value
        cell.border = border_thin
        cell.alignment = Alignment(horizontal="center", vertical="center")
        
        if col_num in [4, 5] and isinstance(value, datetime):
            cell.number_format = 'DD/MM/YYYY'
        elif col_num == 6:
            cell.number_format = '0%'

# Formatação condicional
ws_inicial.conditional_formatting.add('B3:B100',
    CellIsRule(operator='equal', formula=['"Concluída"'], 
               fill=PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
               font=Font(color="006100", bold=True))
)

ws_inicial.conditional_formatting.add('B3:B100',
    CellIsRule(operator='equal', formula=['"Em execução"'], 
               fill=PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
               font=Font(color="9C6500", bold=True))
)

ws_inicial.conditional_formatting.add('B3:B100',
    CellIsRule(operator='equal', formula=['"Planejada"'], 
               fill=PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid"),
               font=Font(color="1F4E78", bold=True))
)

# Escala de cores para % Concluído
ws_inicial.conditional_formatting.add('F3:F100',
    ColorScaleRule(start_type='num', start_value=0, start_color='F8696B',
                   mid_type='num', mid_value=0.5, mid_color='FFEB84',
                   end_type='num', end_value=1, end_color='63BE7B')
)

# Validações
dv_status = DataValidation(type="list", formula1='"Planejada,Em execução,Concluída,Cancelada,Bloqueada"')
ws_inicial.add_data_validation(dv_status)
dv_status.add('B3:B100')

column_widths = [30, 15, 25, 16, 16, 12, 18, 15, 15, 14, 8, 12]
for i, width in enumerate(column_widths, 1):
    ws_inicial.column_dimensions[get_column_letter(i)].width = width

# ==================== ABA 2: PROCESSO 1 (Detalhamento) ====================
ws_proc1 = wb.create_sheet("Processo 1")

# Cabeçalho do projeto
ws_proc1.merge_cells('A1:A2')
ws_proc1['A1'] = "Projeto\nMapeamento dos processos do Gabinete de Governança"
ws_proc1['A1'].font = Font(bold=True, size=11)
ws_proc1['A1'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
ws_proc1['A1'].border = border_thin

ws_proc1.merge_cells('B1:B2')
ws_proc1['B1'] = "Descrição da demanda"
ws_proc1['B1'].font = Font(bold=True)
ws_proc1['B1'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
ws_proc1['B1'].border = border_thin

ws_proc1.merge_cells('C1:H2')
ws_proc1['C1'] = ("Realizar o mapeamento completo dos processos administrativos e operacionais do "
                  "Gabinete de Governança (GGOV), com a finalidade de otimizar o desempenho das "
                  "atividades e garantir maior transparência, eficiência e controle nos fluxos de trabalho.")
ws_proc1['C1'].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
ws_proc1['C1'].border = border_thin

ws_proc1['I1'] = "Status atual"
ws_proc1['I2'] = '=SE(CONT.SE(B7:B12,"Concluída")=6,"Concluída",SE(CONT.SE(B7:B12,"Em execução")>0,"Em execução","Não iniciada"))'
ws_proc1.merge_cells('I1:I2')
ws_proc1['I1'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
ws_proc1['I1'].border = border_thin
ws_proc1['I1'].font = Font(bold=True)

ws_proc1.merge_cells('J1:J2')
ws_proc1['J1'] = "Página inicial"
ws_proc1['J1'].fill = PatternFill(start_color="B6D7A8", end_color="B6D7A8", fill_type="solid")
ws_proc1['J1'].alignment = Alignment(horizontal="center", vertical="center")
ws_proc1['J1'].border = border_thin

# % Conclusão Geral (novo)
ws_proc1['K1'] = "% Conclusão"
ws_proc1['K2'] = '=MÉDIA(G7:G12)'
ws_proc1.merge_cells('K1:K2')
ws_proc1['K1'].alignment = Alignment(horizontal="center", vertical="center")
ws_proc1['K1'].border = border_thin
ws_proc1['K1'].font = Font(bold=True, size=12)
ws_proc1['K2'].number_format = '0%'

# Número SEI
ws_proc1['A3'] = "Número SEI:"
ws_proc1['A4'] = "0000000000000"
ws_proc1.merge_cells('A3:A4')
ws_proc1['A3'].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
ws_proc1['A3'].border = border_thin

# Prioridade e Categoria
ws_proc1['B3'] = "Prioridade:"
ws_proc1['B4'] = "Alta"
ws_proc1.merge_cells('B3:B4')
ws_proc1['B3'].alignment = Alignment(horizontal="center", vertical="center")
ws_proc1['B3'].border = border_thin
ws_proc1['B4'].fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
ws_proc1['B4'].font = Font(bold=True, color="FFFFFF")

ws_proc1['C3'] = "Categoria:"
ws_proc1['C4'] = "Mapeamento"
ws_proc1.merge_cells('C3:C4')
ws_proc1['C3'].alignment = Alignment(horizontal="center", vertical="center")
ws_proc1['C3'].border = border_thin

# Indicador de Risco
ws_proc1['D3'] = "Risco:"
ws_proc1['D4'] = '=SE(CONT.SE(B7:B12,"Não iniciada")>3,"🔴 Alto",SE(MÉDIA(G7:G12)<0.5,"🟡 Médio","🟢 Baixo"))'
ws_proc1.merge_cells('D3:D4')
ws_proc1['D3'].alignment = Alignment(horizontal="center", vertical="center")
ws_proc1['D3'].border = border_thin
ws_proc1['D3'].font = Font(bold=True)

# Cabeçalhos da tabela de passos
ws_proc1['A6'] = "Passos para entregar a demanda:"
ws_proc1['B6'] = "Status"
ws_proc1['C6'] = "Responsável"
ws_proc1['D6'] = "Data de início"
ws_proc1['E6'] = "Data de término"
ws_proc1['F6'] = "Produtos"
ws_proc1['G6'] = "% Progresso"
ws_proc1['H6'] = "Dias Rest."
ws_proc1['I6'] = "Alerta"
ws_proc1['J6'] = "Tarefas"

for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']:
    cell = ws_proc1[f'{col}6']
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border_thin

# Dados dos passos
passos_dados = [
    ["Levantamento de Informações", "Em execução", "Luma Damon de Oliveira Melo", 
     datetime(2025, 12, 10), datetime(2026, 1, 16), "Plano do projeto", 
     0.70, "=E7-HOJE()", "=SE(H7<0,\"🔴\",SE(H7<5,\"🟡\",\"🟢\"))",
     "1. Realizar entrevistas\n2. Coletar documentação"],
    
    ["Mapeamento de Processos", "Em execução", "Suerlei Gondim Dutra", 
     datetime(2025, 12, 10), datetime(2026, 1, 31), "Relatório de Levantamento", 
     0.60, "=E8-HOJE()", "=SE(H8<0,\"🔴\",SE(H8<5,\"🟡\",\"🟢\"))",
     "1. Documentar processos\n2. Identificar gargalos"],
    
    ["Análise de Processos", "Não iniciada", "", 
     None, datetime(2026, 2, 15), "Mapas de Processos (Diagramas)", 
     0.00, "=E9-HOJE()", "=SE(H9<0,\"🔴\",SE(H9<5,\"🟡\",\"🟢\"))",
     "1. Analisar eficiência\n2. Propor melhorias"],
    
    ["Documentação e Relatório Final", "Não iniciada", "", 
     None, datetime(2026, 2, 28), "Relatório de Análise", 
     0.00, "=E10-HOJE()", "=SE(H10<0,\"🔴\",SE(H10<5,\"🟡\",\"🟢\"))",
     "1. Consolidar documentação\n2. Preparar apresentação"],
    
    ["Validação e Aprovação", "Não iniciada", "", 
     None, datetime(2026, 3, 10), "Relatório Final Validado", 
     0.00, "=E11-HOJE()", "=SE(H11<0,\"🔴\",SE(H11<5,\"🟡\",\"🟢\"))",
     "1. Apresentar para stakeholders\n2. Ajustar feedback"],
    
    ["Entrega e Implementação", "Não iniciada", "", 
     None, datetime(2026, 3, 31), "Processos Implementados", 
     0.00, "=E12-HOJE()", "=SE(H12<0,\"🔴\",SE(H12<5,\"🟡\",\"🟢\"))",
     "1. Treinar equipe\n2. Monitorar implementação"],
]

for row_idx, passo in enumerate(passos_dados, 7):
    for col_idx, value in enumerate(passo, 1):
        cell = ws_proc1.cell(row=row_idx, column=col_idx)
        cell.value = value
        cell.border = border_thin
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        
        if col_idx in [4, 5] and isinstance(value, datetime):
            cell.number_format = 'DD/MM/YYYY'
        elif col_idx == 7:
            cell.number_format = '0%'
            cell.alignment = Alignment(horizontal="center", vertical="center")
        elif col_idx in [8, 9]:
            cell.alignment = Alignment(horizontal="center", vertical="center")

# Validações
dv_status_proc = DataValidation(type="list", formula1='"Não iniciada,Em execução,Concluída"')
ws_proc1.add_data_validation(dv_status_proc)
dv_status_proc.add('B7:B20')

# Formatação condicional nos passos
ws_proc1.conditional_formatting.add('G7:G100',
    ColorScaleRule(start_type='num', start_value=0, start_color='F8696B',
                   mid_type='num', mid_value=0.5, mid_color='FFEB84',
                   end_type='num', end_value=1, end_color='63BE7B')
)

column_widths_proc = [32, 15, 28, 16, 16, 25, 12, 10, 8, 40]
for i, width in enumerate(column_widths_proc, 1):
    ws_proc1.column_dimensions[get_column_letter(i)].width = width

ws_proc1.row_dimensions[1].height = 45
ws_proc1.row_dimensions[2].height = 45
for row in range(7, 13):
    ws_proc1.row_dimensions[row].height = 35

# ==================== ABA 3: DASHBOARD EXECUTIVO ====================
ws_dashboard = wb.create_sheet("Dashboard")

ws_dashboard.merge_cells('A1:H2')
ws_dashboard['A1'] = "📊 DASHBOARD EXECUTIVO - GABINETE DE GOVERNANÇA"
ws_dashboard['A1'].font = Font(size=18, bold=True, color="FFFFFF")
ws_dashboard['A1'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
ws_dashboard['A1'].alignment = Alignment(horizontal="center", vertical="center")

ws_dashboard['A3'] = f"📅 Atualizado em: {datetime.now().strftime('%d/%m/%Y às %H:%M')}"
ws_dashboard['A3'].font = Font(italic=True, size=10)
ws_dashboard.merge_cells('A3:H3')

# KPIs
ws_dashboard.merge_cells('A5:H5')
ws_dashboard['A5'] = "🎯 INDICADORES-CHAVE DE PERFORMANCE (KPIs)"
ws_dashboard['A5'].font = Font(size=14, bold=True, color="1F4E78")
ws_dashboard['A5'].alignment = Alignment(horizontal="center")

kpi_row = 6
kpi_data = [
    ["📋 Total Processos", "=CONT.VALORES('Página inicial'!A3:A100)", "4472D4"],
    ["🚀 Em Execução", "=CONT.SE('Página inicial'!B:B,\"Em execução\")", "FFA500"],
    ["✅ Concluídos", "=CONT.SE('Página inicial'!B:B,\"Concluída\")", "00B050"],
    ["⏸️ Bloqueados", "=CONT.SE('Página inicial'!B:B,\"Bloqueada\")", "C00000"],
    ["📈 Taxa Conclusão", "=SE(B6=0,\"0%\",TEXTO(D6/B6,\"0%\"))", "00B0F0"],
    ["💯 Saúde Média", "=MÉDIA('Página inicial'!F3:F20)", "92D050"],
    ["⚠️ Processos Risco", "=CONT.SE('Página inicial'!J:J,\"*Alto*\")", "FF0000"],
    ["⏱️ Prazo Médio", "=MÉDIA('Página inicial'!I3:I20)&\" dias\"", "7030A0"],
]

col = 1
for label, formula, color in kpi_data:
    label_cell = ws_dashboard.cell(row=kpi_row, column=col)
    label_cell.value = label
    label_cell.font = Font(bold=True, size=10, color="FFFFFF")
    label_cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    label_cell.alignment = Alignment(horizontal="center", vertical="center")
    label_cell.border = border_thin
    
    value_cell = ws_dashboard.cell(row=kpi_row+1, column=col)
    value_cell.value = formula
    value_cell.font = Font(bold=True, size=16, color=color)
    value_cell.alignment = Alignment(horizontal="center", vertical="center")
    value_cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    value_cell.border = border_thin
    
    col += 1

ws_dashboard['F7'].number_format = '0%'

# Alertas
ws_dashboard.merge_cells('A10:F10')
ws_dashboard['A10'] = "🚨 PROCESSOS QUE REQUEREM ATENÇÃO"
ws_dashboard['A10'].font = Font(size=14, bold=True, color="C00000")
ws_dashboard['A10'].alignment = Alignment(horizontal="center")

alert_headers = ["Processo", "Responsável", "Status", "% Progresso", "Dias Rest.", "Ação"]
for col, header in enumerate(alert_headers, 1):
    cell = ws_dashboard.cell(row=11, column=col)
    cell.value = header
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    cell.alignment = Alignment(horizontal="center")
    cell.border = border_thin

# Gráficos - Status
ws_dashboard['A16'] = "Status"
ws_dashboard['A16'].font = Font(bold=True)
ws_dashboard['B16'] = "Quantidade"
ws_dashboard['B16'].font = Font(bold=True)

status_labels = ["Planejada", "Em execução", "Concluída", "Bloqueada"]
for i, status in enumerate(status_labels, 17):
    ws_dashboard[f'A{i}'] = status
    ws_dashboard[f'B{i}'] = f'=CONT.SE(\'Página inicial\'!B:B,"{status}")'

donut = DoughnutChart()
labels = Reference(ws_dashboard, min_col=1, min_row=17, max_row=20)
data = Reference(ws_dashboard, min_col=2, min_row=16, max_row=20)
donut.add_data(data, titles_from_data=True)
donut.set_categories(labels)
donut.title = "📊 Distribuição por Status"
donut.style = 10
donut.height = 12
donut.width = 16
donut.dataLabels = DataLabelList()
donut.dataLabels.showPercent = True
ws_dashboard.add_chart(donut, "D16")

# Gráfico de evolução temporal
ws_dashboard['A24'] = "Período"
ws_dashboard['B24'] = "Concluídas"
ws_dashboard['C24'] = "Iniciadas"

timeline = [
    ["Semana 1", 1, 2],
    ["Semana 2", 2, 1],
    ["Semana 3", 1, 1],
    ["Semana 4", 1, 1],
]

for i, (periodo, conc, inic) in enumerate(timeline, 25):
    ws_dashboard[f'A{i}'] = periodo
    ws_dashboard[f'B{i}'] = conc
    ws_dashboard[f'C{i}'] = inic

line = LineChart()
line.title = "📈 Evolução dos Processos"
line.y_axis.title = 'Quantidade'
line.x_axis.title = 'Período'
line.style = 13
labels_time = Reference(ws_dashboard, min_col=1, min_row=25, max_row=28)
data_time = Reference(ws_dashboard, min_col=2, min_row=24, max_row=28, max_col=3)
line.add_data(data_time, titles_from_data=True)
line.set_categories(labels_time)
line.height = 12
line.width = 16
ws_dashboard.add_chart(line, "D31")

for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
    ws_dashboard.column_dimensions[col].width = 18

# ==================== ABA 4: ANÁLISE DE DESEMPENHO ====================
ws_desempenho = wb.create_sheet("Análise de Desempenho")

ws_desempenho.merge_cells('A1:F1')
ws_desempenho['A1'] = "📊 ANÁLISE DE DESEMPENHO E PRODUTIVIDADE"
ws_desempenho['A1'].font = Font(size=16, bold=True, color="FFFFFF")
ws_desempenho['A1'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
ws_desempenho['A1'].alignment = Alignment(horizontal="center", vertical="center")

# Por responsável
ws_desempenho['A3'] = "👤 DESEMPENHO POR RESPONSÁVEL"
ws_desempenho['A3'].font = Font(size=12, bold=True)
ws_desempenho.merge_cells('A3:F3')

desemp_headers = ["Responsável", "Processos Ativos", "Concluídos", "% Conclusão", "Progresso Médio", "Avaliação"]
for col, header in enumerate(desemp_headers, 1):
    cell = ws_desempenho.cell(row=4, column=col)
    cell.value = header
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    cell.alignment = Alignment(horizontal="center")
    cell.border = border_thin

responsaveis = ["Luma Damon de Oliveira Melo", "Suerlei Gondim Dutra", "Maria Santos", "Carlos Pereira", "Ana Oliveira"]
for i, resp in enumerate(responsaveis, 5):
    ws_desempenho[f'A{i}'] = resp
    ws_desempenho[f'B{i}'] = f'=CONT.SE(\'Página inicial\'!C:C,"*{resp.split()[0]}*")'
    ws_desempenho[f'C{i}'] = f'=CONT.SES(\'Página inicial\'!C:C,"*{resp.split()[0]}*",\'Página inicial\'!B:B,"Concluída")'
    ws_desempenho[f'D{i}'] = f'=SE(B{i}=0,0,C{i}/B{i})'
    ws_desempenho[f'E{i}'] = "=MÉDIASE('Página inicial'!C:C,\"*\"&A" + str(i) + "&\"*\",'Página inicial'!F:F)"
    ws_desempenho[f'F{i}'] = f'=SE(E{i}>=0.7,"🟢 Excelente",SE(E{i}>=0.5,"🟡 Bom","🔴 Atenção"))'
    
    ws_desempenho[f'D{i}'].number_format = '0%'
    ws_desempenho[f'E{i}'].number_format = '0%'

for col in ['A', 'B', 'C', 'D', 'E', 'F']:
    ws_desempenho.column_dimensions[col].width = 28

# ==================== ABA 5: INSTRUÇÕES ====================
ws_instrucoes = wb.create_sheet("Instruções")

ws_instrucoes.merge_cells('A1:D1')
ws_instrucoes['A1'] = "📖 GUIA DE USO - SISTEMA LUMA LIGHT"
ws_instrucoes['A1'].font = Font(size=16, bold=True, color="FFFFFF")
ws_instrucoes['A1'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
ws_instrucoes['A1'].alignment = Alignment(horizontal="center", vertical="center")

instrucoes = [
    "", "⚡ SISTEMA INTEGRADO GABINETE DE GOVERNANÇA - VERSÃO LIGHT", "",
    "Este sistema foi desenvolvido especificamente para o GGOV com base no modelo original,", 
    "adicionando recursos avançados de visualização, análise e alertas.", "",
    "📊 ABAS DO SISTEMA:", "",
    "1. PÁGINA INICIAL:", "   • Visão consolidada de todos os processos",
    "   • Status, responsáveis, prazos e indicadores de saúde",
    "   • Cálculos automáticos de tempo e risco", "   • Formatação por cores para identificação rápida", "",
    "2. PROCESSO 1 (Template):", "   • Detalhamento completo do projeto",
    "   • Passos com status, responsável e progresso individual",
    "   • % de conclusão calculado automaticamente", "   • Alertas visuais por etapa",
    "   • Indicadores de prioridade e risco do projeto", "",
    "3. DASHBOARD EXECUTIVO:", "   • 8 KPIs principais atualizados em tempo real",
    "   • Gráfico de rosca para status dos processos", "   • Gráfico de linha para evolução temporal",
    "   • Seção de alertas para processos críticos", "",
    "4. ANÁLISE DE DESEMPENHO:", "   • Produtividade por responsável",
    "   • Métricas de conclusão e progresso", "   • Avaliação automática de performance", "",
    "🎨 CÓDIGO DE CORES:", "", "Status:", "   🟢 Verde = Concluída",
    "   🟡 Amarelo = Em execução", "   🔵 Azul = Planejada", "   🔴 Vermelho = Bloqueada", "",
    "Indicadores:", "   🟢 = Saudável (>80%)", "   🟡 = Atenção (50-80%)", "   🔴 = Crítico (<50%)", "",
    "🚀 COMO USAR:", "", "   1. Adicione processos na Página inicial",
    "   2. Para cada processo, crie uma aba detalhada (copie Processo 1)",
    "   3. Atualize status e % de progresso regularmente",
    "   4. Monitore o Dashboard para visão estratégica",
    "   5. Use Análise de Desempenho para acompanhar equipe", "",
    "💡 Sistema desenvolvido com Python + openpyxl para o Gabinete de Governança",
]

for i, texto in enumerate(instrucoes, 2):
    ws_instrucoes[f'A{i}'] = texto
    ws_instrucoes.merge_cells(f'A{i}:D{i}')
    cell = ws_instrucoes[f'A{i}']
    
    if any(texto.startswith(x) for x in ["⚡", "📊", "🎨", "🚀", "💡"]):
        cell.font = Font(bold=True, size=12, color="1F4E78")
    elif any(texto.startswith(x) for x in ["1.", "2.", "3.", "4."]):
        cell.font = Font(bold=True, size=11)
    elif texto.startswith("   "):
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=2)
    else:
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

ws_instrucoes.column_dimensions['A'].width = 85

# Salvar
filename = "Sistema_GGOV_Luma_Light.xlsx"
wb.save(filename)

print("="*80)
print("⚡ SISTEMA LUMA LIGHT CRIADO COM SUCESSO!")
print("="*80)
print(f"\n📁 Arquivo: {filename}")
print("\n📊 BASEADO NO MODELO: modelo_luma.xlsx")
print("\n✨ RECURSOS IMPLEMENTADOS:")
print("   ✓ Estrutura adaptada do modelo original do GGOV")
print("   ✓ Página inicial com visão consolidada")
print("   ✓ Template de processo detalhado (Processo 1)")
print("   ✓ Dashboard executivo com 8 KPIs")
print("   ✓ Gráficos de rosca e linha do tempo")
print("   ✓ Análise de desempenho por responsável")
print("   ✓ Sistema de alertas visuais (🟢🟡🔴)")
print("   ✓ Formatação condicional por cores")
print("   ✓ Cálculos automáticos de prazos e riscos")
print("   ✓ % de conclusão automático por processo")
print("\n" + "="*80)
print("🎯 Pronto para uso no Gabinete de Governança!")
print("="*80)
