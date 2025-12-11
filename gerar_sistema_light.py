"""
Sistema de Gerenciamento de Demandas - VERSÃO LIGHT ⚡
Recursos: Gráficos Avançados + KPIs Inteligentes + Alertas Visuais
Criado com Python + openpyxl
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Color
from openpyxl.chart import BarChart, PieChart, DoughnutChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, IconSetRule
from openpyxl.styles.differential import DifferentialStyle
from datetime import datetime, timedelta
import random

# Criar workbook
wb = Workbook()

# Estilos padrão
border_thin = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

border_thick = Border(
    left=Side(style='medium'),
    right=Side(style='medium'),
    top=Side(style='medium'),
    bottom=Side(style='medium')
)

# ==================== ABA 1: CONTROLE DE DEMANDAS ====================
ws_demandas = wb.active
ws_demandas.title = "Controle de Demandas"

# Título principal
ws_demandas.merge_cells('A1:N1')
title = ws_demandas['A1']
title.value = "🎯 SISTEMA DE GERENCIAMENTO DE DEMANDAS - VERSÃO LIGHT"
title.font = Font(size=16, bold=True, color="FFFFFF")
title.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
title.alignment = Alignment(horizontal="center", vertical="center")
title.border = border_thick

# Cabeçalhos
headers = ["ID", "Demanda", "Responsável", "Prioridade", "Status", 
           "Data Abertura", "Prazo", "Data Conclusão", "Dias Decorridos", 
           "Dias Restantes", "% Progresso", "Categoria", "Risco", "Saúde"]

# Estilo do cabeçalho
header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True, size=11)

# Aplicar cabeçalhos
for col_num, header in enumerate(headers, 1):
    cell = ws_demandas.cell(row=2, column=col_num)
    cell.value = header
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border_thin

# Dados de exemplo expandidos
dados_exemplo = [
    [1, "Desenvolvimento do módulo de login", "João Silva", "Alta", "Em Andamento", 
     datetime(2025, 12, 1), datetime(2025, 12, 15), None, "=HOJE()-F3", "=G3-HOJE()", 
     0.60, "Desenvolvimento", "=SE(J3<0,\"🔴 Alto\",SE(J3<3,\"🟡 Médio\",\"🟢 Baixo\"))", 
     "=SE(K3>=0.8,\"🟢\",SE(K3>=0.5,\"🟡\",\"🔴\"))"],
    
    [2, "Correção de bugs no relatório", "Maria Santos", "Crítica", "Em Andamento", 
     datetime(2025, 12, 5), datetime(2025, 12, 12), None, "=HOJE()-F4", "=G4-HOJE()", 
     0.85, "Correção", "=SE(J4<0,\"🔴 Alto\",SE(J4<3,\"🟡 Médio\",\"🟢 Baixo\"))", 
     "=SE(K4>=0.8,\"🟢\",SE(K4>=0.5,\"🟡\",\"🔴\"))"],
    
    [3, "Implementar dashboard analytics", "Pedro Costa", "Média", "Planejada", 
     datetime(2025, 12, 8), datetime(2025, 12, 25), None, "=HOJE()-F5", "=G5-HOJE()", 
     0.20, "Desenvolvimento", "=SE(J5<0,\"🔴 Alto\",SE(J5<3,\"🟡 Médio\",\"🟢 Baixo\"))", 
     "=SE(K5>=0.8,\"🟢\",SE(K5>=0.5,\"🟡\",\"🔴\"))"],
    
    [4, "Atualização da documentação", "Ana Oliveira", "Baixa", "Concluída", 
     datetime(2025, 11, 20), datetime(2025, 12, 5), datetime(2025, 12, 3), "=HOJE()-F6", "=G6-HOJE()", 
     1.00, "Documentação", "=SE(J6<0,\"🔴 Alto\",SE(J6<3,\"🟡 Médio\",\"🟢 Baixo\"))", 
     "=SE(K6>=0.8,\"🟢\",SE(K6>=0.5,\"🟡\",\"🔴\"))"],
    
    [5, "Teste de integração API", "Carlos Lima", "Alta", "Em Andamento", 
     datetime(2025, 12, 7), datetime(2025, 12, 14), None, "=HOJE()-F7", "=G7-HOJE()", 
     0.45, "Teste", "=SE(J7<0,\"🔴 Alto\",SE(J7<3,\"🟡 Médio\",\"🟢 Baixo\"))", 
     "=SE(K7>=0.8,\"🟢\",SE(K7>=0.5,\"🟡\",\"🔴\"))"],
    
    [6, "Migração de banco de dados", "João Silva", "Crítica", "Em Andamento", 
     datetime(2025, 12, 3), datetime(2025, 12, 13), None, "=HOJE()-F8", "=G8-HOJE()", 
     0.70, "Infraestrutura", "=SE(J8<0,\"🔴 Alto\",SE(J8<3,\"🟡 Médio\",\"🟢 Baixo\"))", 
     "=SE(K8>=0.8,\"🟢\",SE(K8>=0.5,\"🟡\",\"🔴\"))"],
    
    [7, "Implementação de autenticação 2FA", "Maria Santos", "Alta", "Planejada", 
     datetime(2025, 12, 10), datetime(2025, 12, 22), None, "=HOJE()-F9", "=G9-HOJE()", 
     0.10, "Desenvolvimento", "=SE(J9<0,\"🔴 Alto\",SE(J9<3,\"🟡 Médio\",\"🟢 Baixo\"))", 
     "=SE(K9>=0.8,\"🟢\",SE(K9>=0.5,\"🟡\",\"🔴\"))"],
    
    [8, "Análise de performance do sistema", "Pedro Costa", "Média", "Em Andamento", 
     datetime(2025, 12, 6), datetime(2025, 12, 18), None, "=HOJE()-F10", "=G10-HOJE()", 
     0.55, "Análise", "=SE(J10<0,\"🔴 Alto\",SE(J10<3,\"🟡 Médio\",\"🟢 Baixo\"))", 
     "=SE(K10>=0.8,\"🟢\",SE(K10>=0.5,\"🟡\",\"🔴\"))"],
    
    [9, "Criação de templates de email", "Ana Oliveira", "Baixa", "Concluída", 
     datetime(2025, 11, 28), datetime(2025, 12, 8), datetime(2025, 12, 7), "=HOJE()-F11", "=G11-HOJE()", 
     1.00, "Documentação", "=SE(J11<0,\"🔴 Alto\",SE(J11<3,\"🟡 Médio\",\"🟢 Baixo\"))", 
     "=SE(K11>=0.8,\"🟢\",SE(K11>=0.5,\"🟡\",\"🔴\"))"],
    
    [10, "Configuração de ambiente de staging", "Carlos Lima", "Alta", "Bloqueada", 
     datetime(2025, 12, 4), datetime(2025, 12, 11), None, "=HOJE()-F12", "=G12-HOJE()", 
     0.30, "Infraestrutura", "=SE(J12<0,\"🔴 Alto\",SE(J12<3,\"🟡 Médio\",\"🟢 Baixo\"))", 
     "=SE(K12>=0.8,\"🟢\",SE(K12>=0.5,\"🟡\",\"🔴\"))"],
]

# Inserir dados
for row_num, row_data in enumerate(dados_exemplo, 3):
    for col_num, value in enumerate(row_data, 1):
        cell = ws_demandas.cell(row=row_num, column=col_num)
        cell.value = value
        cell.border = border_thin
        cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Formatar datas
        if col_num in [6, 7, 8] and isinstance(value, datetime):
            cell.number_format = 'DD/MM/YYYY'
        # Formatar percentual
        elif col_num == 11:
            cell.number_format = '0%'

# Ajustar largura das colunas
column_widths = [6, 35, 18, 12, 15, 14, 14, 14, 13, 13, 12, 16, 14, 8]
for i, width in enumerate(column_widths, 1):
    ws_demandas.column_dimensions[get_column_letter(i)].width = width

# Ajustar altura da linha do título
ws_demandas.row_dimensions[1].height = 25
ws_demandas.row_dimensions[2].height = 30

# Validações de dados
dv_prioridade = DataValidation(type="list", formula1='"Crítica,Alta,Média,Baixa"', allow_blank=False)
ws_demandas.add_data_validation(dv_prioridade)
dv_prioridade.add('D3:D100')

dv_status = DataValidation(type="list", formula1='"Planejada,Em Andamento,Concluída,Cancelada,Bloqueada"', allow_blank=False)
ws_demandas.add_data_validation(dv_status)
dv_status.add('E3:E100')

dv_categoria = DataValidation(type="list", formula1='"Desenvolvimento,Correção,Teste,Documentação,Infraestrutura,Análise"', allow_blank=False)
ws_demandas.add_data_validation(dv_categoria)
dv_categoria.add('L3:L100')

# 🎨 FORMATAÇÃO CONDICIONAL AVANÇADA

# Status com cores
ws_demandas.conditional_formatting.add(
    'E3:E100',
    CellIsRule(operator='equal', formula=['"Concluída"'], 
               fill=PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
               font=Font(color="006100", bold=True))
)

ws_demandas.conditional_formatting.add(
    'E3:E100',
    CellIsRule(operator='equal', formula=['"Em Andamento"'], 
               fill=PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
               font=Font(color="9C6500", bold=True))
)

ws_demandas.conditional_formatting.add(
    'E3:E100',
    CellIsRule(operator='equal', formula=['"Bloqueada"'], 
               fill=PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
               font=Font(color="9C0006", bold=True))
)

ws_demandas.conditional_formatting.add(
    'E3:E100',
    CellIsRule(operator='equal', formula=['"Planejada"'], 
               fill=PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid"),
               font=Font(color="1F4E78", bold=True))
)

# Escala de cores para % Progresso
ws_demandas.conditional_formatting.add('K3:K100',
    ColorScaleRule(start_type='num', start_value=0, start_color='F8696B',
                   mid_type='num', mid_value=0.5, mid_color='FFEB84',
                   end_type='num', end_value=1, end_color='63BE7B')
)

# Escala de cores para Dias Restantes (vermelho=negativo, verde=positivo)
ws_demandas.conditional_formatting.add('J3:J100',
    ColorScaleRule(start_type='num', start_value=-10, start_color='F8696B',
                   mid_type='num', mid_value=0, mid_color='FFEB84',
                   end_type='num', end_value=15, end_color='63BE7B')
)

# Ícones para Prioridade
ws_demandas.conditional_formatting.add('D3:D100',
    CellIsRule(operator='equal', formula=['"Crítica"'],
               fill=PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid"),
               font=Font(color="FFFFFF", bold=True))
)

ws_demandas.conditional_formatting.add('D3:D100',
    CellIsRule(operator='equal', formula=['"Alta"'],
               fill=PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid"),
               font=Font(color="FFFFFF", bold=True))
)

# ==================== ABA 2: DASHBOARD INTELIGENTE ====================
ws_dashboard = wb.create_sheet("Dashboard")

# Título
ws_dashboard.merge_cells('A1:H2')
title_dash = ws_dashboard['A1']
title_dash.value = "📊 DASHBOARD INTELIGENTE - VISÃO EXECUTIVA"
title_dash.font = Font(size=20, bold=True, color="FFFFFF")
title_dash.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
title_dash.alignment = Alignment(horizontal="center", vertical="center")

# Data de atualização
ws_dashboard['A3'] = f"📅 Última atualização: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
ws_dashboard['A3'].font = Font(italic=True, size=10)
ws_dashboard.merge_cells('A3:H3')

# ===== SEÇÃO DE KPIs INTELIGENTES =====
ws_dashboard.merge_cells('A5:H5')
ws_dashboard['A5'] = "🎯 KPIs PRINCIPAIS"
ws_dashboard['A5'].font = Font(size=14, bold=True, color="1F4E78")
ws_dashboard['A5'].alignment = Alignment(horizontal="center")

kpi_row = 6
kpi_data = [
    ["📋 Total", "=CONT.VALORES('Controle de Demandas'!A3:A100)", "4472D4"],
    ["🚀 Em Andamento", "=CONT.SE('Controle de Demandas'!E:E,\"Em Andamento\")", "FFA500"],
    ["✅ Concluídas", "=CONT.SE('Controle de Demandas'!E:E,\"Concluída\")", "00B050"],
    ["⏸️ Bloqueadas", "=CONT.SE('Controle de Demandas'!E:E,\"Bloqueada\")", "C00000"],
    ["⚠️ Atrasadas", "=CONT.SE('Controle de Demandas'!J:J,\"<0\")", "FF0000"],
    ["📈 Taxa Conclusão", "=SE(B6=0,\"0%\",TEXTO(D6/B6,\"0%\"))", "00B0F0"],
    ["⏱️ Prazo Médio", "=SE(B6=0,0,MÉDIA('Controle de Demandas'!J3:J12))&\" dias\"", "7030A0"],
    ["💯 Saúde Média", "=MÉDIA('Controle de Demandas'!K3:K12)", "92D050"],
]

col = 1
for label, formula, color in kpi_data:
    # Label
    label_cell = ws_dashboard.cell(row=kpi_row, column=col)
    label_cell.value = label
    label_cell.font = Font(bold=True, size=10, color="FFFFFF")
    label_cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    label_cell.alignment = Alignment(horizontal="center", vertical="center")
    label_cell.border = border_thin
    
    # Valor
    value_cell = ws_dashboard.cell(row=kpi_row+1, column=col)
    value_cell.value = formula
    value_cell.font = Font(bold=True, size=16, color=color)
    value_cell.alignment = Alignment(horizontal="center", vertical="center")
    value_cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    value_cell.border = border_thin
    
    col += 1

# Formatação especial para Saúde Média (percentual)
ws_dashboard['H7'].number_format = '0%'

# ===== ALERTAS VISUAIS =====
ws_dashboard.merge_cells('A10:H10')
ws_dashboard['A10'] = "🚨 ALERTAS E NOTIFICAÇÕES"
ws_dashboard['A10'].font = Font(size=14, bold=True, color="C00000")
ws_dashboard['A10'].alignment = Alignment(horizontal="center")

alert_headers = ["Tipo", "Demanda", "Responsável", "Status", "Ação Necessária"]
for col, header in enumerate(alert_headers, 1):
    cell = ws_dashboard.cell(row=11, column=col)
    cell.value = header
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    cell.alignment = Alignment(horizontal="center")
    cell.border = border_thin

# Fórmulas para alertas (exemplo - em produção usar FILTER)
alert_row = 12
alert_examples = [
    ["🔴 ATRASADA", "='Controle de Demandas'!B3", "='Controle de Demandas'!C3", 
     "='Controle de Demandas'!E3", "Reprogramar ou acelerar"],
    ["🟡 BLOQUEADA", "='Controle de Demandas'!B12", "='Controle de Demandas'!C12", 
     "='Controle de Demandas'!E12", "Resolver impedimento"],
    ["⚠️ PRAZO CURTO", "='Controle de Demandas'!B4", "='Controle de Demandas'!C4", 
     "='Controle de Demandas'!E4", "Priorizar recurso"],
]

for row_idx, alert in enumerate(alert_examples, alert_row):
    for col_idx, value in enumerate(alert, 1):
        cell = ws_dashboard.cell(row=row_idx, column=col_idx)
        cell.value = value
        cell.border = border_thin
        cell.alignment = Alignment(horizontal="left", vertical="center")
        
        # Cor de fundo para tipos de alerta
        if col_idx == 1:
            if "ATRASADA" in str(value):
                cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            elif "BLOQUEADA" in str(value):
                cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            elif "PRAZO" in str(value):
                cell.fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")

# ===== GRÁFICOS AVANÇADOS =====

# Dados para gráficos - Status
chart_row = 17
ws_dashboard['A17'] = "Status"
ws_dashboard['A17'].font = Font(bold=True)
ws_dashboard['B17'] = "Quantidade"
ws_dashboard['B17'].font = Font(bold=True)

status_labels = ["Planejada", "Em Andamento", "Concluída", "Cancelada", "Bloqueada"]
for i, status in enumerate(status_labels, 18):
    ws_dashboard[f'A{i}'] = status
    ws_dashboard[f'B{i}'] = f'=CONT.SE(\'Controle de Demandas\'!E:E,"{status}")'

# Gráfico de Rosca (Donut Chart) - Status
donut = DoughnutChart()
labels = Reference(ws_dashboard, min_col=1, min_row=18, max_row=22)
data = Reference(ws_dashboard, min_col=2, min_row=17, max_row=22)
donut.add_data(data, titles_from_data=True)
donut.set_categories(labels)
donut.title = "📊 Distribuição por Status"
donut.style = 10
donut.height = 12
donut.width = 16

# Labels com valores
donut.dataLabels = DataLabelList()
donut.dataLabels.showCatName = True
donut.dataLabels.showVal = True
donut.dataLabels.showPercent = True

ws_dashboard.add_chart(donut, "D17")

# Dados para gráfico - Prioridade
ws_dashboard['A25'] = "Prioridade"
ws_dashboard['A25'].font = Font(bold=True)
ws_dashboard['B25'] = "Quantidade"
ws_dashboard['B25'].font = Font(bold=True)

prioridade_labels = ["Crítica", "Alta", "Média", "Baixa"]
prioridade_colors = ["C00000", "FFA500", "FFFF00", "92D050"]

for i, prioridade in enumerate(prioridade_labels, 26):
    ws_dashboard[f'A{i}'] = prioridade
    ws_dashboard[f'B{i}'] = f'=CONT.SE(\'Controle de Demandas\'!D:D,"{prioridade}")'

# Gráfico de Barras Horizontal - Prioridade
bar = BarChart()
bar.type = "bar"
bar.title = "🎯 Demandas por Prioridade"
bar.y_axis.title = 'Prioridade'
bar.x_axis.title = 'Quantidade'
bar.style = 11

labels_pri = Reference(ws_dashboard, min_col=1, min_row=26, max_row=29)
data_pri = Reference(ws_dashboard, min_col=2, min_row=25, max_row=29)
bar.add_data(data_pri, titles_from_data=True)
bar.set_categories(labels_pri)
bar.height = 12
bar.width = 16

ws_dashboard.add_chart(bar, "D32")

# Dados para gráfico - Categoria
ws_dashboard['A48'] = "Categoria"
ws_dashboard['A48'].font = Font(bold=True)
ws_dashboard['B48'] = "Concluídas"
ws_dashboard['B48'].font = Font(bold=True)
ws_dashboard['C48'] = "Em Andamento"
ws_dashboard['C48'].font = Font(bold=True)
ws_dashboard['D48'] = "Pendentes"
ws_dashboard['D48'].font = Font(bold=True)

categoria_labels = ["Desenvolvimento", "Correção", "Teste", "Documentação", "Infraestrutura", "Análise"]
for i, categoria in enumerate(categoria_labels, 49):
    ws_dashboard[f'A{i}'] = categoria
    ws_dashboard[f'B{i}'] = f'=CONT.SES(\'Controle de Demandas\'!L:L,"{categoria}",\'Controle de Demandas\'!E:E,"Concluída")'
    ws_dashboard[f'C{i}'] = f'=CONT.SES(\'Controle de Demandas\'!L:L,"{categoria}",\'Controle de Demandas\'!E:E,"Em Andamento")'
    ws_dashboard[f'D{i}'] = f'=CONT.SES(\'Controle de Demandas\'!L:L,"{categoria}",\'Controle de Demandas\'!E:E,"Planejada")'

# Gráfico de Barras Empilhadas - Categoria
bar_cat = BarChart()
bar_cat.type = "col"
bar_cat.grouping = "stacked"
bar_cat.overlap = 100
bar_cat.title = "📂 Progresso por Categoria"
bar_cat.y_axis.title = 'Quantidade'
bar_cat.x_axis.title = 'Categoria'
bar_cat.style = 12

labels_cat = Reference(ws_dashboard, min_col=1, min_row=49, max_row=54)
data_cat = Reference(ws_dashboard, min_col=2, min_row=48, max_row=54, max_col=4)
bar_cat.add_data(data_cat, titles_from_data=True)
bar_cat.set_categories(labels_cat)
bar_cat.height = 13
bar_cat.width = 18

ws_dashboard.add_chart(bar_cat, "A55")

# Dados para linha do tempo - Evolução
ws_dashboard['F48'] = "Semana"
ws_dashboard['G48'] = "Concluídas"
ws_dashboard['H48'] = "Iniciadas"

timeline_data = [
    ["Sem 1", 2, 3],
    ["Sem 2", 3, 2],
    ["Sem 3", 1, 4],
    ["Sem 4", 2, 1],
]

for i, (semana, concluidas, iniciadas) in enumerate(timeline_data, 49):
    ws_dashboard[f'F{i}'] = semana
    ws_dashboard[f'G{i}'] = concluidas
    ws_dashboard[f'H{i}'] = iniciadas

# Gráfico de Linha - Evolução Temporal
line = LineChart()
line.title = "📈 Evolução ao Longo do Tempo"
line.y_axis.title = 'Quantidade'
line.x_axis.title = 'Período'
line.style = 13

labels_time = Reference(ws_dashboard, min_col=6, min_row=49, max_row=52)
data_time = Reference(ws_dashboard, min_col=7, min_row=48, max_row=52, max_col=8)
line.add_data(data_time, titles_from_data=True)
line.set_categories(labels_time)
line.height = 13
line.width = 18

ws_dashboard.add_chart(line, "F55")

# Ajustar larguras
for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
    ws_dashboard.column_dimensions[col].width = 18

# ==================== ABA 3: ANÁLISE DE RISCOS ====================
ws_riscos = wb.create_sheet("Análise de Riscos")

ws_riscos.merge_cells('A1:F1')
ws_riscos['A1'] = "⚠️ ANÁLISE DE RISCOS E SAÚDE DO PROJETO"
ws_riscos['A1'].font = Font(size=16, bold=True, color="FFFFFF")
ws_riscos['A1'].fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
ws_riscos['A1'].alignment = Alignment(horizontal="center", vertical="center")

# Matriz de Riscos
ws_riscos['A3'] = "🎯 MATRIZ DE RISCO (Impacto vs Urgência)"
ws_riscos['A3'].font = Font(size=12, bold=True)
ws_riscos.merge_cells('A3:F3')

risk_headers = ["ID", "Demanda", "Nível de Risco", "Dias Restantes", "% Progresso", "Recomendação"]
for col, header in enumerate(risk_headers, 1):
    cell = ws_riscos.cell(row=4, column=col)
    cell.value = header
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    cell.alignment = Alignment(horizontal="center")
    cell.border = border_thin

# Fórmulas para análise de risco
risk_formulas = [
    ["='Controle de Demandas'!A3", "='Controle de Demandas'!B3", "='Controle de Demandas'!M3", 
     "='Controle de Demandas'!J3", "='Controle de Demandas'!K3", 
     "=SE(D5<0,\"🔴 URGENTE: Reprogramar\",SE(E5<0.3,\"🟡 Acelerar execução\",\"🟢 Monitorar\"))"],
    ["='Controle de Demandas'!A4", "='Controle de Demandas'!B4", "='Controle de Demandas'!M4", 
     "='Controle de Demandas'!J4", "='Controle de Demandas'!K4", 
     "=SE(D6<0,\"🔴 URGENTE: Reprogramar\",SE(E6<0.3,\"🟡 Acelerar execução\",\"🟢 Monitorar\"))"],
    ["='Controle de Demandas'!A5", "='Controle de Demandas'!B5", "='Controle de Demandas'!M5", 
     "='Controle de Demandas'!J5", "='Controle de Demandas'!K5", 
     "=SE(D7<0,\"🔴 URGENTE: Reprogramar\",SE(E7<0.3,\"🟡 Acelerar execução\",\"🟢 Monitorar\"))"],
]

for row_idx, risk_row in enumerate(risk_formulas, 5):
    for col_idx, formula in enumerate(risk_row, 1):
        cell = ws_riscos.cell(row=row_idx, column=col_idx)
        cell.value = formula
        cell.border = border_thin
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        if col_idx == 5:  # % Progresso
            cell.number_format = '0%'

# Indicadores de Saúde
ws_riscos['A10'] = "💚 INDICADORES DE SAÚDE DO PROJETO"
ws_riscos['A10'].font = Font(size=12, bold=True)
ws_riscos.merge_cells('A10:D10')

health_metrics = [
    ["Métrica", "Valor Atual", "Meta", "Status"],
    ["Taxa de Conclusão no Prazo", "=CONT.SES('Controle de Demandas'!E:E,\"Concluída\",'Controle de Demandas'!J:J,\">0\")/CONT.SE('Controle de Demandas'!E:E,\"Concluída\")", "80%", 
     "=SE(B12>=0.8,\"🟢 Excelente\",SE(B12>=0.6,\"🟡 Atenção\",\"🔴 Crítico\"))"],
    ["Progresso Médio Geral", "=MÉDIA('Controle de Demandas'!K3:K12)", "70%", 
     "=SE(B13>=0.7,\"🟢 Excelente\",SE(B13>=0.5,\"🟡 Atenção\",\"🔴 Crítico\"))"],
    ["Taxa de Bloqueio", "=CONT.SE('Controle de Demandas'!E:E,\"Bloqueada\")/CONT.VALORES('Controle de Demandas'!A3:A12)", "10%", 
     "=SE(B14<=0.1,\"🟢 Excelente\",SE(B14<=0.2,\"🟡 Atenção\",\"🔴 Crítico\"))"],
    ["Demandas Críticas em Risco", "=CONT.SES('Controle de Demandas'!D:D,\"Crítica\",'Controle de Demandas'!J:J,\"<3\")", "0", 
     "=SE(B15=0,\"🟢 Excelente\",SE(B15<=2,\"🟡 Atenção\",\"🔴 Crítico\"))"],
]

for row_idx, metric in enumerate(health_metrics, 11):
    for col_idx, value in enumerate(metric, 1):
        cell = ws_riscos.cell(row=row_idx, column=col_idx)
        cell.value = value
        cell.border = border_thin
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        if row_idx == 11:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
        
        if col_idx == 2 and row_idx > 11:
            cell.number_format = '0%'

# Ajustar larguras
ws_riscos.column_dimensions['A'].width = 12
ws_riscos.column_dimensions['B'].width = 35
ws_riscos.column_dimensions['C'].width = 18
ws_riscos.column_dimensions['D'].width = 15
ws_riscos.column_dimensions['E'].width = 15
ws_riscos.column_dimensions['F'].width = 30

# ==================== ABA 4: INSTRUÇÕES ====================
ws_instrucoes = wb.create_sheet("Instruções")

ws_instrucoes.merge_cells('A1:D1')
ws_instrucoes['A1'] = "📖 GUIA DE USO - VERSÃO LIGHT"
ws_instrucoes['A1'].font = Font(size=16, bold=True, color="FFFFFF")
ws_instrucoes['A1'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
ws_instrucoes['A1'].alignment = Alignment(horizontal="center", vertical="center")

instrucoes = [
    "",
    "⚡ BEM-VINDO À VERSÃO LIGHT!",
    "",
    "Esta versão inclui recursos avançados de visualização e análise:",
    "",
    "📊 RECURSOS PRINCIPAIS:",
    "",
    "1. CONTROLE DE DEMANDAS:",
    "   • Tabela completa com todas as informações",
    "   • Cálculos automáticos de prazos e riscos",
    "   • Formatação condicional com cores inteligentes",
    "   • Indicadores visuais de saúde (🟢🟡🔴)",
    "   • Sistema de semáforo para riscos",
    "",
    "2. DASHBOARD INTELIGENTE:",
    "   • 8 KPIs principais atualizados em tempo real",
    "   • Gráfico de Rosca para distribuição de status",
    "   • Gráfico de Barras para análise de prioridades",
    "   • Gráfico Empilhado para progresso por categoria",
    "   • Gráfico de Linha para evolução temporal",
    "   • Seção de alertas e notificações críticas",
    "",
    "3. ANÁLISE DE RISCOS:",
    "   • Matriz de risco automatizada",
    "   • Recomendações inteligentes por demanda",
    "   • 4 indicadores de saúde do projeto",
    "   • Alertas de demandas críticas em risco",
    "",
    "🎨 CÓDIGO DE CORES:",
    "",
    "Status:",
    "   🟢 Verde = Concluída",
    "   🟡 Amarelo = Em Andamento",
    "   🔵 Azul = Planejada",
    "   🔴 Vermelho = Bloqueada",
    "",
    "Risco:",
    "   🟢 Baixo = Mais de 5 dias até prazo",
    "   🟡 Médio = 3-5 dias até prazo",
    "   🔴 Alto = Menos de 3 dias ou atrasada",
    "",
    "Saúde:",
    "   🟢 = Progresso > 80%",
    "   🟡 = Progresso 50-80%",
    "   🔴 = Progresso < 50%",
    "",
    "📈 DICAS DE USO:",
    "",
    "   ✓ Use as listas suspensas para garantir padronização",
    "   ✓ Os KPIs atualizam automaticamente ao modificar dados",
    "   ✓ Monitore a aba de Riscos diariamente",
    "   ✓ Cores vermelhas indicam necessidade de ação imediata",
    "   ✓ Gráficos são dinâmicos e refletem mudanças instantaneamente",
    "",
    "🚀 PRÓXIMOS PASSOS:",
    "",
    "   → Adicione novas demandas na aba 'Controle de Demandas'",
    "   → Atualize status e % de progresso regularmente",
    "   → Revise o Dashboard para insights estratégicos",
    "   → Use a Análise de Riscos para tomada de decisão",
    "",
    "💡 Desenvolvido com Python + openpyxl",
]

for i, texto in enumerate(instrucoes, 2):
    ws_instrucoes[f'A{i}'] = texto
    ws_instrucoes.merge_cells(f'A{i}:D{i}')
    cell = ws_instrucoes[f'A{i}']
    
    if texto.startswith("⚡") or texto.startswith("📊") or texto.startswith("🎨") or texto.startswith("📈") or texto.startswith("🚀"):
        cell.font = Font(bold=True, size=12, color="1F4E78")
    elif any(texto.startswith(x) for x in ["1.", "2.", "3."]):
        cell.font = Font(bold=True, size=11)
    elif texto.startswith("   "):
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=2)
    else:
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

ws_instrucoes.column_dimensions['A'].width = 90

# Salvar arquivo
filename = "Sistema_Gerenciamento_Light.xlsx"
wb.save(filename)

print("="*80)
print("⚡ SISTEMA VERSÃO LIGHT CRIADO COM SUCESSO!")
print("="*80)
print(f"\n📁 Arquivo: {filename}")
print("\n📊 RECURSOS IMPLEMENTADOS:")
print("\n   🎯 KPIs Inteligentes:")
print("      ✓ 8 indicadores principais em tempo real")
print("      ✓ Cálculos automáticos de saúde e risco")
print("      ✓ Métricas de eficiência e performance")
print("\n   📈 Gráficos Avançados:")
print("      ✓ Gráfico de Rosca (Donut) para status")
print("      ✓ Barras horizontais para prioridades")
print("      ✓ Barras empilhadas para categorias")
print("      ✓ Linha do tempo para evolução")
print("\n   🚨 Alertas Visuais:")
print("      ✓ Sistema de semáforo (🟢🟡🔴)")
print("      ✓ Formatação condicional inteligente")
print("      ✓ Escalas de cores para progresso")
print("      ✓ Notificações de demandas críticas")
print("\n   ⚠️ Análise de Riscos:")
print("      ✓ Matriz de risco automatizada")
print("      ✓ Recomendações por demanda")
print("      ✓ Indicadores de saúde do projeto")
print("\n" + "="*80)
print("🎉 Sistema pronto para uso! Abra no Excel para explorar.")
print("="*80)
