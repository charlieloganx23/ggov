"""
Planilha Excel para Sincronização com Sistema Web GGOV
Estrutura otimizada para Google Sheets API
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from datetime import datetime

wb = Workbook()

border_thin = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

# ==================== ABA: PROCESSO 1 ====================
ws = wb.active
ws.title = "Processo 1"

# ===== CABEÇALHO INFORMATIVO =====
ws.merge_cells('A1:L1')
ws['A1'] = "📂 PROCESSO 1: Mapeamento dos processos do Gabinete de Governança"
ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
ws['A1'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 35

# ===== INFORMAÇÕES DO PROJETO =====
ws.merge_cells('A2:L2')
ws['A2'] = "📋 INFORMAÇÕES DO PROJETO"
ws['A2'].font = Font(size=12, bold=True, color="1F4E78")
ws['A2'].fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
ws['A2'].alignment = Alignment(horizontal="center", vertical="center")

# Informações básicas (apenas para referência visual)
info_data = [
    ["SEI:", "0000000000000", "Prioridade:", "Alta", "Categoria:", "Mapeamento"],
    ["Data Início:", datetime(2025, 12, 10), "Data Término:", datetime(2026, 1, 31), "Orçamento:", "R$ 15.000"],
]

for row_idx, row_data in enumerate(info_data, 3):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.value = value
        cell.border = border_thin
        
        if col_idx in [1, 3, 5]:  # Labels
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        
        if isinstance(value, datetime):
            cell.number_format = 'DD/MM/YYYY'

ws.row_dimensions[3].height = 25
ws.row_dimensions[4].height = 25

# ===== DESCRIÇÃO =====
ws['A5'] = "Descrição:"
ws['A5'].font = Font(bold=True)
ws['A5'].fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
ws['A5'].border = border_thin

ws.merge_cells('B5:L5')
ws['B5'] = ("Realizar o mapeamento completo dos processos administrativos e operacionais do Gabinete de Governança (GGOV), "
            "com a finalidade de otimizar o desempenho das atividades e garantir maior transparência, eficiência e controle nos "
            "fluxos de trabalho.")
ws['B5'].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
ws['B5'].border = border_thin
ws.row_dimensions[5].height = 60

# ===== ESPAÇADOR =====
ws.merge_cells('A6:L8')
ws['A6'] = "⬇️ PREENCHA OS DADOS ABAIXO - ESTES SERÃO SINCRONIZADOS COM O SISTEMA WEB ⬇️"
ws['A6'].font = Font(size=12, bold=True, color="C00000")
ws['A6'].fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
ws['A6'].alignment = Alignment(horizontal="center", vertical="center")

# ===== INSTRUÇÕES RÁPIDAS =====
ws.merge_cells('A9:L10')
ws['A9'] = ("💡 INSTRUÇÕES: Preencha as células em BRANCO. Use os dropdowns para Status e Prioridade. "
            "Valores de % devem ser decimais (ex: 0.7 = 70%). Não altere os cabeçalhos (linha 11)!")
ws['A9'].font = Font(size=10, italic=True, color="666666")
ws['A9'].fill = PatternFill(start_color="E7F3FF", end_color="E7F3FF", fill_type="solid")
ws['A9'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

# ===== CABEÇALHOS DAS ETAPAS (LINHA 11) =====
headers_etapas = [
    "Etapa", "Status", "Responsável", "Dt. Início", "Dt. Término", 
    "Produtos/Entregas", "Dependências", "% Progresso", 
    "Horas Est.", "Horas Real", "Peso"
]

for col, header in enumerate(headers_etapas, 1):
    cell = ws.cell(row=11, column=col)
    cell.value = header
    cell.font = Font(bold=True, size=10, color="FFFFFF")
    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border_thin

ws.row_dimensions[11].height = 35

# ===== DADOS DAS ETAPAS (LINHAS 12-17) - MODELO COM 6 ETAPAS =====
etapas_modelo = [
    ["Levantamento de Informações", "Em execução", "Luma Damon de Oliveira Melo", 
     datetime(2025, 12, 10), datetime(2026, 1, 16), "Plano do projeto", 
     "-", 0.70, 80, 56, 0.15],
    
    ["Mapeamento de Processos", "Em execução", "Suerlei Gondim Dutra", 
     datetime(2025, 12, 10), datetime(2026, 1, 31), "Relatório de Levantamento\nMapas de Processos", 
     "Etapa 1", 0.60, 120, 72, 0.25],
    
    ["Análise de Processos", "Não iniciada", "Equipe GGOV", 
     datetime(2026, 1, 17), datetime(2026, 2, 15), "Análise de eficiência e gargalos", 
     "Etapa 1, 2", 0.00, 100, 0, 0.20],
    
    ["Documentação e Relatório Final", "Não iniciada", "Equipe Técnica", 
     datetime(2026, 2, 1), datetime(2026, 2, 28), "Relatório Final Consolidado", 
     "Etapa 3", 0.00, 80, 0, 0.20],
    
    ["Validação e Aprovação", "Não iniciada", "Direção GGOV", 
     datetime(2026, 2, 20), datetime(2026, 3, 10), "Aprovação formal", 
     "Etapa 4", 0.00, 40, 0, 0.10],
    
    ["Entrega e Implementação", "Não iniciada", "Equipe GGOV Completa", 
     datetime(2026, 3, 1), datetime(2026, 3, 31), "Processos implementados", 
     "Etapa 5", 0.00, 60, 0, 0.10],
]

for row_idx, etapa in enumerate(etapas_modelo, 12):
    for col_idx, value in enumerate(etapa, 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.value = value
        cell.border = border_thin
        cell.alignment = Alignment(horizontal="center" if col_idx > 3 else "left", 
                                   vertical="center", wrap_text=True)
        
        # Formatação por tipo de dado
        if col_idx in [4, 5]:  # Datas
            cell.number_format = 'DD/MM/YYYY'
        elif col_idx in [8, 11]:  # Percentuais
            cell.number_format = '0.00'
        elif col_idx in [9, 10]:  # Horas
            cell.number_format = '0'
        
        # Destaque para células editáveis
        if col_idx in [2, 3, 8, 10]:  # Status, Responsável, % Progresso, Horas Real
            cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    
    ws.row_dimensions[row_idx].height = 40

# ===== VALIDAÇÕES DE DADOS =====
# Dropdown para Status
dv_status = DataValidation(type="list", 
                            formula1='"Não iniciada,Em execução,Concluída,Bloqueada,Cancelada"',
                            allow_blank=False)
dv_status.error = 'Selecione um status válido da lista'
dv_status.errorTitle = 'Valor Inválido'
ws.add_data_validation(dv_status)
dv_status.add('B12:B17')

# ===== ESPAÇADOR ANTES DAS TAREFAS =====
ws.merge_cells('A18:L18')
ws['A18'] = ""
ws.row_dimensions[18].height = 10

ws.merge_cells('A19:L19')
ws['A19'] = "📝 TAREFAS DETALHADAS POR ETAPA"
ws['A19'].font = Font(size=12, bold=True, color="FFFFFF")
ws['A19'].fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
ws['A19'].alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[19].height = 28

ws.merge_cells('A20:L20')
ws['A20'] = "💡 Adicione tarefas específicas de cada etapa aqui. Mantenha o formato das colunas!"
ws['A20'].font = Font(size=10, italic=True, color="666666")
ws['A20'].fill = PatternFill(start_color="E7F3FF", end_color="E7F3FF", fill_type="solid")
ws['A20'].alignment = Alignment(horizontal="center", vertical="center")

# ===== CABEÇALHOS DAS TAREFAS (LINHA 21) =====
headers_tarefas = [
    "Etapa", "Tarefa", "Status", "Responsável", "Prioridade", 
    "Prazo", "% Conclusão", "Horas", "Observações"
]

for col, header in enumerate(headers_tarefas, 1):
    cell = ws.cell(row=21, column=col)
    cell.value = header
    cell.font = Font(bold=True, size=10, color="FFFFFF")
    cell.fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border_thin

ws.row_dimensions[21].height = 30

# ===== DADOS DAS TAREFAS (LINHAS 22-30) - MODELO COM 9 TAREFAS =====
tarefas_modelo = [
    ["Etapa 1", "1. Realizar entrevistas com os responsáveis de cada área", 
     "Em execução", "Luma Damon", "Alta", datetime(2025, 12, 15), 0.80, 20, "Entrevistas em andamento"],
    
    ["Etapa 1", "2. Analisar documentos existentes, como manuais e fluxos anteriores", 
     "Em execução", "Luma Damon", "Alta", datetime(2025, 12, 20), 0.70, 16, "70% dos docs revisados"],
    
    ["Etapa 1", "3. Observar e registrar as atividades nas áreas de governança", 
     "Em execução", "Luma Damon", "Média", datetime(2026, 1, 5), 0.60, 24, "Observação em campo"],
    
    ["Etapa 1", "4. Criar questionário para coletar dados com responsáveis", 
     "Concluída", "Luma Damon", "Alta", datetime(2025, 12, 12), 1.00, 8, "Questionário aplicado"],
    
    ["Etapa 1", "5. Identificar entradas, saídas e responsáveis de cada processo", 
     "Em execução", "Luma Damon", "Alta", datetime(2026, 1, 10), 0.50, 12, "50% identificado"],
    
    ["Etapa 2", "1. Documentar processos no formato AS-IS", 
     "Em execução", "Suerlei Gondim", "Alta", datetime(2026, 1, 15), 0.70, 40, "Documentação em progresso"],
    
    ["Etapa 2", "2. Criar diagramas de fluxo (BPMN)", 
     "Em execução", "Suerlei Gondim", "Alta", datetime(2026, 1, 20), 0.60, 30, "Diagramas iniciados"],
    
    ["Etapa 2", "3. Identificar gargalos e ineficiências", 
     "Não iniciada", "Suerlei Gondim", "Média", datetime(2026, 1, 25), 0.00, 25, "Aguardando mapeamento"],
    
    ["Etapa 2", "4. Consolidar relatório de levantamento", 
     "Não iniciada", "Suerlei Gondim", "Média", datetime(2026, 1, 31), 0.00, 25, "Etapa final"],
]

for row_idx, tarefa in enumerate(tarefas_modelo, 22):
    for col_idx, value in enumerate(tarefa, 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.value = value
        cell.border = border_thin
        cell.alignment = Alignment(horizontal="center" if col_idx not in [1, 2, 9] else "left", 
                                   vertical="center", wrap_text=True)
        
        # Formatação por tipo
        if col_idx == 6:  # Prazo
            cell.number_format = 'DD/MM/YYYY'
        elif col_idx == 7:  # % Conclusão
            cell.number_format = '0.00'
        elif col_idx == 8:  # Horas
            cell.number_format = '0'
        
        # Destaque para células editáveis
        if col_idx in [3, 5, 7, 9]:  # Status, Prioridade, % Conclusão, Observações
            cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    
    ws.row_dimensions[row_idx].height = 35

# Validações para tarefas
dv_status_tarefa = DataValidation(type="list", 
                                   formula1='"Não iniciada,Em execução,Concluída,Bloqueada,Cancelada"',
                                   allow_blank=False)
ws.add_data_validation(dv_status_tarefa)
dv_status_tarefa.add('C22:C30')

dv_prioridade = DataValidation(type="list", 
                               formula1='"Alta,Média,Baixa"',
                               allow_blank=False)
ws.add_data_validation(dv_prioridade)
dv_prioridade.add('E22:E30')

# ===== LEGENDA E INSTRUÇÕES FINAIS =====
ws.merge_cells('A32:L32')
ws['A32'] = "📖 LEGENDA E INSTRUÇÕES DE PREENCHIMENTO"
ws['A32'].font = Font(size=11, bold=True, color="FFFFFF")
ws['A32'].fill = PatternFill(start_color="7030A0", end_color="7030A0", fill_type="solid")
ws['A32'].alignment = Alignment(horizontal="center", vertical="center")

instrucoes = [
    ["", ""],
    ["✅ Células AMARELAS", "São editáveis - preencha conforme necessário"],
    ["📊 % Progresso/Conclusão", "Use valores decimais: 0.5 = 50%, 0.75 = 75%, 1.0 = 100%"],
    ["📅 Datas", "Use formato DD/MM/YYYY ou clique no calendário"],
    ["⚖️ Peso", "Soma total deve ser 1.0 (100%). Representa importância da etapa"],
    ["🔄 Status", "Use o dropdown para selecionar (Não iniciada, Em execução, Concluída)"],
    ["⭐ Prioridade", "Use o dropdown (Alta, Média, Baixa)"],
    ["⏱️ Horas Real", "Atualize conforme trabalho é executado"],
    ["🔗 Sincronização", "Os dados são lidos pelo sistema web automaticamente"],
    ["💾 IMPORTANTE", "Sempre salve a planilha após fazer alterações!"],
]

for row_idx, instrucao in enumerate(instrucoes, 33):
    ws[f'A{row_idx}'] = instrucao[0]
    ws[f'A{row_idx}'].font = Font(bold=True, size=9)
    ws[f'A{row_idx}'].border = border_thin
    
    ws.merge_cells(f'B{row_idx}:L{row_idx}')
    ws[f'B{row_idx}'] = instrucao[1]
    ws[f'B{row_idx}'].font = Font(size=9)
    ws[f'B{row_idx}'].alignment = Alignment(horizontal="left", vertical="center")
    ws[f'B{row_idx}'].border = border_thin

# ===== LARGURAS DAS COLUNAS =====
column_widths = [25, 50, 20, 20, 12, 30, 15, 12, 12, 12, 10, 35]
for i, width in enumerate(column_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = width

# ===== ABA DE INSTRUÇÕES =====
ws_instr = wb.create_sheet("📖 Instruções")

ws_instr.merge_cells('A1:D1')
ws_instr['A1'] = "📖 GUIA COMPLETO DE PREENCHIMENTO E SINCRONIZAÇÃO"
ws_instr['A1'].font = Font(size=16, bold=True, color="FFFFFF")
ws_instr['A1'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
ws_instr['A1'].alignment = Alignment(horizontal="center", vertical="center")
ws_instr.row_dimensions[1].height = 35

instrucoes_detalhadas = [
    "", "🎯 OBJETIVO DESTA PLANILHA", "",
    "Esta planilha foi estruturada especialmente para sincronização automática com o Sistema Web GGOV.",
    "Ao preencher os dados aqui, eles serão automaticamente refletidos no sistema web em tempo real!", "",
    
    "📋 ESTRUTURA DA PLANILHA", "",
    "A aba 'Processo 1' contém:", "",
    
    "1️⃣ SEÇÃO DE ETAPAS (Linhas 12-17):",
    "   • 6 etapas principais do processo",
    "   • Cada etapa tem: Nome, Status, Responsável, Datas, Produtos, % Progresso, Horas, Peso",
    "   • Peso: Importância da etapa (soma deve ser 1.0 = 100%)", "",
    
    "2️⃣ SEÇÃO DE TAREFAS (Linhas 22-30):",
    "   • Tarefas detalhadas de cada etapa",
    "   • Cada tarefa tem: Etapa, Descrição, Status, Responsável, Prioridade, Prazo, %, Horas", "",
    
    "✏️ COMO PREENCHER", "",
    "Células AMARELAS são editáveis:",
    "   • Status: Use o dropdown (Não iniciada, Em execução, Concluída, Bloqueada, Cancelada)",
    "   • Responsável: Digite o nome da pessoa/equipe",
    "   • % Progresso: Digite decimal (0.5 = 50%, 0.75 = 75%, 1.0 = 100%)",
    "   • Horas Real: Atualize conforme trabalho avança",
    "   • Prioridade (tarefas): Use dropdown (Alta, Média, Baixa)", "",
    
    "⚠️ NÃO ALTERE:", "",
    "   ❌ Cabeçalhos das colunas (linha 11 e 21)",
    "   ❌ Estrutura das linhas (não insira/delete linhas entre dados)",
    "   ❌ Nomes das abas ('Processo 1' é obrigatório)",
    "   ❌ Ordem das colunas", "",
    
    "🔄 SINCRONIZAÇÃO COM SISTEMA WEB", "",
    "Passo 1: Preencha/edite os dados nesta planilha",
    "Passo 2: Salve a planilha (Ctrl+S)",
    "Passo 3: Se usando Google Sheets, a sincronização é automática!",
    "Passo 4: O sistema web atualiza a cada 30 segundos ou ao clicar em 'Atualizar'", "",
    
    "📤 COMO USAR COM GOOGLE SHEETS", "",
    "1. Faça upload desta planilha para Google Drive",
    "2. Abra com Google Sheets",
    "3. Compartilhe a planilha (Qualquer pessoa com link → Leitor)",
    "4. Copie o ID da planilha (da URL)",
    "5. No sistema web, clique em ⚙️ e configure API Key + Spreadsheet ID",
    "6. Pronto! Agora está sincronizado automaticamente", "",
    
    "💡 DICAS IMPORTANTES", "",
    "✅ Sempre use valores decimais para porcentagens (0.5 não 50)",
    "✅ Mantenha os status exatos do dropdown",
    "✅ Não deixe células de status vazias",
    "✅ A soma dos pesos deve ser 1.0 (100%)",
    "✅ Salve frequentemente para não perder dados",
    "✅ Use fórmulas se quiser calcular automaticamente", "",
    
    "🆘 SOLUÇÃO DE PROBLEMAS", "",
    "❌ Dados não aparecem no sistema web:",
    "   → Verifique se salvou a planilha",
    "   → Confirme que a aba se chama exatamente 'Processo 1'",
    "   → Verifique se os dados estão nas linhas corretas (12-17 e 22-30)",
    "   → No sistema web, clique em 'Atualizar' manualmente", "",
    
    "❌ Erro ao sincronizar:",
    "   → Verifique se a planilha está compartilhada (pública ou com link)",
    "   → Confirme que o Spreadsheet ID está correto",
    "   → Verifique se a Google Sheets API está ativada", "",
    
    "📞 SUPORTE", "",
    "Se tiver dúvidas, consulte o arquivo GUIA_GOOGLE_SHEETS.md",
    "ou verifique o Console do navegador (F12) para mensagens de erro.", "",
    
    "🚀 Sistema desenvolvido para o Gabinete de Governança (GGOV)",
    "💎 Versão: 1.0 | Data: Dezembro 2025",
]

for i, texto in enumerate(instrucoes_detalhadas, 2):
    ws_instr[f'A{i}'] = texto
    ws_instr.merge_cells(f'A{i}:D{i}')
    cell = ws_instr[f'A{i}']
    
    if any(texto.startswith(x) for x in ["🎯", "📋", "✏️", "⚠️", "🔄", "📤", "💡", "🆘", "📞", "🚀"]):
        cell.font = Font(bold=True, size=12, color="1F4E78")
        cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    elif texto.startswith("   "):
        cell.alignment = Alignment(horizontal="left", indent=2, wrap_text=True)
    else:
        cell.alignment = Alignment(horizontal="left", wrap_text=True)

ws_instr.column_dimensions['A'].width = 100

# ===== SALVAR PLANILHA =====
filename = "Modelo_Sincronizacao_GGOV.xlsx"
wb.save(filename)

print("="*90)
print("✅ PLANILHA EXCEL CRIADA COM SUCESSO!")
print("="*90)
print(f"\n📁 Arquivo: {filename}")
print("\n📊 ESTRUTURA CRIADA:")
print("\n   📂 Aba 'Processo 1':")
print("      ✓ Informações do projeto (linhas 1-5)")
print("      ✓ Instruções de preenchimento (linhas 6-10)")
print("      ✓ Cabeçalhos das ETAPAS (linha 11)")
print("      ✓ 6 ETAPAS modelo (linhas 12-17)")
print("         → Colunas: Etapa, Status, Responsável, Datas, Produtos,")
print("                    Dependências, % Progresso, Horas Est., Horas Real, Peso")
print("      ✓ Cabeçalhos das TAREFAS (linha 21)")
print("      ✓ 9 TAREFAS modelo (linhas 22-30)")
print("         → Colunas: Etapa, Tarefa, Status, Responsável, Prioridade,")
print("                    Prazo, % Conclusão, Horas, Observações")
print("      ✓ Legenda e instruções (linhas 32+)")
print("\n   📖 Aba 'Instruções':")
print("      ✓ Guia completo de preenchimento")
print("      ✓ Como sincronizar com Google Sheets")
print("      ✓ Dicas e solução de problemas")
print("\n🎨 RECURSOS IMPLEMENTADOS:")
print("      ✓ Células editáveis destacadas em AMARELO")
print("      ✓ Dropdowns para Status e Prioridade")
print("      ✓ Formatação automática de datas e percentuais")
print("      ✓ Validação de dados")
print("      ✓ Cabeçalhos com cores e bordas")
print("      ✓ Instruções integradas na planilha")
print("\n📤 PRÓXIMOS PASSOS:")
print("      1. Abra a planilha e preencha os dados reais")
print("      2. Faça upload para Google Drive")
print("      3. Abra com Google Sheets")
print("      4. Compartilhe (Qualquer pessoa com link → Leitor)")
print("      5. Configure o sistema web com API Key + Spreadsheet ID")
print("      6. Sincronização automática ativada!")
print("\n" + "="*90)
print("🎯 Planilha pronta para uso e sincronização!")
print("="*90)
