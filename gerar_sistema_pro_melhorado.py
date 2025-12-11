"""
Sistema de Gerenciamento de Processos LUMA - VERSÃO PRO MELHORADA 🚀
Com Overview Dinâmico + Automações Completas + Sincronização Automática
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, PieChart, DoughnutChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, DataBarRule, IconSetRule
from datetime import datetime, timedelta

wb = Workbook()

border_thin = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

border_thick = Border(
    left=Side(style='medium'), right=Side(style='medium'),
    top=Side(style='medium'), bottom=Side(style='medium')
)

# ==================== ABA 1: PÁGINA INICIAL MELHORADA ====================
ws_inicial = wb.active
ws_inicial.title = "Página inicial"

# Título principal
ws_inicial.merge_cells('A1:P1')
ws_inicial['A1'] = "🚀 GABINETE DE GOVERNANÇA - OVERVIEW EXECUTIVO DE PROCESSOS"
ws_inicial['A1'].font = Font(size=16, bold=True, color="FFFFFF")
ws_inicial['A1'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
ws_inicial['A1'].alignment = Alignment(horizontal="center", vertical="center")
ws_inicial.row_dimensions[1].height = 30

# ===== SEÇÃO DE INDICADORES RÁPIDOS (KPIs no topo) =====
ws_inicial.merge_cells('A2:P2')
ws_inicial['A2'] = "📊 RESUMO EXECUTIVO EM TEMPO REAL"
ws_inicial['A2'].font = Font(size=12, bold=True, color="1F4E78")
ws_inicial['A2'].alignment = Alignment(horizontal="center", vertical="center")
ws_inicial['A2'].fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")

# KPIs em linha
kpi_labels_row3 = ["📋 Total", "🚀 Ativos", "✅ Concluídos", "⏸️ Planejados", 
                   "🔴 Críticos", "💯 Saúde Geral", "💰 Orçamento Total", "📈 Eficiência Média"]

for col, label in enumerate(kpi_labels_row3, 1):
    cell = ws_inicial.cell(row=3, column=col)
    cell.value = label
    cell.font = Font(bold=True, size=9, color="FFFFFF")
    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border_thin

ws_inicial.row_dimensions[3].height = 25

# Fórmulas dos KPIs
kpi_formulas = [
    "=CONT.VALORES(A6:A100)-CONT.SE(A6:A100,\"\")",  # Total
    "=CONT.SE(C6:C100,\"Em execução\")",  # Ativos
    "=CONT.SE(C6:C100,\"Concluída\")",  # Concluídos
    "=CONT.SE(C6:C100,\"Planejada\")",  # Planejados
    "=CONT.SE(L6:L100,\"🔴 Alto\")",  # Críticos
    "=MÉDIA(G6:G100)",  # Saúde Geral
    "=SOMA(N6:N100)",  # Orçamento Total
    "=MÉDIA(O6:O100)"  # Eficiência Média
]

for col, formula in enumerate(kpi_formulas, 1):
    cell = ws_inicial.cell(row=4, column=col)
    cell.value = formula
    cell.font = Font(bold=True, size=13, color="1F4E78")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    cell.border = border_thin
    
    if col == 6:
        cell.number_format = '0%'
    elif col == 7:
        cell.number_format = 'R$ #,##0'
    elif col == 8:
        cell.number_format = '0%'

ws_inicial.row_dimensions[4].height = 28

# Cabeçalhos da tabela principal
headers_inicial = ["ID", "Processo", "Status", "Responsável Principal", "Data Início", 
                   "Data Término", "% Concluído", "Etapas Concluídas", "Tempo Est.", 
                   "Tempo Real", "Dias Rest.", "Risco", "Saúde", "Custo Est.", 
                   "Eficiência", "Link"]

header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True, size=10)

for col_num, header in enumerate(headers_inicial, 1):
    cell = ws_inicial.cell(row=5, column=col_num)
    cell.value = header
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border_thin

ws_inicial.row_dimensions[5].height = 35

# Dados dos processos com automações
processos_exemplo = [
    [1, "Mapeamento Processos GGOV", "='Processo 1'!I2", "='Processo 1'!C15", 
     datetime(2025, 12, 10), datetime(2026, 1, 31), "='Processo 1'!K2", 
     "='Processo 1'!L2", "=F6-E6", "=HOJE()-E6", "=F6-HOJE()",
     "=SE(K6<0,\"🔴 Alto\",SE(K6<3,\"🟡 Médio\",\"🟢 Baixo\"))",
     "=SE(G6>=0.8,\"🟢 Excelente\",SE(G6>=0.6,\"🟡 Bom\",SE(G6>=0.3,\"🟠 Regular\",\"🔴 Crítico\")))", 
     15000, "='Processo 1'!J4", "Processo 1"],
    
    [2, "Atualização Normativas Internas", "='Processo 2'!I2", "='Processo 2'!C15", 
     datetime(2025, 12, 5), datetime(2025, 12, 20), "='Processo 2'!K2",
     "='Processo 2'!L2", "=F7-E7", "=HOJE()-E7", "=F7-HOJE()",
     "=SE(K7<0,\"🔴 Alto\",SE(K7<3,\"🟡 Médio\",\"🟢 Baixo\"))",
     "=SE(G7>=0.8,\"🟢 Excelente\",SE(G7>=0.6,\"🟡 Bom\",SE(G7>=0.3,\"🟠 Regular\",\"🔴 Crítico\")))", 
     8000, "='Processo 2'!J4", "Processo 2"],
    
    [3, "Implantação Sistema GED", "='Processo 3'!I2", "='Processo 3'!C15", 
     datetime(2025, 12, 15), datetime(2026, 2, 15), "='Processo 3'!K2",
     "='Processo 3'!L2", "=F8-E8", "=HOJE()-E8", "=F8-HOJE()",
     "=SE(K8<0,\"🔴 Alto\",SE(K8<3,\"🟡 Médio\",\"🟢 Baixo\"))",
     "=SE(G8>=0.8,\"🟢 Excelente\",SE(G8>=0.6,\"🟡 Bom\",SE(G8>=0.3,\"🟠 Regular\",\"🔴 Crítico\")))", 
     25000, "='Processo 3'!J4", "Processo 3"],
]

for row_num, row_data in enumerate(processos_exemplo, 6):
    for col_num, value in enumerate(row_data, 1):
        cell = ws_inicial.cell(row=row_num, column=col_num)
        cell.value = value
        cell.border = border_thin
        cell.alignment = Alignment(horizontal="center", vertical="center")
        
        if col_num in [5, 6] and isinstance(value, datetime):
            cell.number_format = 'DD/MM/YYYY'
        elif col_num in [7, 15]:
            cell.number_format = '0%'
        elif col_num == 14:
            cell.number_format = 'R$ #,##0.00'

# Formatação condicional avançada
# Status com cores
ws_inicial.conditional_formatting.add('C6:C100',
    CellIsRule(operator='equal', formula=['"Concluída"'], 
               fill=PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
               font=Font(color="006100", bold=True))
)

ws_inicial.conditional_formatting.add('C6:C100',
    CellIsRule(operator='equal', formula=['"Em execução"'], 
               fill=PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
               font=Font(color="9C6500", bold=True))
)

ws_inicial.conditional_formatting.add('C6:C100',
    CellIsRule(operator='equal', formula=['"Bloqueada"'], 
               fill=PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
               font=Font(color="9C0006", bold=True))
)

ws_inicial.conditional_formatting.add('C6:C100',
    CellIsRule(operator='equal', formula=['"Planejada"'], 
               fill=PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid"),
               font=Font(color="1F4E78", bold=True))
)

# Data Bar para % Concluído
ws_inicial.conditional_formatting.add('G6:G100',
    DataBarRule(start_type='num', start_value=0, end_type='num', end_value=1,
                color="4472C4", showValue=True)
)

# Escala de cores para Dias Restantes
ws_inicial.conditional_formatting.add('K6:K100',
    ColorScaleRule(start_type='num', start_value=-10, start_color='F8696B',
                   mid_type='num', mid_value=0, mid_color='FFEB84',
                   end_type='num', end_value=20, end_color='63BE7B')
)

# Ícones para Etapas Concluídas
ws_inicial.conditional_formatting.add('H6:H100',
    IconSetRule('3Symbols', 'num', [0, 33, 67], showValue=True, percent=True, reverse=False)
)

# Validações
dv_status = DataValidation(type="list", formula1='"Planejada,Em execução,Concluída,Cancelada,Bloqueada"')
ws_inicial.add_data_validation(dv_status)
dv_status.add('C6:C100')

column_widths = [5, 30, 14, 25, 12, 12, 11, 13, 10, 10, 10, 12, 16, 12, 11, 12]
for i, width in enumerate(column_widths, 1):
    ws_inicial.column_dimensions[get_column_letter(i)].width = width

# ==================== PROCESSOS AUTOMATIZADOS (1, 2, 3) ====================
for proc_num in range(1, 4):
    ws_proc = wb.create_sheet(f"Processo {proc_num}")
    
    # Cabeçalho do projeto
    ws_proc.merge_cells('A1:A2')
    ws_proc['A1'] = f"Projeto\n{['Mapeamento dos processos do GGOV', 'Atualização de Normativas Internas', 'Implantação do Sistema GED'][proc_num-1]}"
    ws_proc['A1'].font = Font(bold=True, size=11)
    ws_proc['A1'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws_proc['A1'].border = border_thin
    
    ws_proc.merge_cells('B1:B2')
    ws_proc['B1'] = "Descrição da demanda"
    ws_proc['B1'].font = Font(bold=True)
    ws_proc['B1'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws_proc['B1'].border = border_thin
    
    ws_proc.merge_cells('C1:H2')
    descricoes = [
        ("Realizar o mapeamento completo dos processos administrativos e operacionais do "
         "Gabinete de Governança (GGOV), com a finalidade de otimizar o desempenho."),
        ("Revisar e atualizar todas as normativas internas do GGOV para garantir conformidade "
         "com as novas diretrizes e legislação vigente."),
        ("Implementar sistema de Gestão Eletrônica de Documentos para digitalização e controle "
         "de todos os documentos do gabinete.")
    ]
    ws_proc['C1'] = descricoes[proc_num-1]
    ws_proc['C1'].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws_proc['C1'].border = border_thin
    
    # Status calculado automaticamente
    ws_proc['I1'] = "Status Geral"
    ws_proc['I2'] = '=SE(CONT.SE(B8:B13,"Concluída")=CONT.VALORES(A8:A13),"Concluída",SE(CONT.SE(B8:B13,"Em execução")>0,"Em execução",SE(CONT.SE(B8:B13,"Não iniciada")=CONT.VALORES(A8:A13),"Planejada","Em execução")))'
    ws_proc.merge_cells('I1:I2')
    ws_proc['I1'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws_proc['I1'].border = border_thin
    ws_proc['I1'].font = Font(bold=True, size=10)
    
    ws_proc.merge_cells('J1:J2')
    ws_proc['J1'] = "⬅️ Voltar"
    ws_proc['J1'].fill = PatternFill(start_color="B6D7A8", end_color="B6D7A8", fill_type="solid")
    ws_proc['J1'].alignment = Alignment(horizontal="center", vertical="center")
    ws_proc['J1'].border = border_thin
    ws_proc['J1'].font = Font(bold=True)
    
    # % Conclusão Automático
    ws_proc['K1'] = "% Conclusão"
    ws_proc['K2'] = '=MÉDIA(H8:H13)'
    ws_proc.merge_cells('K1:K2')
    ws_proc['K1'].alignment = Alignment(horizontal="center", vertical="center")
    ws_proc['K1'].border = border_thin
    ws_proc['K1'].font = Font(bold=True, size=12, color="FFFFFF")
    ws_proc['K1'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws_proc['K2'].number_format = '0%'
    ws_proc['K2'].font = Font(bold=True, size=14, color="4472C4")
    
    # Etapas Concluídas/Total
    ws_proc['L1'] = "Etapas"
    ws_proc['L2'] = '=CONT.SE(B8:B13,"Concluída")&"/"&CONT.VALORES(A8:A13)'
    ws_proc.merge_cells('L1:L2')
    ws_proc['L1'].alignment = Alignment(horizontal="center", vertical="center")
    ws_proc['L1'].border = border_thin
    ws_proc['L1'].font = Font(bold=True, size=10)
    ws_proc['L2'].font = Font(bold=True, size=12)
    ws_proc['L2'].alignment = Alignment(horizontal="center", vertical="center")
    
    # Informações do projeto (Linha 3-4)
    ws_proc['A3'] = "SEI:"
    ws_proc['B3'] = "0000000000000"
    ws_proc['A3'].font = Font(bold=True)
    ws_proc['A3'].border = border_thin
    ws_proc['B3'].border = border_thin
    
    ws_proc['C3'] = "Prioridade:"
    prioridades = ["Alta", "Crítica", "Média"]
    ws_proc['D3'] = prioridades[proc_num-1]
    cores_prioridade = ["FFA500", "FF0000", "FFFF00"]
    ws_proc['C3'].font = Font(bold=True)
    ws_proc['C3'].border = border_thin
    ws_proc['D3'].fill = PatternFill(start_color=cores_prioridade[proc_num-1], 
                                     end_color=cores_prioridade[proc_num-1], fill_type="solid")
    ws_proc['D3'].font = Font(bold=True, color="FFFFFF" if proc_num <= 2 else "000000")
    ws_proc['D3'].alignment = Alignment(horizontal="center")
    ws_proc['D3'].border = border_thin
    
    ws_proc['E3'] = "Categoria:"
    categorias = ["Mapeamento", "Normativas", "TI/Sistemas"]
    ws_proc['F3'] = categorias[proc_num-1]
    ws_proc['E3'].font = Font(bold=True)
    ws_proc['E3'].border = border_thin
    ws_proc['F3'].border = border_thin
    
    ws_proc['G3'] = "Risco:"
    ws_proc['H3'] = '=SE(K2<0.3,"🔴 Alto",SE(K2<0.7,"🟡 Médio","🟢 Baixo"))'
    ws_proc['G3'].font = Font(bold=True)
    ws_proc['G3'].border = border_thin
    ws_proc['H3'].border = border_thin
    ws_proc['H3'].alignment = Alignment(horizontal="center")
    
    ws_proc['I3'] = "Orçamento:"
    orcamentos = [15000, 8000, 25000]
    ws_proc['J3'] = orcamentos[proc_num-1]
    ws_proc['I3'].font = Font(bold=True)
    ws_proc['I3'].border = border_thin
    ws_proc['J3'].number_format = 'R$ #,##0.00'
    ws_proc['J3'].border = border_thin
    
    ws_proc['K3'] = "Gasto Real:"
    ws_proc['L3'] = '=SOMA(K8:K13)'
    ws_proc['K3'].font = Font(bold=True)
    ws_proc['K3'].border = border_thin
    ws_proc['L3'].number_format = 'R$ #,##0.00'
    ws_proc['L3'].border = border_thin
    
    # Linha 4 - Datas e responsáveis
    ws_proc['A4'] = "Início:"
    datas_inicio = [datetime(2025, 12, 10), datetime(2025, 12, 5), datetime(2025, 12, 15)]
    ws_proc['B4'] = datas_inicio[proc_num-1]
    ws_proc['A4'].font = Font(bold=True)
    ws_proc['A4'].border = border_thin
    ws_proc['B4'].number_format = 'DD/MM/YYYY'
    ws_proc['B4'].border = border_thin
    
    ws_proc['C4'] = "Término:"
    datas_fim = [datetime(2026, 1, 31), datetime(2025, 12, 20), datetime(2026, 2, 15)]
    ws_proc['D4'] = datas_fim[proc_num-1]
    ws_proc['C4'].font = Font(bold=True)
    ws_proc['C4'].border = border_thin
    ws_proc['D4'].number_format = 'DD/MM/YYYY'
    ws_proc['D4'].border = border_thin
    
    ws_proc['E4'] = "Duração:"
    ws_proc['F4'] = "=D4-B4"
    ws_proc['E4'].font = Font(bold=True)
    ws_proc['E4'].border = border_thin
    ws_proc['F4'].border = border_thin
    ws_proc['F4'].alignment = Alignment(horizontal="center")
    
    ws_proc['G4'] = "Dias Rest.:"
    ws_proc['H4'] = "=D4-HOJE()"
    ws_proc['G4'].font = Font(bold=True)
    ws_proc['G4'].border = border_thin
    ws_proc['H4'].border = border_thin
    ws_proc['H4'].alignment = Alignment(horizontal="center")
    
    ws_proc['I4'] = "Eficiência:"
    ws_proc['J4'] = '=SE(J3=0,0,1-(L3/J3))'
    ws_proc['I4'].font = Font(bold=True)
    ws_proc['I4'].border = border_thin
    ws_proc['J4'].number_format = '0%'
    ws_proc['J4'].border = border_thin
    
    # AUTOMAÇÃO: Responsável Principal (calculado pela etapa com maior % concluído)
    ws_proc['K4'] = "Coord. Atual:"
    ws_proc['L4'] = "=TEXTOJUNTAR(\", \",VERDADEIRO,ÚNICO(FILTRAR(C8:C13,C8:C13<>\"\")))"
    ws_proc['K4'].font = Font(bold=True)
    ws_proc['K4'].border = border_thin
    ws_proc['L4'].border = border_thin
    ws_proc['L4'].font = Font(bold=True, color="1F4E78")
    ws_proc['L4'].alignment = Alignment(horizontal="center")
    
    # AUTOMAÇÃO: Linha 15 - Responsável que mais aparece nas etapas
    ws_proc['A15'] = "Responsável Principal (Auto):"
    ws_proc['C15'] = '=ÍNDICE(C8:C13,MODO(CORRESP(C8:C13,C8:C13,0),CORRESP(C8:C13,C8:C13,0)))'
    ws_proc['A15'].font = Font(bold=True, size=9, color="7030A0")
    ws_proc.merge_cells('A15:B15')
    ws_proc.merge_cells('C15:D15')
    ws_proc['C15'].font = Font(bold=True, size=10, color="7030A0")
    ws_proc['C15'].fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")
    
    # Cabeçalhos da tabela de etapas
    headers_etapas = ["Etapa", "Status", "Responsável", "Dt. Início", "Dt. Término", 
                      "Produtos/Entregas", "Dependência", "% Progresso", "Horas Est.", 
                      "Horas Real", "Custo", "Tarefas Detalhadas"]
    
    for col, header in enumerate(headers_etapas, 1):
        cell = ws_proc.cell(row=7, column=col)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF", size=9)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border_thin
    
    ws_proc.row_dimensions[7].height = 32
    
    # Dados das etapas (personalizados por processo)
    etapas_processo1 = [
        ["Levantamento", "Em execução", "Luma Damon", 
         datetime(2025, 12, 10), datetime(2026, 1, 16), "Plano do projeto", 
         "-", 0.70, 80, 56, 2800, "1. Entrevistas\n2. Coleta docs"],
        ["Mapeamento", "Em execução", "Suerlei Gondim", 
         datetime(2025, 12, 10), datetime(2026, 1, 31), "Relatório", 
         "Etapa 1", 0.60, 120, 72, 3600, "1. Documentar\n2. Gargalos"],
        ["Análise", "Não iniciada", "Pedro Analista", 
         datetime(2026, 1, 17), datetime(2026, 2, 15), "Mapas", 
         "Etapa 1,2", 0.00, 100, 0, 0, "1. Eficiência\n2. Melhorias"],
        ["Documentação", "Não iniciada", "Ana Documentadora", 
         datetime(2026, 2, 1), datetime(2026, 2, 28), "Relatório Final", 
         "Etapa 3", 0.00, 80, 0, 0, "1. Consolidar\n2. Apresentar"],
        ["Validação", "Não iniciada", "Diretor GGOV", 
         datetime(2026, 2, 20), datetime(2026, 3, 10), "Aprovação", 
         "Etapa 4", 0.00, 40, 0, 0, "1. Revisar\n2. Aprovar"],
        ["Implementação", "Não iniciada", "Equipe GGOV", 
         datetime(2026, 3, 1), datetime(2026, 3, 31), "Processos Ativos", 
         "Etapa 5", 0.00, 60, 0, 0, "1. Treinar\n2. Monitorar"],
    ]
    
    etapas_processo2 = [
        ["Revisão Legislação", "Em execução", "Maria Santos", 
         datetime(2025, 12, 5), datetime(2025, 12, 12), "Checklist Legal", 
         "-", 0.80, 40, 32, 1600, "1. Leis\n2. Decretos"],
        ["Análise Impacto", "Em execução", "João Jurídico", 
         datetime(2025, 12, 8), datetime(2025, 12, 15), "Relatório Impacto", 
         "Etapa 1", 0.60, 30, 18, 900, "1. Mudanças\n2. Riscos"],
        ["Redação Normativas", "Não iniciada", "Equipe Jurídica", 
         datetime(2025, 12, 13), datetime(2025, 12, 18), "Minutas", 
         "Etapa 2", 0.00, 50, 0, 0, "1. Redigir\n2. Revisar"],
        ["Validação Interna", "Não iniciada", "Comitê GGOV", 
         datetime(2025, 12, 16), datetime(2025, 12, 19), "Aprovação", 
         "Etapa 3", 0.00, 20, 0, 0, "1. Apresentar\n2. Ajustar"],
        ["Publicação", "Não iniciada", "Comunicação", 
         datetime(2025, 12, 19), datetime(2025, 12, 20), "Normativas Publicadas", 
         "Etapa 4", 0.00, 10, 0, 0, "1. Diagramar\n2. Publicar"],
        ["Treinamento", "Não iniciada", "RH", 
         datetime(2025, 12, 20), datetime(2025, 12, 20), "Equipe Capacitada", 
         "Etapa 5", 0.00, 15, 0, 0, "1. Workshop\n2. Material"],
    ]
    
    etapas_processo3 = [
        ["Levantamento Requisitos", "Planejada", "Carlos Pereira", 
         datetime(2025, 12, 15), datetime(2025, 12, 28), "Especificações", 
         "-", 0.20, 60, 12, 600, "1. Reuniões\n2. Documentar"],
        ["Seleção Fornecedor", "Não iniciada", "Compras", 
         datetime(2025, 12, 20), datetime(2026, 1, 10), "Contrato", 
         "Etapa 1", 0.00, 40, 0, 0, "1. Pesquisa\n2. Cotação"],
        ["Customização Sistema", "Não iniciada", "Equipe TI", 
         datetime(2026, 1, 11), datetime(2026, 1, 31), "Sistema Config", 
         "Etapa 2", 0.00, 120, 0, 0, "1. Parametrizar\n2. Integrar"],
        ["Migração Dados", "Não iniciada", "DBA", 
         datetime(2026, 1, 25), datetime(2026, 2, 5), "Base Migrada", 
         "Etapa 3", 0.00, 80, 0, 0, "1. Extrair\n2. Importar"],
        ["Testes", "Não iniciada", "QA Team", 
         datetime(2026, 2, 1), datetime(2026, 2, 10), "Homologação", 
         "Etapa 4", 0.00, 50, 0, 0, "1. Testar\n2. Ajustar"],
        ["Go-Live", "Não iniciada", "TI + GGOV", 
         datetime(2026, 2, 11), datetime(2026, 2, 15), "Sistema Prod", 
         "Etapa 5", 0.00, 30, 0, 0, "1. Deploy\n2. Suporte"],
    ]
    
    etapas_data = [etapas_processo1, etapas_processo2, etapas_processo3][proc_num-1]
    
    for row_idx, etapa in enumerate(etapas_data, 8):
        for col_idx, value in enumerate(etapa, 1):
            cell = ws_proc.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.border = border_thin
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            
            if col_idx in [4, 5] and isinstance(value, datetime):
                cell.number_format = 'DD/MM/YYYY'
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col_idx == 8:
                cell.number_format = '0%'
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col_idx in [9, 10]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col_idx == 11:
                cell.number_format = 'R$ #,##0.00'
                cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Validações
    dv_status_etapa = DataValidation(type="list", formula1='"Não iniciada,Em execução,Concluída,Bloqueada"')
    ws_proc.add_data_validation(dv_status_etapa)
    dv_status_etapa.add('B8:B20')
    
    # Data Bar para progresso
    ws_proc.conditional_formatting.add('H8:H20',
        DataBarRule(start_type='num', start_value=0, end_type='num', end_value=1,
                    color="4472C4", showValue=True)
    )
    
    # Formatação condicional no % total
    ws_proc.conditional_formatting.add('K2',
        ColorScaleRule(start_type='num', start_value=0, start_color='F8696B',
                       mid_type='num', mid_value=0.5, mid_color='FFEB84',
                       end_type='num', end_value=1, end_color='63BE7B')
    )
    
    column_widths_proc = [18, 14, 20, 11, 11, 22, 11, 11, 9, 9, 11, 30]
    for i, width in enumerate(column_widths_proc, 1):
        ws_proc.column_dimensions[get_column_letter(i)].width = width
    
    for row in range(1, 14):
        if row in [1, 2]:
            ws_proc.row_dimensions[row].height = 35
        elif row >= 8:
            ws_proc.row_dimensions[row].height = 38

# ==================== ABA: DASHBOARD EXECUTIVO ====================
ws_dashboard = wb.create_sheet("Dashboard")

ws_dashboard.merge_cells('A1:H2')
ws_dashboard['A1'] = "📊 DASHBOARD EXECUTIVO - VISÃO ESTRATÉGICA 360°"
ws_dashboard['A1'].font = Font(size=18, bold=True, color="FFFFFF")
ws_dashboard['A1'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
ws_dashboard['A1'].alignment = Alignment(horizontal="center", vertical="center")

ws_dashboard['A3'] = f"📅 Atualizado automaticamente em: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
ws_dashboard['A3'].font = Font(italic=True, size=10)
ws_dashboard.merge_cells('A3:H3')

# KPIs
ws_dashboard.merge_cells('A5:H5')
ws_dashboard['A5'] = "🎯 INDICADORES ESTRATÉGICOS EM TEMPO REAL"
ws_dashboard['A5'].font = Font(size=14, bold=True, color="1F4E78")
ws_dashboard['A5'].alignment = Alignment(horizontal="center")

kpi_data = [
    ["📋 Total", "=CONT.VALORES('Página inicial'!A6:A100)", "4472D4"],
    ["🚀 Ativos", "=CONT.SE('Página inicial'!C:C,\"Em execução\")", "FFA500"],
    ["✅ Concluídos", "=CONT.SE('Página inicial'!C:C,\"Concluída\")", "00B050"],
    ["⚠️ Críticos", "=CONT.SE('Página inicial'!L:L,\"*Alto*\")", "C00000"],
    ["📈 Taxa Sucesso", "=TEXTO(D6/B6,\"0%\")", "00B0F0"],
    ["💯 Saúde", "=MÉDIA('Página inicial'!G6:G20)", "92D050"],
    ["💰 Orçamento", "=SOMA('Página inicial'!N6:N20)", "7030A0"],
    ["⚡ Eficiência", "=MÉDIA('Página inicial'!O6:O20)", "00B050"],
]

col = 1
for label, formula, color in kpi_data:
    label_cell = ws_dashboard.cell(row=6, column=col)
    label_cell.value = label
    label_cell.font = Font(bold=True, size=9, color="FFFFFF")
    label_cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    label_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    label_cell.border = border_thin
    
    value_cell = ws_dashboard.cell(row=7, column=col)
    value_cell.value = formula
    value_cell.font = Font(bold=True, size=14, color=color)
    value_cell.alignment = Alignment(horizontal="center", vertical="center")
    value_cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    value_cell.border = border_thin
    
    if col in [5, 6, 8]:
        value_cell.number_format = '0%'
    elif col == 7:
        value_cell.number_format = 'R$ #,##0'
    
    col += 1

# Gráficos
ws_dashboard['A10'] = "Status"
ws_dashboard['B10'] = "Quantidade"
status_list = ["Em execução", "Concluída", "Planejada", "Bloqueada"]
for i, st in enumerate(status_list, 11):
    ws_dashboard[f'A{i}'] = st
    ws_dashboard[f'B{i}'] = f'=CONT.SE(\'Página inicial\'!C:C,"{st}")'

donut = DoughnutChart()
labels_d = Reference(ws_dashboard, min_col=1, min_row=11, max_row=14)
data_d = Reference(ws_dashboard, min_col=2, min_row=10, max_row=14)
donut.add_data(data_d, titles_from_data=True)
donut.set_categories(labels_d)
donut.title = "📊 Distribuição por Status"
donut.style = 10
donut.height = 12
donut.width = 16
donut.dataLabels = DataLabelList()
donut.dataLabels.showPercent = True
ws_dashboard.add_chart(donut, "D10")

for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
    ws_dashboard.column_dimensions[col].width = 16

# ==================== ABA: INSTRUÇÕES ====================
ws_instr = wb.create_sheet("Instruções")

ws_instr.merge_cells('A1:D1')
ws_instr['A1'] = "📖 GUIA DO SISTEMA MELHORADO"
ws_instr['A1'].font = Font(size=16, bold=True, color="FFFFFF")
ws_instr['A1'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
ws_instr['A1'].alignment = Alignment(horizontal="center", vertical="center")

instrucoes = [
    "", "🚀 SISTEMA PRO MELHORADO - AUTOMAÇÕES COMPLETAS", "",
    "✨ NOVIDADES DESTA VERSÃO:", "",
    "📊 PÁGINA INICIAL APRIMORADA:",
    "   • KPIs em tempo real no topo da página",
    "   • Coluna 'Etapas Concluídas' mostra progresso visual",
    "   • Responsável Principal sincronizado automaticamente",
    "   • % Conclusão calculado pela média das etapas",
    "   • Indicadores de Saúde em 4 níveis (Excelente/Bom/Regular/Crítico)",
    "   • Data Bars visuais para % de progresso",
    "   • Ícones de status para etapas concluídas", "",
    "🤖 AUTOMAÇÕES NOS PROCESSOS:",
    "   • Status Geral atualizado conforme etapas",
    "   • % Conclusão calculado automaticamente pela média",
    "   • Responsável Principal identificado automaticamente",
    "   • Coordenador Atual lista todos os envolvidos",
    "   • Etapas Concluídas contadas automaticamente (ex: 2/6)",
    "   • Risco calculado pelo % de conclusão",
    "   • Eficiência orçamentária automática", "",
    "🔗 SINCRONIZAÇÃO AUTOMÁTICA:",
    "   1. Atualize o status das etapas nos processos",
    "   2. O Status Geral do processo atualiza sozinho",
    "   3. O % Conclusão é calculado automaticamente",
    "   4. O Responsável Principal é identificado",
    "   5. A Página inicial reflete TUDO automaticamente",
    "   6. Dashboard atualiza em tempo real", "",
    "📈 COMO USAR:",
    "   1. Vá em Processo 1, 2 ou 3",
    "   2. Atualize Status e % Progresso das etapas",
    "   3. Adicione responsáveis nas etapas",
    "   4. Volte à Página inicial e veja a mágica!",
    "   5. Confira o Dashboard para visão executiva", "",
    "💡 Tudo está conectado e sincronizado automaticamente!",
]

for i, texto in enumerate(instrucoes, 2):
    ws_instr[f'A{i}'] = texto
    ws_instr.merge_cells(f'A{i}:D{i}')
    cell = ws_instr[f'A{i}']
    
    if any(texto.startswith(x) for x in ["🚀", "✨", "📊", "🤖", "🔗", "📈", "💡"]):
        cell.font = Font(bold=True, size=12, color="1F4E78")
    elif texto.startswith("   •") or texto.startswith("   "):
        cell.alignment = Alignment(horizontal="left", indent=2)
    elif len(texto) > 0 and texto[0].isdigit():
        cell.alignment = Alignment(horizontal="left", indent=1)
    else:
        cell.alignment = Alignment(horizontal="left")

ws_instr.column_dimensions['A'].width = 80

# Salvar
filename = "Sistema_GGOV_Pro_Melhorado.xlsx"
wb.save(filename)

print("="*80)
print("🚀 SISTEMA PRO MELHORADO CRIADO COM SUCESSO!")
print("="*80)
print(f"\n📁 Arquivo: {filename}")
print("\n✨ MELHORIAS IMPLEMENTADAS:")
print("\n   📊 PÁGINA INICIAL DINÂMICA:")
print("      ✓ KPIs em tempo real no topo")
print("      ✓ Coluna 'Etapas Concluídas' (ex: 3/6)")
print("      ✓ Responsável sincronizado automaticamente")
print("      ✓ Indicador de Saúde em 4 níveis")
print("      ✓ Data Bars visuais para progresso")
print("      ✓ Ícones para etapas concluídas")
print("\n   🤖 AUTOMAÇÕES NOS PROCESSOS:")
print("      ✓ Status calculado automaticamente")
print("      ✓ % Conclusão = Média das etapas")
print("      ✓ Responsável Principal auto-identificado")
print("      ✓ Contador de etapas automático")
print("      ✓ Risco baseado em progresso")
print("      ✓ Eficiência orçamentária")
print("\n   🔗 SINCRONIZAÇÃO COMPLETA:")
print("      ✓ Processos ↔ Página inicial")
print("      ✓ Etapas ↔ Status geral")
print("      ✓ Responsáveis ↔ Coordenação")
print("      ✓ Tudo atualiza em tempo real!")
print("\n" + "="*80)
print("🎯 Sistema 100% automatizado e sincronizado!")
print("="*80)
