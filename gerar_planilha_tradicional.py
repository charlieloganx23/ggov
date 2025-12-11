"""
Planilha Excel Tradicional para Sistema GGOV
Estrutura linear e simplificada para fácil preenchimento
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from datetime import datetime

wb = Workbook()
ws = wb.active
ws.title = "Processo 1"

border_thin = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

# ==================== TÍTULO ====================
ws['A1'] = "📂 PROCESSO 1: Mapeamento dos processos do Gabinete de Governança"
ws['A1'].font = Font(size=14, bold=True, color="FFFFFF")
ws['A1'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 30

# ==================== INFORMAÇÕES DO PROJETO ====================
# Cabeçalhos (linha 2)
info_headers = ["SEI", "Prioridade", "Categoria", "Data Início", "Data Término", "Orçamento", "Descrição"]
for col, header in enumerate(info_headers, 1):
    cell = ws.cell(row=2, column=col)
    cell.value = header
    cell.font = Font(bold=True, size=10, color="FFFFFF")
    cell.fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = border_thin

# Dados (linha 3)
info_data = [
    "0000000000000",
    "Alta", 
    "Mapeamento",
    datetime(2025, 12, 10),
    datetime(2026, 1, 31),
    "R$ 15.000",
    "Realizar o mapeamento completo dos processos administrativos e operacionais do GGOV"
]

for col, value in enumerate(info_data, 1):
    cell = ws.cell(row=3, column=col)
    cell.value = value
    cell.border = border_thin
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    
    if col in [4, 5]:  # Datas
        cell.number_format = 'DD/MM/YYYY'

ws.row_dimensions[2].height = 25
ws.row_dimensions[3].height = 50

# ==================== ETAPAS ====================
ws['A5'] = "📊 ETAPAS DO PROCESSO"
ws['A5'].font = Font(size=12, bold=True, color="FFFFFF")
ws['A5'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
ws['A5'].alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[5].height = 25

# Cabeçalhos das etapas (linha 6)
etapas_headers = [
    "Nome da Etapa", "Status", "Responsável", "Data Início", "Data Término",
    "Produtos/Entregas", "Dependências", "% Progresso", "Horas Est.", "Horas Real", "Peso"
]

for col, header in enumerate(etapas_headers, 1):
    cell = ws.cell(row=6, column=col)
    cell.value = header
    cell.font = Font(bold=True, size=10, color="FFFFFF")
    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border_thin

ws.row_dimensions[6].height = 30

# Dados das etapas (linhas 7+)
etapas_data = [
    ["Levantamento de Informações", "Em execução", "Luma Damon de Oliveira Melo",
     datetime(2025, 12, 10), datetime(2026, 1, 16), "Plano do projeto",
     "-", 0.70, 80, 56, 0.15],
    
    ["Mapeamento de Processos", "Em execução", "Suerlei Gondim Dutra",
     datetime(2025, 12, 10), datetime(2026, 1, 31), "Relatório de Levantamento / Mapas de Processos",
     "Etapa 1", 0.60, 120, 72, 0.25],
    
    ["Análise de Processos", "Não iniciada", "Equipe GGOV",
     datetime(2026, 1, 17), datetime(2026, 2, 15), "Análise de eficiência e gargalos",
     "Etapa 1, 2", 0.00, 100, 0, 0.20],
    
    ["Documentação e Relatório Final", "Não iniciada", "Equipe Técnica",
     datetime(2026, 2, 1), datetime(2026, 2, 28), "Relatório Final Consolidado",
     "Etapa 3", 0.00, 80, 0, 0.20],
    
    ["Validação e Aprovação", "Não iniciada", "Direção GGOV",
     datetime(2026, 2, 20), datetime(2026, 3, 10), "Aprovação formal",
     "Etapa 4", 0.00, 40, 0, 0.10],
    
    ["Entrega e Implementação", "Não iniciada", "Equipe GGOV Completa",
     datetime(2026, 3, 1), datetime(2026, 3, 31), "Processos implementados",
     "Etapa 5", 0.00, 60, 0, 0.10],
]

for row_idx, etapa in enumerate(etapas_data, 7):
    for col_idx, value in enumerate(etapa, 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.value = value
        cell.border = border_thin
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        
        if col_idx in [4, 5]:  # Datas
            cell.number_format = 'DD/MM/YYYY'
        elif col_idx in [8, 11]:  # Percentuais
            cell.number_format = '0.00'
        elif col_idx in [9, 10]:  # Horas
            cell.number_format = '0'
        
        # Células editáveis em amarelo
        if col_idx in [2, 3, 8, 10]:
            cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    
    ws.row_dimensions[row_idx].height = 35

# Validação de status
dv_status = DataValidation(type="list",
                           formula1='"Não iniciada,Em execução,Concluída,Bloqueada,Cancelada"',
                           allow_blank=False)
ws.add_data_validation(dv_status)
dv_status.add('B7:B50')

# ==================== TAREFAS ====================
tarefas_start_row = len(etapas_data) + 9  # Linha após as etapas + espaço

ws[f'A{tarefas_start_row}'] = "📝 TAREFAS DETALHADAS"
ws[f'A{tarefas_start_row}'].font = Font(size=12, bold=True, color="FFFFFF")
ws[f'A{tarefas_start_row}'].fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
ws[f'A{tarefas_start_row}'].alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[tarefas_start_row].height = 25

# Cabeçalhos das tarefas
tarefas_headers = [
    "Etapa", "Nome da Tarefa", "Status", "Responsável", "Prioridade",
    "Prazo", "% Conclusão", "Horas", "Observações"
]

header_row = tarefas_start_row + 1
for col, header in enumerate(tarefas_headers, 1):
    cell = ws.cell(row=header_row, column=col)
    cell.value = header
    cell.font = Font(bold=True, size=10, color="FFFFFF")
    cell.fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border_thin

ws.row_dimensions[header_row].height = 30

# Dados das tarefas
tarefas_data = [
    ["Etapa 1", "Realizar entrevistas com os responsáveis de cada área",
     "Em execução", "Luma Damon", "Alta", datetime(2025, 12, 15), 0.80, 20, "Entrevistas em andamento"],
    
    ["Etapa 1", "Analisar documentos existentes, como manuais e fluxos anteriores",
     "Em execução", "Luma Damon", "Alta", datetime(2025, 12, 20), 0.70, 16, "70% dos docs revisados"],
    
    ["Etapa 1", "Observar e registrar as atividades nas áreas de governança",
     "Em execução", "Luma Damon", "Média", datetime(2026, 1, 5), 0.60, 24, "Observação em campo"],
    
    ["Etapa 1", "Criar questionário para coletar dados com responsáveis",
     "Concluída", "Luma Damon", "Alta", datetime(2025, 12, 12), 1.00, 8, "Questionário aplicado"],
    
    ["Etapa 1", "Identificar entradas, saídas e responsáveis de cada processo",
     "Em execução", "Luma Damon", "Alta", datetime(2026, 1, 10), 0.50, 12, "50% identificado"],
    
    ["Etapa 2", "Documentar processos no formato AS-IS",
     "Em execução", "Suerlei Gondim", "Alta", datetime(2026, 1, 15), 0.70, 40, "Documentação em progresso"],
    
    ["Etapa 2", "Criar diagramas de fluxo (BPMN)",
     "Em execução", "Suerlei Gondim", "Alta", datetime(2026, 1, 20), 0.60, 30, "Diagramas iniciados"],
    
    ["Etapa 2", "Identificar gargalos e ineficiências",
     "Não iniciada", "Suerlei Gondim", "Média", datetime(2026, 1, 25), 0.00, 25, "Aguardando mapeamento"],
    
    ["Etapa 2", "Consolidar relatório de levantamento",
     "Não iniciada", "Suerlei Gondim", "Média", datetime(2026, 1, 31), 0.00, 25, "Etapa final"],
]

data_start_row = header_row + 1
for row_idx, tarefa in enumerate(tarefas_data, data_start_row):
    for col_idx, value in enumerate(tarefa, 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.value = value
        cell.border = border_thin
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        
        if col_idx == 6:  # Prazo
            cell.number_format = 'DD/MM/YYYY'
        elif col_idx == 7:  # % Conclusão
            cell.number_format = '0.00'
        elif col_idx == 8:  # Horas
            cell.number_format = '0'
        
        # Células editáveis
        if col_idx in [3, 5, 7, 9]:
            cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    
    ws.row_dimensions[row_idx].height = 30

# Validações para tarefas
dv_status_tarefa = DataValidation(type="list",
                                  formula1='"Não iniciada,Em execução,Concluída,Bloqueada,Cancelada"',
                                  allow_blank=False)
ws.add_data_validation(dv_status_tarefa)
dv_status_tarefa.add(f'C{data_start_row}:C100')

dv_prioridade = DataValidation(type="list",
                               formula1='"Alta,Média,Baixa"',
                               allow_blank=False)
ws.add_data_validation(dv_prioridade)
dv_prioridade.add(f'E{data_start_row}:E100')

# ==================== LARGURAS DAS COLUNAS ====================
column_widths = {
    'A': 35, 'B': 15, 'C': 25, 'D': 12, 'E': 12,
    'F': 30, 'G': 15, 'H': 12, 'I': 12, 'J': 12, 'K': 10
}

for col_letter, width in column_widths.items():
    ws.column_dimensions[col_letter].width = width

# ==================== SALVAR ====================
filename = "Planilha_Tradicional_GGOV.xlsx"
wb.save(filename)

print("="*80)
print("✅ PLANILHA TRADICIONAL CRIADA COM SUCESSO!")
print("="*80)
print(f"\n📁 Arquivo: {filename}")
print("\n📊 ESTRUTURA TRADICIONAL:")
print("\n   Linha 1: Título do processo")
print("   Linha 2-3: INFORMAÇÕES DO PROJETO")
print("      → SEI | Prioridade | Categoria | Datas | Orçamento | Descrição")
print("\n   Linha 6+: ETAPAS")
print("      → Nome | Status | Responsável | Datas | Produtos | % | Horas | Peso")
print(f"\n   Linha {header_row}+: TAREFAS")
print("      → Etapa | Tarefa | Status | Responsável | Prioridade | Prazo | % | Horas")
print("\n✨ VANTAGENS:")
print("   ✅ Layout linear e intuitivo")
print("   ✅ Fácil de copiar/colar linhas")
print("   ✅ Sem células mescladas complexas")
print("   ✅ Campos editáveis em AMARELO")
print("   ✅ Dropdowns para Status e Prioridade")
print("\n" + "="*80)
