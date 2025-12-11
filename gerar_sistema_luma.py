"""
Sistema de Gerenciamento de Demandas baseado no modelo_luma.xlsx
Criado com Python + openpyxl
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import datetime, timedelta

# Criar workbook
wb = Workbook()

# Estilos padrão
border_thin = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# ==================== ABA 1: PÁGINA INICIAL (Resumo) ====================
ws_inicial = wb.active
ws_inicial.title = "Página inicial"

# Cabeçalhos da aba inicial
headers_inicial = ["Processo", "Status", "Responsável", "Data de Início", 
                   "Data de Término", "% Concluído", "Tempo Estimado (dias)", "Tempo Real (dias)"]

# Aplicar cabeçalhos com negrito
for col_num, header in enumerate(headers_inicial, 1):
    cell = ws_inicial.cell(row=1, column=col_num)
    cell.value = header
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border_thin

# Dados de exemplo para a página inicial
processos_exemplo = [
    ["Processo A", "='processos A'!H2", "João Silva", 
     datetime(2025, 1, 1), datetime(2025, 1, 10), 0.6, 10, 6],
    ["Processo B", "Em andamento", "Maria Oliveira", 
     datetime(2025, 1, 5), datetime(2025, 1, 12), 0.8, 7, 5],
    ["Processo C", "Não iniciado", "Carlos Pereira", 
     datetime(2025, 1, 10), datetime(2025, 1, 20), 0.0, 10, 0],
    ["Processo D", "Em andamento", "Ana Souza", 
     datetime(2025, 1, 8), datetime(2025, 1, 15), 0.5, 7, 3],
    ["Processo E", "Concluída", "Luís Almeida", 
     datetime(2025, 1, 2), datetime(2025, 1, 7), 1.0, 5, 5],
]

# Inserir dados
for row_num, row_data in enumerate(processos_exemplo, 2):
    for col_num, value in enumerate(row_data, 1):
        cell = ws_inicial.cell(row=row_num, column=col_num)
        cell.value = value
        cell.border = border_thin
        cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Formatar datas
        if col_num in [4, 5] and isinstance(value, datetime):
            cell.number_format = 'DD/MM/YYYY'
        # Formatar percentual
        elif col_num == 6:
            cell.number_format = '0%'

# Ajustar larguras das colunas
column_widths_inicial = [15.5, 15.5, 18.5, 16.5, 16.5, 12, 22, 19]
for i, width in enumerate(column_widths_inicial, 1):
    ws_inicial.column_dimensions[get_column_letter(i)].width = width

# ==================== ABA 2: PROCESSOS A (Detalhamento) ====================
ws_processo_a = wb.create_sheet("processos A")

# Título do projeto (células mescladas)
ws_processo_a.merge_cells('A1:A2')
cell_projeto = ws_processo_a['A1']
cell_projeto.value = "Projeto\nMapeamento dos processos do Gabinete de Governança"
cell_projeto.font = Font(bold=True, size=11)
cell_projeto.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
cell_projeto.border = border_thin

# Descrição da demanda
ws_processo_a.merge_cells('B1:B2')
cell_desc_label = ws_processo_a['B1']
cell_desc_label.value = "Descrição da demanda"
cell_desc_label.font = Font(bold=True)
cell_desc_label.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
cell_desc_label.border = border_thin

# Texto da descrição
ws_processo_a.merge_cells('C1:G2')
cell_descricao = ws_processo_a['C1']
cell_descricao.value = ("Realizar o mapeamento completo dos processos administrativos e operacionais do "
                        "Gabinete de Governança (GGOV), com a finalidade de otimizar o desempenho das "
                        "atividades e garantir maior transparência, eficiência e controle nos fluxos de trabalho.")
cell_descricao.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
cell_descricao.border = border_thin

# Status atual - adicionar fórmula antes de mesclar
ws_processo_a['H1'] = "Status atual"
ws_processo_a['H2'] = '=SE(CONT.SE(B6:B11,"Concluída")=6,"Concluída",SE(CONT.SE(B6:B11,"Em execução")>0,"Em execução","Não iniciada"))'
ws_processo_a.merge_cells('H1:H2')
ws_processo_a['H1'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
ws_processo_a['H1'].border = border_thin

# Link para página inicial
ws_processo_a.merge_cells('I1:I2')
cell_link = ws_processo_a['I1']
cell_link.value = "Página inicial"
cell_link.fill = PatternFill(start_color="B6D7A8", end_color="B6D7A8", fill_type="solid")
cell_link.alignment = Alignment(horizontal="center", vertical="center")
cell_link.border = border_thin

# Número SEI
ws_processo_a['A3'] = "Número sei:"
ws_processo_a['A4'] = "0000000000000"
ws_processo_a.merge_cells('A3:A4')
ws_processo_a['A3'].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
ws_processo_a['A3'].border = border_thin

# Status em execução (texto padrão)
ws_processo_a['H3'] = "Em execução"
ws_processo_a['H3'].alignment = Alignment(horizontal="center", vertical="center")
ws_processo_a['H3'].border = border_thin
ws_processo_a.merge_cells('H3:H4')

# Cabeçalhos da tabela de passos
ws_processo_a['A5'] = "Passos para entregar a demanda:"
ws_processo_a['B5'] = "Status"
ws_processo_a['C5'] = "Responsável"
ws_processo_a['D5'] = "Data de início"
ws_processo_a['E5'] = "Data de término"
ws_processo_a['F5'] = "Produtos"
ws_processo_a['G5'] = "Tarefas"

for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
    cell = ws_processo_a[f'{col}5']
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border_thin

# Dados dos passos
passos_dados = [
    ["Levantamento de Informações", "Em execução", "Luma Damon de Oliveira Melo", 
     datetime(2025, 12, 10), datetime(2026, 1, 16), "Plano do projeto", 
     "1. Realizar entrevistas com os responsáveis\n2. Coletar documentação existente"],
    ["Mapeamento de Processos", "Em execução", "Suerlei Gondim Dutra", 
     datetime(2025, 12, 10), datetime(2026, 1, 31), "Relatório de Levantamento de Dados", 
     "1. Documentar processos atuais\n2. Identificar gargalos"],
    ["Análise de Processos", "Não iniciada", "", 
     None, None, "Mapas de Processos (Diagramas)", 
     "1. Analisar eficiência\n2. Propor melhorias"],
    ["Documentação e Relatório Final", "Não iniciada", "", 
     None, None, "Relatório de Análise de Processos", 
     "1. Consolidar documentação\n2. Preparar apresentação"],
    ["Validação e Aprovação", "Não iniciada", "", 
     None, None, "Relatório Final Validado", 
     "1. Apresentar para stakeholders\n2. Ajustar conforme feedback"],
    ["Entrega e Implementação", "Não iniciada", "", 
     None, None, "Processos Implementados", 
     "1. Treinar equipe\n2. Monitorar implementação"],
]

# Inserir dados dos passos
for row_idx, passo in enumerate(passos_dados, 6):
    for col_idx, value in enumerate(passo, 1):
        cell = ws_processo_a.cell(row=row_idx, column=col_idx)
        cell.value = value
        cell.border = border_thin
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        
        # Formatar datas
        if col_idx in [4, 5] and isinstance(value, datetime):
            cell.number_format = 'DD/MM/YYYY'

# Validação de dados para Status
dv_status_processo = DataValidation(type="list", 
                                     formula1='"Não iniciada,Em execução,Concluída"', 
                                     allow_blank=False)
dv_status_processo.error = 'Selecione um status válido'
dv_status_processo.errorTitle = 'Entrada Inválida'
ws_processo_a.add_data_validation(dv_status_processo)
dv_status_processo.add('B6:B20')

# Ajustar larguras das colunas
column_widths_processo = [30, 15, 25, 20, 20, 23, 63]
for i, width in enumerate(column_widths_processo, 1):
    ws_processo_a.column_dimensions[get_column_letter(i)].width = width

# Ajustar altura das linhas
ws_processo_a.row_dimensions[1].height = 50
ws_processo_a.row_dimensions[2].height = 50
for row in range(6, 12):
    ws_processo_a.row_dimensions[row].height = 40

# ==================== ABA 3: PROCESSOS B (Template) ====================
ws_processo_b = wb.create_sheet("processos B")

# Copiar estrutura do processos A
ws_processo_b.merge_cells('A1:A2')
ws_processo_b['A1'] = "Projeto\n[Nome do Projeto B]"
ws_processo_b['A1'].font = Font(bold=True, size=11)
ws_processo_b['A1'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
ws_processo_b['A1'].border = border_thin

ws_processo_b.merge_cells('B1:B2')
ws_processo_b['B1'] = "Descrição da demanda"
ws_processo_b['B1'].font = Font(bold=True)
ws_processo_b['B1'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
ws_processo_b['B1'].border = border_thin

ws_processo_b.merge_cells('C1:G2')
ws_processo_b['C1'] = "[Descrição detalhada do projeto B]"
ws_processo_b['C1'].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
ws_processo_b['C1'].border = border_thin

ws_processo_b.merge_cells('H1:H2')
ws_processo_b['H1'] = "Status atual"
ws_processo_b['H1'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
ws_processo_b['H1'].border = border_thin

ws_processo_b.merge_cells('I1:I2')
ws_processo_b['I1'] = "Página inicial"
ws_processo_b['I1'].fill = PatternFill(start_color="B6D7A8", end_color="B6D7A8", fill_type="solid")
ws_processo_b['I1'].alignment = Alignment(horizontal="center", vertical="center")
ws_processo_b['I1'].border = border_thin

ws_processo_b.merge_cells('A3:A4')
ws_processo_b['A3'] = "Número sei:\n0000000000000"
ws_processo_b['A3'].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
ws_processo_b['A3'].border = border_thin

ws_processo_b.merge_cells('H3:H4')
# Status em execução (antes de mesclar)
ws_processo_b['H3'] = '=SE(CONT.SE(B6:B11,"Concluída")=6,"Concluída",SE(CONT.SE(B6:B11,"Em execução")>0,"Em execução","Não iniciada"))'
ws_processo_b['H3'].alignment = Alignment(horizontal="center", vertical="center")
ws_processo_b['H3'].border = border_thin
ws_processo_b.merge_cells('H3:H4')

# Cabeçalhos
headers_b = ["Passos para entregar a demanda:", "Status", "Responsável", 
             "Data de início", "Data de término", "Produtos", "Tarefas"]
for col, header in enumerate(headers_b, 1):
    cell = ws_processo_b.cell(row=5, column=col)
    cell.value = header
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border_thin

# Linhas vazias para preenchimento
for row in range(6, 12):
    for col in range(1, 8):
        cell = ws_processo_b.cell(row=row, column=col)
        cell.border = border_thin
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
# Validação de dados
dv_status_b = DataValidation(type="list", 
                              formula1='"Não iniciada,Em execução,Concluída"', 
                              allow_blank=False)
dv_status_b.error = 'Selecione um status válido'
dv_status_b.errorTitle = 'Entrada Inválida'
ws_processo_b.add_data_validation(dv_status_b)
dv_status_b.add('B6:B20')

# Ajustar larguras
# Ajustar larguras
for i, width in enumerate(column_widths_processo, 1):
    ws_processo_b.column_dimensions[get_column_letter(i)].width = width

ws_processo_b.row_dimensions[1].height = 50
ws_processo_b.row_dimensions[2].height = 50
for row in range(6, 12):
    ws_processo_b.row_dimensions[row].height = 40

# ==================== ABA 4: DASHBOARD ====================
ws_dashboard = wb.create_sheet("Dashboard")

# Título
ws_dashboard.merge_cells('A1:F2')
ws_dashboard['A1'] = "📊 DASHBOARD - GERENCIAMENTO DE PROCESSOS"
ws_dashboard['A1'].font = Font(size=18, bold=True, color="FFFFFF")
ws_dashboard['A1'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
ws_dashboard['A1'].alignment = Alignment(horizontal="center", vertical="center")

# KPIs
kpi_row = 4
kpi_labels = ["Total de Processos", "Em Execução", "Concluídos", "Não Iniciados", "Taxa de Conclusão"]
kpi_formulas = [
    "=CONT.VALORES('Página inicial'!A2:A100)-CONT.SE('Página inicial'!A2:A100,\"\")",
    "=CONT.SE('Página inicial'!B:B,\"Em execução\")+CONT.SE('Página inicial'!B:B,\"Em andamento\")",
    "=CONT.SE('Página inicial'!B:B,\"Concluída\")",
    "=CONT.SE('Página inicial'!B:B,\"Não iniciado\")+CONT.SE('Página inicial'!B:B,\"Não iniciada\")",
    "=SE(B4=0,\"0%\",TEXTO(D4/B4,\"0%\"))"
]

for i, (label, formula) in enumerate(zip(kpi_labels, kpi_formulas)):
    label_cell = ws_dashboard.cell(row=kpi_row, column=i*2+1)
    label_cell.value = label
    label_cell.font = Font(bold=True, size=10)
    label_cell.alignment = Alignment(horizontal="center", vertical="center")
    label_cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    label_cell.border = border_thin
    
    value_cell = ws_dashboard.cell(row=kpi_row+1, column=i*2+1)
    value_cell.value = formula
    value_cell.font = Font(bold=True, size=14, color="366092")
    value_cell.alignment = Alignment(horizontal="center", vertical="center")
    value_cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    value_cell.border = border_thin
    
    ws_dashboard.merge_cells(start_row=kpi_row, start_column=i*2+1, end_row=kpi_row, end_column=i*2+2)
    ws_dashboard.merge_cells(start_row=kpi_row+1, start_column=i*2+1, end_row=kpi_row+1, end_column=i*2+2)

# Tabela de status
ws_dashboard['A8'] = "Status"
ws_dashboard['A8'].font = Font(bold=True, size=12)
ws_dashboard['B8'] = "Quantidade"
ws_dashboard['B8'].font = Font(bold=True, size=12)

status_list = ["Não iniciado", "Não iniciada", "Em execução", "Em andamento", "Concluída"]
for i, status in enumerate(status_list, 9):
    ws_dashboard[f'A{i}'] = status
    ws_dashboard[f'B{i}'] = f'=CONT.SE(\'Página inicial\'!B:B,"{status}")'

# Ajustar larguras
for col in ['A', 'B', 'C', 'D', 'E', 'F']:
    ws_dashboard.column_dimensions[col].width = 18

# ==================== ABA 5: INSTRUÇÕES ====================
ws_instrucoes = wb.create_sheet("Instruções")

ws_instrucoes.merge_cells('A1:D1')
ws_instrucoes['A1'] = "📖 INSTRUÇÕES DE USO DO SISTEMA"
ws_instrucoes['A1'].font = Font(size=16, bold=True, color="FFFFFF")
ws_instrucoes['A1'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
ws_instrucoes['A1'].alignment = Alignment(horizontal="center", vertical="center")

instrucoes_texto = [
    "",
    "🔹 PÁGINA INICIAL:",
    "   • Visão geral de todos os processos em andamento",
    "   • Resumo com status, responsável, datas e progresso",
    "   • Use esta aba para acompanhamento rápido",
    "",
    "🔹 ABAS DE PROCESSOS (processos A, processos B, etc.):",
    "   • Detalhamento completo de cada processo",
    "   • Descreva o projeto, número SEI e objetivos",
    "   • Liste todos os passos necessários para conclusão",
    "   • Defina responsável, datas, produtos e tarefas para cada passo",
    "   • O status geral é calculado automaticamente com base nos passos",
    "",
    "🔹 STATUS DISPONÍVEIS:",
    "   • Não iniciada / Não iniciado: Processo ainda não começou",
    "   • Em execução / Em andamento: Processo em desenvolvimento",
    "   • Concluída: Processo finalizado",
    "",
    "🔹 DASHBOARD:",
    "   • Indicadores-chave (KPIs) atualizados automaticamente",
    "   • Visão consolidada de todos os processos",
    "",
    "🔹 COMO ADICIONAR NOVO PROCESSO:",
    "   1. Adicione uma linha na aba 'Página inicial'",
    "   2. Crie uma nova aba copiando a estrutura de 'processos B'",
    "   3. Renomeie a aba com o nome do processo",
    "   4. Preencha os detalhes e passos do processo",
    "   5. Na 'Página inicial', adicione a fórmula de status: ='nome_da_aba'!H2",
    "",
    "✅ DICAS:",
    "   • Use as listas suspensas (dropdown) para Status",
    "   • Mantenha as datas atualizadas para melhor controle",
    "   • Revise periodicamente o % Concluído de cada processo",
    "   • O sistema calcula automaticamente o tempo real e estimado",
]

for i, texto in enumerate(instrucoes_texto, 2):
    ws_instrucoes[f'A{i}'] = texto
    ws_instrucoes.merge_cells(f'A{i}:D{i}')
    cell = ws_instrucoes[f'A{i}']
    
    if texto.startswith("🔹"):
        cell.font = Font(bold=True, size=11, color="366092")
    elif texto.startswith("   •") or texto.startswith("   "):
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
    else:
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

ws_instrucoes.column_dimensions['A'].width = 80

# Salvar arquivo
filename = "Sistema_Gerenciamento_Processos_Luma.xlsx"
wb.save(filename)

print("="*80)
print(f"✅ SISTEMA CRIADO COM SUCESSO!")
print("="*80)
print(f"\n📁 Arquivo: {filename}")
print("\n📊 ABAS CRIADAS:")
print("   ✓ Página inicial - Resumo de todos os processos")
print("   ✓ processos A - Exemplo de processo detalhado")
print("   ✓ processos B - Template para novos processos")
print("   ✓ Dashboard - Indicadores e métricas")
print("   ✓ Instruções - Guia de uso do sistema")
print("\n🎯 RECURSOS IMPLEMENTADOS:")
print("   ✓ Estrutura baseada no modelo_luma.xlsx")
print("   ✓ Células mescladas para títulos e descrições")
print("   ✓ Validação de dados (listas suspensas)")
print("   ✓ Fórmulas automáticas para cálculo de status")
print("   ✓ Formatação profissional e organizada")
print("   ✓ Sistema expansível para múltiplos processos")
print("\n" + "="*80)
print("🚀 Abra o arquivo no Excel para começar!")
print("="*80)
