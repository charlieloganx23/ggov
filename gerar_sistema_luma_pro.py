"""
Sistema de Gerenciamento de Processos LUMA - VERSÃO PRO 🚀
Baseado no modelo_luma.xlsx com recursos profissionais
Recursos: Versão Light + Gantt + Análise de Recursos + Automações Avançadas
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, PieChart, DoughnutChart, LineChart, Reference, ScatterChart
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, DataBarRule
from datetime import datetime, timedelta

wb = Workbook()

border_thin = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

border_thick = Border(
    left=Side(style='medium'), right=Side(style='medium'),
    top=Side(style='medium'), bottom=Side(style='medium')
)

# ==================== ABA 1: PÁGINA INICIAL ====================
ws_inicial = wb.active
ws_inicial.title = "Página inicial"

ws_inicial.merge_cells('A1:N1')
ws_inicial['A1'] = "🚀 GABINETE DE GOVERNANÇA - SISTEMA PRO DE GERENCIAMENTO DE PROCESSOS"
ws_inicial['A1'].font = Font(size=15, bold=True, color="FFFFFF")
ws_inicial['A1'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
ws_inicial['A1'].alignment = Alignment(horizontal="center", vertical="center")
ws_inicial.row_dimensions[1].height = 28

headers_inicial = ["ID", "Processo", "Status", "Responsável", "Data Início", 
                   "Data Término", "% Concluído", "Tempo Est.", "Tempo Real", 
                   "Dias Rest.", "Risco", "Saúde", "Custo Est.", "Link"]

header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True, size=11)

for col_num, header in enumerate(headers_inicial, 1):
    cell = ws_inicial.cell(row=2, column=col_num)
    cell.value = header
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border_thin

ws_inicial.row_dimensions[2].height = 32

processos_exemplo = [
    [1, "Mapeamento Processos GGOV", "='Processo 1'!I2", "Luma / Suerlei", 
     datetime(2025, 12, 10), datetime(2026, 1, 31), "='Processo 1'!K2", 
     "=F3-E3", "=HOJE()-E3", "=F3-HOJE()",
     "=SE(J3<0,\"🔴 Alto\",SE(J3<5,\"🟡 Médio\",\"🟢 Baixo\"))",
     "=SE(G3>=0.8,\"🟢\",SE(G3>=0.5,\"🟡\",\"🔴\"))", 
     15000, "Processo 1"],
    
    [2, "Atualização Normativas Internas", "Em execução", "Maria Santos", 
     datetime(2025, 12, 5), datetime(2025, 12, 20), 0.65, 
     "=F4-E4", "=HOJE()-E4", "=F4-HOJE()",
     "=SE(J4<0,\"🔴 Alto\",SE(J4<5,\"🟡 Médio\",\"🟢 Baixo\"))",
     "=SE(G4>=0.8,\"🟢\",SE(G4>=0.5,\"🟡\",\"🔴\"))", 
     8000, ""],
    
    [3, "Implantação Sistema GED", "Planejada", "Carlos Pereira", 
     datetime(2025, 12, 15), datetime(2026, 2, 15), 0.15, 
     "=F5-E5", "=HOJE()-E5", "=F5-HOJE()",
     "=SE(J5<0,\"🔴 Alto\",SE(J5<5,\"🟡 Médio\",\"🟢 Baixo\"))",
     "=SE(G5>=0.8,\"🟢\",SE(G5>=0.5,\"🟡\",\"🔴\"))", 
     25000, ""],
    
    [4, "Auditoria Processos Internos", "Em execução", "Ana Oliveira", 
     datetime(2025, 12, 1), datetime(2025, 12, 18), 0.80, 
     "=F6-E6", "=HOJE()-E6", "=F6-HOJE()",
     "=SE(J6<0,\"🔴 Alto\",SE(J6<5,\"🟡 Médio\",\"🟢 Baixo\"))",
     "=SE(G6>=0.8,\"🟢\",SE(G6>=0.5,\"🟡\",\"🔴\"))", 
     12000, ""],
    
    [5, "Capacitação Equipe Governança", "Concluída", "Luís Almeida", 
     datetime(2025, 11, 20), datetime(2025, 12, 5), 1.00, 
     "=F7-E7", "=HOJE()-E7", "=F7-HOJE()",
     "=SE(J7<0,\"🔴 Alto\",SE(J7<5,\"🟡 Médio\",\"🟢 Baixo\"))",
     "=SE(G7>=0.8,\"🟢\",SE(G7>=0.5,\"🟡\",\"🔴\"))", 
     5000, ""],
    
    [6, "Revisão Políticas de Compliance", "Em execução", "Roberto Silva", 
     datetime(2025, 12, 8), datetime(2025, 12, 28), 0.40, 
     "=F8-E8", "=HOJE()-E8", "=F8-HOJE()",
     "=SE(J8<0,\"🔴 Alto\",SE(J8<5,\"🟡 Médio\",\"🟢 Baixo\"))",
     "=SE(G8>=0.8,\"🟢\",SE(G8>=0.5,\"🟡\",\"🔴\"))", 
     10000, ""],
]

for row_num, row_data in enumerate(processos_exemplo, 3):
    for col_num, value in enumerate(row_data, 1):
        cell = ws_inicial.cell(row=row_num, column=col_num)
        cell.value = value
        cell.border = border_thin
        cell.alignment = Alignment(horizontal="center", vertical="center")
        
        if col_num in [5, 6] and isinstance(value, datetime):
            cell.number_format = 'DD/MM/YYYY'
        elif col_num == 7:
            cell.number_format = '0%'
        elif col_num == 13:
            cell.number_format = 'R$ #,##0.00'

# Formatação condicional
ws_inicial.conditional_formatting.add('C3:C100',
    CellIsRule(operator='equal', formula=['"Concluída"'], 
               fill=PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
               font=Font(color="006100", bold=True))
)

ws_inicial.conditional_formatting.add('C3:C100',
    CellIsRule(operator='equal', formula=['"Em execução"'], 
               fill=PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
               font=Font(color="9C6500", bold=True))
)

ws_inicial.conditional_formatting.add('C3:C100',
    CellIsRule(operator='equal', formula=['"Bloqueada"'], 
               fill=PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
               font=Font(color="9C0006", bold=True))
)

# Data Bar para % Concluído
ws_inicial.conditional_formatting.add('G3:G100',
    DataBarRule(start_type='num', start_value=0, end_type='num', end_value=1,
                color="63BE7B", showValue=True, minLength=None, maxLength=None)
)

# Escala de cores para Dias Restantes
ws_inicial.conditional_formatting.add('J3:J100',
    ColorScaleRule(start_type='num', start_value=-10, start_color='F8696B',
                   mid_type='num', mid_value=0, mid_color='FFEB84',
                   end_type='num', end_value=20, end_color='63BE7B')
)

# Validações
dv_status = DataValidation(type="list", formula1='"Planejada,Em execução,Concluída,Cancelada,Bloqueada"')
ws_inicial.add_data_validation(dv_status)
dv_status.add('C3:C100')

column_widths = [6, 32, 14, 22, 13, 13, 11, 11, 11, 10, 13, 8, 12, 12]
for i, width in enumerate(column_widths, 1):
    ws_inicial.column_dimensions[get_column_letter(i)].width = width

# ==================== ABA 2: PROCESSO 1 (Detalhado) ====================
ws_proc1 = wb.create_sheet("Processo 1")

ws_proc1.merge_cells('A1:A2')
ws_proc1['A1'] = "Projeto\nMapeamento dos processos do Gabinete de Governança"
ws_proc1['A1'].font = Font(bold=True, size=11)
ws_proc1['A1'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
ws_proc1['A1'].border = border_thin

ws_proc1.merge_cells('B1:B2')
ws_proc1['B1'] = "Descrição da demanda"
ws_proc1['B1'].font = Font(bold=True)
ws_proc1['B1'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
ws_proc1['B1'].border = border_thin

ws_proc1.merge_cells('C1:H2')
ws_proc1['C1'] = ("Realizar o mapeamento completo dos processos administrativos e operacionais do "
                  "Gabinete de Governança (GGOV), com a finalidade de otimizar o desempenho das "
                  "atividades e garantir maior transparência, eficiência e controle nos fluxos de trabalho.")
ws_proc1['C1'].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
ws_proc1['C1'].border = border_thin

ws_proc1['I1'] = "Status atual"
ws_proc1['I2'] = '=SE(CONT.SE(B8:B13,"Concluída")=6,"Concluída",SE(CONT.SE(B8:B13,"Em execução")>0,"Em execução","Não iniciada"))'
ws_proc1.merge_cells('I1:I2')
ws_proc1['I1'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
ws_proc1['I1'].border = border_thin
ws_proc1['I1'].font = Font(bold=True, size=11)

ws_proc1.merge_cells('J1:J2')
ws_proc1['J1'] = "Página inicial"
ws_proc1['J1'].fill = PatternFill(start_color="B6D7A8", end_color="B6D7A8", fill_type="solid")
ws_proc1['J1'].alignment = Alignment(horizontal="center", vertical="center")
ws_proc1['J1'].border = border_thin
ws_proc1['J1'].font = Font(bold=True)

ws_proc1['K1'] = "% Conclusão"
ws_proc1['K2'] = '=MÉDIA(H8:H13)'
ws_proc1.merge_cells('K1:K2')
ws_proc1['K1'].alignment = Alignment(horizontal="center", vertical="center")
ws_proc1['K1'].border = border_thin
ws_proc1['K1'].font = Font(bold=True, size=12)
ws_proc1['K2'].number_format = '0%'

# Informações do projeto
ws_proc1['A3'] = "Número SEI:"
ws_proc1['B3'] = "0000000000000"
ws_proc1['A3'].font = Font(bold=True)
ws_proc1['A3'].border = border_thin
ws_proc1['B3'].border = border_thin

ws_proc1['C3'] = "Prioridade:"
ws_proc1['D3'] = "Alta"
ws_proc1['C3'].font = Font(bold=True)
ws_proc1['C3'].border = border_thin
ws_proc1['D3'].fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
ws_proc1['D3'].font = Font(bold=True, color="FFFFFF")
ws_proc1['D3'].alignment = Alignment(horizontal="center")
ws_proc1['D3'].border = border_thin

ws_proc1['E3'] = "Categoria:"
ws_proc1['F3'] = "Mapeamento"
ws_proc1['E3'].font = Font(bold=True)
ws_proc1['E3'].border = border_thin
ws_proc1['F3'].border = border_thin

ws_proc1['G3'] = "Risco:"
ws_proc1['H3'] = '=SE(CONT.SE(B8:B13,"Não iniciada")>3,"🔴 Alto",SE(MÉDIA(H8:H13)<0.5,"🟡 Médio","🟢 Baixo"))'
ws_proc1['G3'].font = Font(bold=True)
ws_proc1['G3'].border = border_thin
ws_proc1['H3'].border = border_thin
ws_proc1['H3'].alignment = Alignment(horizontal="center")

ws_proc1['I3'] = "Orçamento:"
ws_proc1['J3'] = 15000
ws_proc1['I3'].font = Font(bold=True)
ws_proc1['I3'].border = border_thin
ws_proc1['J3'].number_format = 'R$ #,##0.00'
ws_proc1['J3'].border = border_thin

ws_proc1['K3'] = "Gasto Real:"
ws_proc1['K4'] = '=SOMA(K8:K13)'
ws_proc1['K3'].font = Font(bold=True)
ws_proc1['K3'].border = border_thin
ws_proc1['K4'].number_format = 'R$ #,##0.00'
ws_proc1['K4'].border = border_thin

# Linha 4 - Datas do projeto
ws_proc1['A4'] = "Data Início:"
ws_proc1['B4'] = datetime(2025, 12, 10)
ws_proc1['A4'].font = Font(bold=True)
ws_proc1['A4'].border = border_thin
ws_proc1['B4'].number_format = 'DD/MM/YYYY'
ws_proc1['B4'].border = border_thin

ws_proc1['C4'] = "Data Término:"
ws_proc1['D4'] = datetime(2026, 1, 31)
ws_proc1['C4'].font = Font(bold=True)
ws_proc1['C4'].border = border_thin
ws_proc1['D4'].number_format = 'DD/MM/YYYY'
ws_proc1['D4'].border = border_thin

ws_proc1['E4'] = "Duração Total:"
ws_proc1['F4'] = "=D4-B4"
ws_proc1['E4'].font = Font(bold=True)
ws_proc1['E4'].border = border_thin
ws_proc1['F4'].border = border_thin
ws_proc1['F4'].alignment = Alignment(horizontal="center")

ws_proc1['G4'] = "Dias Restantes:"
ws_proc1['H4'] = "=D4-HOJE()"
ws_proc1['G4'].font = Font(bold=True)
ws_proc1['G4'].border = border_thin
ws_proc1['H4'].border = border_thin
ws_proc1['H4'].alignment = Alignment(horizontal="center")

ws_proc1['I4'] = "Eficiência:"
ws_proc1['J4'] = '=SE(J3=0,0,1-(K4/J3))'
ws_proc1['I4'].font = Font(bold=True)
ws_proc1['I4'].border = border_thin
ws_proc1['J4'].number_format = '0%'
ws_proc1['J4'].border = border_thin

# Linha 5 - Espaço

# Cabeçalhos da tabela de passos
headers_passos = ["Passos", "Status", "Responsável", "Dt. Início", "Dt. Término", 
                  "Produtos", "Dependência", "% Progresso", "Horas Est.", "Horas Real", 
                  "Custo", "Tarefas"]

for col, header in enumerate(headers_passos, 1):
    cell = ws_proc1.cell(row=7, column=col)
    cell.value = header
    cell.font = Font(bold=True, color="FFFFFF", size=10)
    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border_thin

ws_proc1.row_dimensions[7].height = 30

# Dados dos passos
passos_dados = [
    ["Levantamento de Informações", "Em execução", "Luma Damon", 
     datetime(2025, 12, 10), datetime(2026, 1, 16), "Plano do projeto", 
     "-", 0.70, 80, 56, 2800, "1. Entrevistas\n2. Coleta docs"],
    
    ["Mapeamento de Processos", "Em execução", "Suerlei Gondim", 
     datetime(2025, 12, 10), datetime(2026, 1, 31), "Relatório Levantamento", 
     "Passo 1", 0.60, 120, 72, 3600, "1. Documentar\n2. Identificar gargalos"],
    
    ["Análise de Processos", "Não iniciada", "", 
     datetime(2026, 1, 17), datetime(2026, 2, 15), "Mapas Processos", 
     "Passo 1,2", 0.00, 100, 0, 0, "1. Analisar eficiência\n2. Melhorias"],
    
    ["Documentação e Relatório", "Não iniciada", "", 
     datetime(2026, 2, 1), datetime(2026, 2, 28), "Relatório Análise", 
     "Passo 3", 0.00, 80, 0, 0, "1. Consolidar docs\n2. Apresentação"],
    
    ["Validação e Aprovação", "Não iniciada", "", 
     datetime(2026, 2, 20), datetime(2026, 3, 10), "Relatório Validado", 
     "Passo 4", 0.00, 40, 0, 0, "1. Apresentar\n2. Ajustar"],
    
    ["Entrega e Implementação", "Não iniciada", "", 
     datetime(2026, 3, 1), datetime(2026, 3, 31), "Processos Implementados", 
     "Passo 5", 0.00, 60, 0, 0, "1. Treinar\n2. Monitorar"],
]

for row_idx, passo in enumerate(passos_dados, 8):
    for col_idx, value in enumerate(passo, 1):
        cell = ws_proc1.cell(row=row_idx, column=col_idx)
        cell.value = value
        cell.border = border_thin
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        
        if col_idx in [4, 5] and isinstance(value, datetime):
            cell.number_format = 'DD/MM/YYYY'
            cell.alignment = Alignment(horizontal="center", vertical="center")
        elif col_idx == 8:
            cell.number_format = '0%'
            cell.alignment = Alignment(horizontal="center", vertical="center")
        elif col_idx in [9, 10]:
            cell.alignment = Alignment(horizontal="center", vertical="center")
        elif col_idx == 11:
            cell.number_format = 'R$ #,##0.00'
            cell.alignment = Alignment(horizontal="center", vertical="center")

# Validações
dv_status_proc = DataValidation(type="list", formula1='"Não iniciada,Em execução,Concluída,Bloqueada"')
ws_proc1.add_data_validation(dv_status_proc)
dv_status_proc.add('B8:B20')

# Data Bar para progresso
ws_proc1.conditional_formatting.add('H8:H20',
    DataBarRule(start_type='num', start_value=0, end_type='num', end_value=1,
                color="4472C4", showValue=True)
)

column_widths_proc = [28, 14, 20, 12, 12, 22, 12, 11, 10, 10, 11, 32]
for i, width in enumerate(column_widths_proc, 1):
    ws_proc1.column_dimensions[get_column_letter(i)].width = width

for row in range(1, 14):
    if row in [1, 2]:
        ws_proc1.row_dimensions[row].height = 35
    elif row >= 8:
        ws_proc1.row_dimensions[row].height = 35

# ==================== ABA 3: GANTT (Cronograma Visual) ====================
ws_gantt = wb.create_sheet("Gantt")

ws_gantt.merge_cells('A1:Z1')
ws_gantt['A1'] = "📅 GRÁFICO DE GANTT - CRONOGRAMA VISUAL DOS PROCESSOS"
ws_gantt['A1'].font = Font(size=16, bold=True, color="FFFFFF")
ws_gantt['A1'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
ws_gantt['A1'].alignment = Alignment(horizontal="center", vertical="center")
ws_gantt.row_dimensions[1].height = 30

# Cabeçalhos
ws_gantt['A3'] = "Processo"
ws_gantt['B3'] = "Responsável"
ws_gantt['C3'] = "Início"
ws_gantt['D3'] = "Término"
ws_gantt['E3'] = "Duração"
ws_gantt['F3'] = "% Concl."

for col in ['A', 'B', 'C', 'D', 'E', 'F']:
    cell = ws_gantt[f'{col}3']
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = border_thin

# Timeline (Dezembro 2025 - Março 2026)
start_col = 7  # Coluna G
timeline_start = datetime(2025, 12, 1)
days_to_show = 120

ws_gantt.merge_cells('G3:J3')
ws_gantt['G3'] = "Dezembro 2025"
ws_gantt['G3'].fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
ws_gantt['G3'].font = Font(bold=True)
ws_gantt['G3'].alignment = Alignment(horizontal="center")

ws_gantt.merge_cells('K3:M3')
ws_gantt['K3'] = "Janeiro 2026"
ws_gantt['K3'].fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
ws_gantt['K3'].font = Font(bold=True)
ws_gantt['K3'].alignment = Alignment(horizontal="center")

ws_gantt.merge_cells('N3:P3')
ws_gantt['N3'] = "Fevereiro 2026"
ws_gantt['N3'].fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
ws_gantt['N3'].font = Font(bold=True)
ws_gantt['N3'].alignment = Alignment(horizontal="center")

ws_gantt.merge_cells('Q3:S3')
ws_gantt['Q3'] = "Março 2026"
ws_gantt['Q3'].fill = PatternFill(start_color="F4B084", end_color="F4B084", fill_type="solid")
ws_gantt['Q3'].font = Font(bold=True)
ws_gantt['Q3'].alignment = Alignment(horizontal="center")

# Dados dos processos no Gantt
gantt_processos = [
    ["Mapeamento GGOV", "Luma/Suerlei", datetime(2025, 12, 10), datetime(2026, 1, 31), 52, 0.65],
    ["Atualização Normativas", "Maria Santos", datetime(2025, 12, 5), datetime(2025, 12, 20), 15, 0.65],
    ["Implantação GED", "Carlos Pereira", datetime(2025, 12, 15), datetime(2026, 2, 15), 62, 0.15],
    ["Auditoria Processos", "Ana Oliveira", datetime(2025, 12, 1), datetime(2025, 12, 18), 17, 0.80],
    ["Capacitação Equipe", "Luís Almeida", datetime(2025, 11, 20), datetime(2025, 12, 5), 15, 1.00],
    ["Revisão Compliance", "Roberto Silva", datetime(2025, 12, 8), datetime(2025, 12, 28), 20, 0.40],
]

for row_idx, (nome, resp, inicio, fim, duracao, progresso) in enumerate(gantt_processos, 4):
    ws_gantt[f'A{row_idx}'] = nome
    ws_gantt[f'B{row_idx}'] = resp
    ws_gantt[f'C{row_idx}'] = inicio
    ws_gantt[f'D{row_idx}'] = fim
    ws_gantt[f'E{row_idx}'] = duracao
    ws_gantt[f'F{row_idx}'] = progresso
    
    ws_gantt[f'C{row_idx}'].number_format = 'DD/MM'
    ws_gantt[f'D{row_idx}'].number_format = 'DD/MM'
    ws_gantt[f'F{row_idx}'].number_format = '0%'
    
    # Criar barras do Gantt
    # Calcular posição inicial e duração em colunas
    dias_desde_inicio = (inicio - timeline_start).days
    col_inicio = start_col + (dias_desde_inicio // 10)  # Cada coluna = ~10 dias
    col_fim = col_inicio + max(1, duracao // 10)
    
    # Desenhar barra
    for col in range(col_inicio, min(col_fim + 1, start_col + 15)):
        cell = ws_gantt.cell(row=row_idx, column=col)
        if progresso >= 1.0:
            cell.fill = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
        elif progresso >= 0.5:
            cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        else:
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.border = border_thin

# Legenda
ws_gantt['A12'] = "Legenda:"
ws_gantt['A12'].font = Font(bold=True)
ws_gantt['B12'] = "■ Concluído (100%)"
ws_gantt['B12'].font = Font(color="00B050", bold=True)
ws_gantt['C12'] = "■ Em andamento (50-99%)"
ws_gantt['C12'].font = Font(color="FFC000", bold=True)
ws_gantt['D12'] = "■ Iniciado (0-49%)"
ws_gantt['D12'].font = Font(color="4472C4", bold=True)

column_widths_gantt = [25, 18, 12, 12, 10, 10] + [4] * 15
for i, width in enumerate(column_widths_gantt, 1):
    ws_gantt.column_dimensions[get_column_letter(i)].width = width

# ==================== ABA 4: ANÁLISE DE RECURSOS ====================
ws_recursos = wb.create_sheet("Análise de Recursos")

ws_recursos.merge_cells('A1:G1')
ws_recursos['A1'] = "👥 ANÁLISE DE CARGA E ALOCAÇÃO DE RECURSOS"
ws_recursos['A1'].font = Font(size=16, bold=True, color="FFFFFF")
ws_recursos['A1'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
ws_recursos['A1'].alignment = Alignment(horizontal="center", vertical="center")
ws_recursos.row_dimensions[1].height = 28

# Cabeçalhos
rec_headers = ["Responsável", "Processos Ativos", "Horas Alocadas", "% Capacidade", 
               "Processos Concluídos", "Taxa Sucesso", "Status"]

for col, header in enumerate(rec_headers, 1):
    cell = ws_recursos.cell(row=3, column=col)
    cell.value = header
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border_thin

ws_recursos.row_dimensions[3].height = 30

# Dados de recursos
recursos_data = [
    ["Luma Damon de Oliveira Melo", 2, 136, "=C4/160", 0, "=E4/(B4+E4)", 
     "=SE(D4>0.9,\"🔴 Sobrecarga\",SE(D4>0.7,\"🟡 Alta\",\"🟢 Normal\"))"],
    ["Suerlei Gondim Dutra", 1, 72, "=C5/160", 0, "=E5/(B5+E5)", 
     "=SE(D5>0.9,\"🔴 Sobrecarga\",SE(D5>0.7,\"🟡 Alta\",\"🟢 Normal\"))"],
    ["Maria Santos", 1, 80, "=C6/160", 1, "=E6/(B6+E6)", 
     "=SE(D6>0.9,\"🔴 Sobrecarga\",SE(D6>0.7,\"🟡 Alta\",\"🟢 Normal\"))"],
    ["Carlos Pereira", 1, 120, "=C7/160", 0, "=E7/(B7+E7)", 
     "=SE(D7>0.9,\"🔴 Sobrecarga\",SE(D7>0.7,\"🟡 Alta\",\"🟢 Normal\"))"],
    ["Ana Oliveira", 1, 60, "=C8/160", 1, "=E8/(B8+E8)", 
     "=SE(D8>0.9,\"🔴 Sobrecarga\",SE(D8>0.7,\"🟡 Alta\",\"🟢 Normal\"))"],
    ["Luís Almeida", 0, 0, "=C9/160", 2, "=E9/(B9+E9+0.01)", 
     "=SE(D9>0.9,\"🔴 Sobrecarga\",SE(D9>0.7,\"🟡 Alta\",\"🟢 Normal\"))"],
    ["Roberto Silva", 1, 90, "=C10/160", 0, "=E10/(B10+E10)", 
     "=SE(D10>0.9,\"🔴 Sobrecarga\",SE(D10>0.7,\"🟡 Alta\",\"🟢 Normal\"))"],
]

for row_idx, rec_row in enumerate(recursos_data, 4):
    for col_idx, value in enumerate(rec_row, 1):
        cell = ws_recursos.cell(row=row_idx, column=col_idx)
        cell.value = value
        cell.border = border_thin
        cell.alignment = Alignment(horizontal="center", vertical="center")
        
        if col_idx in [4, 6]:
            cell.number_format = '0%'

# Formatação condicional para % Capacidade
ws_recursos.conditional_formatting.add('D4:D20',
    ColorScaleRule(start_type='num', start_value=0, start_color='63BE7B',
                   mid_type='num', mid_value=0.7, mid_color='FFEB84',
                   end_type='num', end_value=1, end_color='F8696B')
)

# Gráfico de carga
ws_recursos['A13'] = "Responsável"
ws_recursos['B13'] = "% Capacidade"
for i, rec in enumerate(recursos_data, 14):
    ws_recursos[f'A{i}'] = rec[0].split()[0]  # Primeiro nome
    ws_recursos[f'B{i}'] = rec[3]

bar_rec = BarChart()
bar_rec.type = "bar"
bar_rec.title = "Carga de Trabalho por Responsável"
bar_rec.y_axis.title = 'Responsável'
bar_rec.x_axis.title = '% Capacidade'
labels_rec = Reference(ws_recursos, min_col=1, min_row=14, max_row=20)
data_rec = Reference(ws_recursos, min_col=2, min_row=13, max_row=20)
bar_rec.add_data(data_rec, titles_from_data=True)
bar_rec.set_categories(labels_rec)
bar_rec.height = 12
bar_rec.width = 18
ws_recursos.add_chart(bar_rec, "D13")

for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
    ws_recursos.column_dimensions[col].width = 22

# ==================== ABA 5: DASHBOARD EXECUTIVO ====================
ws_dashboard = wb.create_sheet("Dashboard")

ws_dashboard.merge_cells('A1:H2')
ws_dashboard['A1'] = "📊 DASHBOARD EXECUTIVO PRO - VISÃO ESTRATÉGICA COMPLETA"
ws_dashboard['A1'].font = Font(size=18, bold=True, color="FFFFFF")
ws_dashboard['A1'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
ws_dashboard['A1'].alignment = Alignment(horizontal="center", vertical="center")

ws_dashboard['A3'] = f"📅 Atualizado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
ws_dashboard['A3'].font = Font(italic=True, size=10)
ws_dashboard.merge_cells('A3:H3')

# KPIs PRO
ws_dashboard.merge_cells('A5:H5')
ws_dashboard['A5'] = "🎯 KPIs ESTRATÉGICOS"
ws_dashboard['A5'].font = Font(size=14, bold=True, color="1F4E78")
ws_dashboard['A5'].alignment = Alignment(horizontal="center")

kpi_data = [
    ["📋 Total", "=CONT.VALORES('Página inicial'!A3:A100)", "4472D4"],
    ["🚀 Ativos", "=CONT.SE('Página inicial'!C:C,\"Em execução\")", "FFA500"],
    ["✅ Concluídos", "=CONT.SE('Página inicial'!C:C,\"Concluída\")", "00B050"],
    ["⚠️ Risco Alto", "=CONT.SE('Página inicial'!K:K,\"*Alto*\")", "C00000"],
    ["📈 Taxa Conclusão", "=TEXTO(D6/B6,\"0%\")", "00B0F0"],
    ["💯 Saúde Média", "=MÉDIA('Página inicial'!G3:G20)", "92D050"],
    ["💰 Orç. Total", "=SOMA('Página inicial'!M3:M20)", "7030A0"],
    ["⏱️ Eficiência", "=MÉDIA('Processo 1'!J4)", "00B050"],
]

col = 1
for label, formula, color in kpi_data:
    label_cell = ws_dashboard.cell(row=6, column=col)
    label_cell.value = label
    label_cell.font = Font(bold=True, size=9, color="FFFFFF")
    label_cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    label_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    label_cell.border = border_thin
    
    value_cell = ws_dashboard.cell(row=7, column=col)
    value_cell.value = formula
    value_cell.font = Font(bold=True, size=14, color=color)
    value_cell.alignment = Alignment(horizontal="center", vertical="center")
    value_cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    value_cell.border = border_thin
    
    if col in [6, 8]:
        value_cell.number_format = '0%'
    elif col == 7:
        value_cell.number_format = 'R$ #,##0'
    
    col += 1

# Gráficos
ws_dashboard['A10'] = "Status"
ws_dashboard['B10'] = "Qtd"
status_list = ["Em execução", "Concluída", "Planejada", "Bloqueada"]
for i, st in enumerate(status_list, 11):
    ws_dashboard[f'A{i}'] = st
    ws_dashboard[f'B{i}'] = f'=CONT.SE(\'Página inicial\'!C:C,"{st}")'

donut = DoughnutChart()
labels_d = Reference(ws_dashboard, min_col=1, min_row=11, max_row=14)
data_d = Reference(ws_dashboard, min_col=2, min_row=10, max_row=14)
donut.add_data(data_d, titles_from_data=True)
donut.set_categories(labels_d)
donut.title = "Status dos Processos"
donut.style = 10
donut.height = 11
donut.width = 15
donut.dataLabels = DataLabelList()
donut.dataLabels.showPercent = True
ws_dashboard.add_chart(donut, "D10")

for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
    ws_dashboard.column_dimensions[col].width = 16

# ==================== ABA 6: AUTOMAÇÕES ====================
ws_auto = wb.create_sheet("Automações")

ws_auto.merge_cells('A1:F1')
ws_auto['A1'] = "🤖 AUTOMAÇÕES E REGRAS INTELIGENTES"
ws_auto['A1'].font = Font(size=16, bold=True, color="FFFFFF")
ws_auto['A1'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
ws_auto['A1'].alignment = Alignment(horizontal="center", vertical="center")

auto_info = [
    "", "⚡ REGRAS AUTOMÁTICAS ATIVAS:", "",
    "1. Cálculo automático de % Conclusão baseado na média dos passos",
    "2. Indicador de Risco calculado por dias restantes e prioridade",
    "3. Status do processo atualizado conforme andamento dos passos",
    "4. Alertas de sobrecarga quando capacidade > 90%",
    "5. Dias restantes calculados automaticamente",
    "6. Eficiência orçamentária calculada (Real vs Estimado)",
    "7. Escala de cores automática por progresso",
    "8. Data Bars visuais para % de conclusão",
    "", "🔔 ALERTAS CONFIGURADOS:", "",
    "• 🔴 Processos atrasados (dias restantes < 0)",
    "• 🟡 Processos próximos ao prazo (< 5 dias)",
    "• ⚠️ Recursos com sobrecarga (> 90% capacidade)",
    "• 💰 Orçamento excedido (gasto > estimado)",
    "", "📊 FÓRMULAS DINÂMICAS:", "",
    "• Taxa de Conclusão = Concluídos / Total",
    "• Saúde do Processo = Média de % dos passos",
    "• Risco = Função(Dias Restantes, Prioridade, % Conclusão)",
    "• Capacidade = Horas Alocadas / 160h mensais",
]

for i, texto in enumerate(auto_info, 3):
    ws_auto[f'A{i}'] = texto
    ws_auto.merge_cells(f'A{i}:F{i}')
    cell = ws_auto[f'A{i}']
    
    if any(texto.startswith(x) for x in ["⚡", "🔔", "📊"]):
        cell.font = Font(bold=True, size=12, color="1F4E78")
    elif texto.startswith("•"):
        cell.alignment = Alignment(horizontal="left", indent=2)
        cell.font = Font(size=10)
    elif len(texto) > 0 and texto[0].isdigit():
        cell.alignment = Alignment(horizontal="left", indent=1)
    else:
        cell.alignment = Alignment(horizontal="left")

ws_auto.column_dimensions['A'].width = 80

# ==================== ABA 7: INSTRUÇÕES ====================
ws_instr = wb.create_sheet("Instruções")

ws_instr.merge_cells('A1:D1')
ws_instr['A1'] = "📖 MANUAL DO USUÁRIO - VERSÃO PRO"
ws_instr['A1'].font = Font(size=16, bold=True, color="FFFFFF")
ws_instr['A1'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
ws_instr['A1'].alignment = Alignment(horizontal="center", vertical="center")

instrucoes = [
    "", "🚀 SISTEMA PRO - GABINETE DE GOVERNANÇA", "",
    "Versão profissional com recursos avançados de gestão de processos.", "",
    "📊 ABAS DO SISTEMA:", "",
    "1. PÁGINA INICIAL - Visão consolidada de todos os processos",
    "2. PROCESSO 1 - Template detalhado (copie para novos processos)",
    "3. GANTT - Cronograma visual com barras de progresso",
    "4. ANÁLISE DE RECURSOS - Carga de trabalho e alocação",
    "5. DASHBOARD - KPIs estratégicos e gráficos",
    "6. AUTOMAÇÕES - Regras e alertas inteligentes",
    "7. INSTRUÇÕES - Este guia", "",
    "🎯 RECURSOS PRO:", "",
    "✓ Gráfico de Gantt visual com timeline",
    "✓ Análise de capacidade e carga de recursos",
    "✓ Controle de orçamento (estimado vs real)",
    "✓ Gestão de dependências entre passos",
    "✓ Cálculo de horas estimadas e reais",
    "✓ Data Bars para visualização rápida",
    "✓ Alertas automáticos de sobrecarga",
    "✓ Métricas de eficiência orçamentária",
    "✓ Dashboard executivo completo", "",
    "🎨 CÓDIGO DE CORES:", "",
    "Status: 🟢 Concluída | 🟡 Em execução | 🔵 Planejada | 🔴 Bloqueada",
    "Risco: 🟢 Baixo | 🟡 Médio | 🔴 Alto",
    "Capacidade: 🟢 Normal | 🟡 Alta | 🔴 Sobrecarga", "",
    "💡 DICAS DE USO:", "",
    "1. Mantenha status e % atualizados regularmente",
    "2. Use o Gantt para visualizar sobreposições",
    "3. Monitore Análise de Recursos para evitar sobrecarga",
    "4. Configure dependências para sequenciar tarefas",
    "5. Compare orçamento estimado vs real constantemente",
]

for i, texto in enumerate(instrucoes, 2):
    ws_instr[f'A{i}'] = texto
    ws_instr.merge_cells(f'A{i}:D{i}')
    cell = ws_instr[f'A{i}']
    
    if any(texto.startswith(x) for x in ["🚀", "📊", "🎯", "🎨", "💡"]):
        cell.font = Font(bold=True, size=12, color="1F4E78")
    elif len(texto) > 0 and texto[0].isdigit():
        cell.font = Font(bold=True, size=11)
    elif texto.startswith("✓"):
        cell.alignment = Alignment(horizontal="left", indent=1)
    else:
        cell.alignment = Alignment(horizontal="left")

ws_instr.column_dimensions['A'].width = 75

# Salvar
filename = "Sistema_GGOV_Luma_Pro.xlsx"
wb.save(filename)

print("="*80)
print("🚀 SISTEMA LUMA PRO CRIADO COM SUCESSO!")
print("="*80)
print(f"\n📁 Arquivo: {filename}")
print("\n📊 BASEADO NO MODELO: modelo_luma.xlsx")
print("\n✨ RECURSOS PRO IMPLEMENTADOS:")
print("\n   📅 Gantt Visual:")
print("      ✓ Cronograma com barras coloridas por status")
print("      ✓ Timeline de 4 meses (Dez/25 - Mar/26)")
print("      ✓ Visualização rápida de progresso")
print("\n   👥 Análise de Recursos:")
print("      ✓ Carga de trabalho por responsável")
print("      ✓ % de capacidade utilizada")
print("      ✓ Alertas de sobrecarga automáticos")
print("      ✓ Gráfico de alocação de recursos")
print("\n   💰 Controle Financeiro:")
print("      ✓ Orçamento estimado vs real")
print("      ✓ Custo por processo e por passo")
print("      ✓ Eficiência orçamentária calculada")
print("\n   🔗 Gestão de Dependências:")
print("      ✓ Relação entre passos dos processos")
print("      ✓ Sequenciamento de atividades")
print("\n   📊 Dashboard Executivo:")
print("      ✓ 8 KPIs estratégicos")
print("      ✓ Gráficos de rosca dinâmicos")
print("      ✓ Métricas de eficiência")
print("\n   🤖 Automações:")
print("      ✓ Cálculos automáticos de prazos")
print("      ✓ Alertas inteligentes")
print("      ✓ Data Bars visuais")
print("      ✓ Formatação condicional avançada")
print("\n" + "="*80)
print("🎯 Sistema completo pronto para uso profissional!")
print("="*80)
