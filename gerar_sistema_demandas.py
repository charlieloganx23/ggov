"""
Sistema de Gerenciamento de Demandas - Excel
Criado com Python + openpyxl
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import datetime, timedelta

# Criar workbook
wb = Workbook()

# ==================== ABA 1: CONTROLE DE DEMANDAS ====================
ws_demandas = wb.active
ws_demandas.title = "Controle de Demandas"

# Cabeçalhos
headers = ["ID", "Demanda", "Responsável", "Prioridade", "Status", 
           "Data Abertura", "Prazo", "Data Conclusão", "Dias Decorridos", 
           "Dias Restantes", "% Progresso", "Categoria"]

# Estilo do cabeçalho
header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True, size=11)
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Aplicar cabeçalhos
for col_num, header in enumerate(headers, 1):
    cell = ws_demandas.cell(row=1, column=col_num)
    cell.value = header
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = border

# Dados de exemplo
dados_exemplo = [
    [1, "Desenvolvimento do módulo de login", "João Silva", "Alta", "Em Andamento", 
     datetime(2025, 12, 1), datetime(2025, 12, 15), "", "=HOJE()-F2", "=G2-HOJE()", "60%", "Desenvolvimento"],
    [2, "Correção de bugs no relatório", "Maria Santos", "Crítica", "Em Andamento", 
     datetime(2025, 12, 5), datetime(2025, 12, 12), "", "=HOJE()-F3", "=G3-HOJE()", "80%", "Correção"],
    [3, "Implementar dashboard analytics", "Pedro Costa", "Média", "Planejada", 
     datetime(2025, 12, 8), datetime(2025, 12, 20), "", "=HOJE()-F4", "=G4-HOJE()", "30%", "Desenvolvimento"],
    [4, "Atualização da documentação", "Ana Oliveira", "Baixa", "Concluída", 
     datetime(2025, 11, 20), datetime(2025, 12, 5), datetime(2025, 12, 3), "=HOJE()-F5", "=G5-HOJE()", "100%", "Documentação"],
    [5, "Teste de integração API", "Carlos Lima", "Alta", "Em Andamento", 
     datetime(2025, 12, 7), datetime(2025, 12, 14), "", "=HOJE()-F6", "=G6-HOJE()", "45%", "Teste"],
]

# Inserir dados
for row_num, row_data in enumerate(dados_exemplo, 2):
    for col_num, value in enumerate(row_data, 1):
        cell = ws_demandas.cell(row=row_num, column=col_num)
        cell.value = value
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Formatar datas
        if col_num in [6, 7, 8] and isinstance(value, datetime):
            cell.number_format = 'DD/MM/YYYY'

# Ajustar largura das colunas
column_widths = [8, 35, 20, 12, 15, 15, 15, 15, 15, 15, 12, 18]
for i, width in enumerate(column_widths, 1):
    ws_demandas.column_dimensions[get_column_letter(i)].width = width

# Criar validações de dados (listas suspensas)
# Validação de Prioridade
dv_prioridade = DataValidation(type="list", formula1='"Crítica,Alta,Média,Baixa"', allow_blank=False)
dv_prioridade.error = 'Selecione uma prioridade válida'
dv_prioridade.errorTitle = 'Entrada Inválida'
ws_demandas.add_data_validation(dv_prioridade)
dv_prioridade.add(f'D2:D100')

# Validação de Status
dv_status = DataValidation(type="list", formula1='"Planejada,Em Andamento,Concluída,Cancelada,Bloqueada"', allow_blank=False)
dv_status.error = 'Selecione um status válido'
dv_status.errorTitle = 'Entrada Inválida'
ws_demandas.add_data_validation(dv_status)
dv_status.add(f'E2:E100')

# Validação de Categoria
dv_categoria = DataValidation(type="list", formula1='"Desenvolvimento,Correção,Teste,Documentação,Infraestrutura,Análise"', allow_blank=False)
dv_categoria.error = 'Selecione uma categoria válida'
dv_categoria.errorTitle = 'Entrada Inválida'
ws_demandas.add_data_validation(dv_categoria)
dv_categoria.add(f'L2:L100')

# Formatação condicional para Status
from openpyxl.formatting.rule import CellIsRule

# Verde para Concluída
ws_demandas.conditional_formatting.add(
    'E2:E100',
    CellIsRule(operator='equal', formula=['"Concluída"'], 
               fill=PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
               font=Font(color="006100"))
)

# Amarelo para Em Andamento
ws_demandas.conditional_formatting.add(
    'E2:E100',
    CellIsRule(operator='equal', formula=['"Em Andamento"'], 
               fill=PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
               font=Font(color="9C6500"))
)

# Vermelho para Bloqueada
ws_demandas.conditional_formatting.add(
    'E2:E100',
    CellIsRule(operator='equal', formula=['"Bloqueada"'], 
               fill=PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
               font=Font(color="9C0006"))
)

# Azul para Planejada
ws_demandas.conditional_formatting.add(
    'E2:E100',
    CellIsRule(operator='equal', formula=['"Planejada"'], 
               fill=PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid"),
               font=Font(color="1F4E78"))
)

# ==================== ABA 2: DASHBOARD ====================
ws_dashboard = wb.create_sheet("Dashboard")

# Título do Dashboard
ws_dashboard.merge_cells('A1:F2')
title_cell = ws_dashboard['A1']
title_cell.value = "📊 DASHBOARD - GERENCIAMENTO DE DEMANDAS"
title_cell.font = Font(size=18, bold=True, color="FFFFFF")
title_cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
title_cell.alignment = Alignment(horizontal="center", vertical="center")

# Indicadores - KPIs
kpi_row = 4
kpi_labels = ["Total de Demandas", "Em Andamento", "Concluídas", "Atrasadas", "Taxa de Conclusão"]
kpi_formulas = [
    "=CONT.VALORES('Controle de Demandas'!A2:A100)",
    "=CONT.SE('Controle de Demandas'!E:E,\"Em Andamento\")",
    "=CONT.SE('Controle de Demandas'!E:E,\"Concluída\")",
    "=CONT.SE('Controle de Demandas'!J:J,\"<0\")",
    "=SE(B4=0,\"0%\",TEXTO(D4/B4,\"0%\"))"
]

kpi_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
for i, (label, formula) in enumerate(zip(kpi_labels, kpi_formulas)):
    # Label
    label_cell = ws_dashboard.cell(row=kpi_row, column=i*2+1)
    label_cell.value = label
    label_cell.font = Font(bold=True, size=10)
    label_cell.alignment = Alignment(horizontal="center", vertical="center")
    label_cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    label_cell.border = border
    
    # Valor
    value_cell = ws_dashboard.cell(row=kpi_row+1, column=i*2+1)
    value_cell.value = formula
    value_cell.font = Font(bold=True, size=14, color="366092")
    value_cell.alignment = Alignment(horizontal="center", vertical="center")
    value_cell.fill = kpi_fill
    value_cell.border = border
    
    ws_dashboard.merge_cells(start_row=kpi_row, start_column=i*2+1, end_row=kpi_row, end_column=i*2+2)
    ws_dashboard.merge_cells(start_row=kpi_row+1, start_column=i*2+1, end_row=kpi_row+1, end_column=i*2+2)

# Dados para gráficos - Status
chart_data_start = 8
ws_dashboard['A8'] = "Status"
ws_dashboard['A8'].font = Font(bold=True, size=12)
ws_dashboard['B8'] = "Quantidade"
ws_dashboard['B8'].font = Font(bold=True, size=12)

status_labels = ["Planejada", "Em Andamento", "Concluída", "Cancelada", "Bloqueada"]
for i, status in enumerate(status_labels, 9):
    ws_dashboard[f'A{i}'] = status
    ws_dashboard[f'B{i}'] = f'=CONT.SE(\'Controle de Demandas\'!E:E,"{status}")'

# Gráfico de Pizza - Status
pie = PieChart()
labels = Reference(ws_dashboard, min_col=1, min_row=9, max_row=13)
data = Reference(ws_dashboard, min_col=2, min_row=8, max_row=13)
pie.add_data(data, titles_from_data=True)
pie.set_categories(labels)
pie.title = "Distribuição por Status"
pie.height = 10
pie.width = 15
ws_dashboard.add_chart(pie, "D8")

# Dados para gráfico - Prioridade
ws_dashboard['A16'] = "Prioridade"
ws_dashboard['A16'].font = Font(bold=True, size=12)
ws_dashboard['B16'] = "Quantidade"
ws_dashboard['B16'].font = Font(bold=True, size=12)

prioridade_labels = ["Crítica", "Alta", "Média", "Baixa"]
for i, prioridade in enumerate(prioridade_labels, 17):
    ws_dashboard[f'A{i}'] = prioridade
    ws_dashboard[f'B{i}'] = f'=CONT.SE(\'Controle de Demandas\'!D:D,"{prioridade}")'

# Gráfico de Barras - Prioridade
bar = BarChart()
bar.type = "col"
bar.title = "Demandas por Prioridade"
bar.y_axis.title = 'Quantidade'
bar.x_axis.title = 'Prioridade'

labels_pri = Reference(ws_dashboard, min_col=1, min_row=17, max_row=20)
data_pri = Reference(ws_dashboard, min_col=2, min_row=16, max_row=20)
bar.add_data(data_pri, titles_from_data=True)
bar.set_categories(labels_pri)
bar.height = 10
bar.width = 15
ws_dashboard.add_chart(bar, "D22")

# Dados para gráfico - Categoria
ws_dashboard['A24'] = "Categoria"
ws_dashboard['A24'].font = Font(bold=True, size=12)
ws_dashboard['B24'] = "Quantidade"
ws_dashboard['B24'].font = Font(bold=True, size=12)

categoria_labels = ["Desenvolvimento", "Correção", "Teste", "Documentação", "Infraestrutura", "Análise"]
for i, categoria in enumerate(categoria_labels, 25):
    ws_dashboard[f'A{i}'] = categoria
    ws_dashboard[f'B{i}'] = f'=CONT.SE(\'Controle de Demandas\'!L:L,"{categoria}")'

# Gráfico de Barras - Categoria
bar_cat = BarChart()
bar_cat.type = "col"
bar_cat.title = "Demandas por Categoria"
bar_cat.y_axis.title = 'Quantidade'
bar_cat.x_axis.title = 'Categoria'

labels_cat = Reference(ws_dashboard, min_col=1, min_row=25, max_row=30)
data_cat = Reference(ws_dashboard, min_col=2, min_row=24, max_row=30)
bar_cat.add_data(data_cat, titles_from_data=True)
bar_cat.set_categories(labels_cat)
bar_cat.height = 10
bar_cat.width = 15
ws_dashboard.add_chart(bar_cat, "D38")

# Ajustar larguras
for col in ['A', 'B', 'C', 'D', 'E', 'F']:
    ws_dashboard.column_dimensions[col].width = 18

# ==================== ABA 3: RELATÓRIOS ====================
ws_relatorio = wb.create_sheet("Relatórios")

# Título
ws_relatorio.merge_cells('A1:H1')
title_rel = ws_relatorio['A1']
title_rel.value = "📋 RELATÓRIOS E ANÁLISES"
title_rel.font = Font(size=16, bold=True, color="FFFFFF")
title_rel.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
title_rel.alignment = Alignment(horizontal="center", vertical="center")

# Demandas Atrasadas
ws_relatorio['A3'] = "⚠️ DEMANDAS ATRASADAS"
ws_relatorio['A3'].font = Font(size=13, bold=True, color="C00000")
ws_relatorio.merge_cells('A3:H3')

rel_headers = ["ID", "Demanda", "Responsável", "Prazo", "Dias de Atraso", "Status", "Prioridade"]
for i, header in enumerate(rel_headers, 1):
    cell = ws_relatorio.cell(row=4, column=i)
    cell.value = header
    cell.font = Font(bold=True)
    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    cell.alignment = Alignment(horizontal="center")
    cell.border = border

# Fórmula para listar atrasadas (exemplo manual - em produção usar FILTER ou tabela dinâmica)
ws_relatorio['A5'] = "=SE('Controle de Demandas'!J2<0,'Controle de Demandas'!A2,\"\")"
ws_relatorio['B5'] = "=SE('Controle de Demandas'!J2<0,'Controle de Demandas'!B2,\"\")"
ws_relatorio['C5'] = "=SE('Controle de Demandas'!J2<0,'Controle de Demandas'!C2,\"\")"
ws_relatorio['D5'] = "=SE('Controle de Demandas'!J2<0,'Controle de Demandas'!G2,\"\")"
ws_relatorio['E5'] = "=SE('Controle de Demandas'!J2<0,ABS('Controle de Demandas'!J2),\"\")"
ws_relatorio['F5'] = "=SE('Controle de Demandas'!J2<0,'Controle de Demandas'!E2,\"\")"
ws_relatorio['G5'] = "=SE('Controle de Demandas'!J2<0,'Controle de Demandas'!D2,\"\")"

# Demandas por Responsável
ws_relatorio['A10'] = "👤 RESUMO POR RESPONSÁVEL"
ws_relatorio['A10'].font = Font(size=13, bold=True, color="366092")
ws_relatorio.merge_cells('A10:D10')

ws_relatorio['A11'] = "Responsável"
ws_relatorio['B11'] = "Total"
ws_relatorio['C11'] = "Concluídas"
ws_relatorio['D11'] = "Em Andamento"

for col in ['A', 'B', 'C', 'D']:
    cell = ws_relatorio[f'{col}11']
    cell.font = Font(bold=True)
    cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    cell.alignment = Alignment(horizontal="center")
    cell.border = border

# Exemplos de dados (em produção, usar fórmulas dinâmicas)
responsaveis_exemplo = ["João Silva", "Maria Santos", "Pedro Costa", "Ana Oliveira", "Carlos Lima"]
for i, resp in enumerate(responsaveis_exemplo, 12):
    ws_relatorio[f'A{i}'] = resp
    ws_relatorio[f'B{i}'] = f'=CONT.SE(\'Controle de Demandas\'!C:C,"{resp}")'
    ws_relatorio[f'C{i}'] = f'=CONT.SES(\'Controle de Demandas\'!C:C,"{resp}",\'Controle de Demandas\'!E:E,"Concluída")'
    ws_relatorio[f'D{i}'] = f'=CONT.SES(\'Controle de Demandas\'!C:C,"{resp}",\'Controle de Demandas\'!E:E,"Em Andamento")'

# Ajustar larguras
for i in range(1, 9):
    ws_relatorio.column_dimensions[get_column_letter(i)].width = 20

# ==================== ABA 4: CONFIGURAÇÕES ====================
ws_config = wb.create_sheet("Configurações")

ws_config['A1'] = "⚙️ CONFIGURAÇÕES DO SISTEMA"
ws_config['A1'].font = Font(size=14, bold=True)
ws_config.merge_cells('A1:C1')

config_items = [
    ["", "Opção", "Valor"],
    ["Prioridades:", "Crítica", ""],
    ["", "Alta", ""],
    ["", "Média", ""],
    ["", "Baixa", ""],
    ["", "", ""],
    ["Status:", "Planejada", ""],
    ["", "Em Andamento", ""],
    ["", "Concluída", ""],
    ["", "Cancelada", ""],
    ["", "Bloqueada", ""],
    ["", "", ""],
    ["Categorias:", "Desenvolvimento", ""],
    ["", "Correção", ""],
    ["", "Teste", ""],
    ["", "Documentação", ""],
    ["", "Infraestrutura", ""],
    ["", "Análise", ""],
]

for row_idx, row_data in enumerate(config_items, 3):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws_config.cell(row=row_idx, column=col_idx)
        cell.value = value
        if row_idx == 3:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

ws_config.column_dimensions['A'].width = 15
ws_config.column_dimensions['B'].width = 20
ws_config.column_dimensions['C'].width = 30

# Instruções
ws_config['A23'] = "📖 INSTRUÇÕES DE USO:"
ws_config['A23'].font = Font(bold=True, size=12)

instrucoes = [
    "1. Use a aba 'Controle de Demandas' para adicionar e gerenciar suas tarefas",
    "2. Preencha todos os campos obrigatórios (ID, Demanda, Responsável, etc.)",
    "3. Use as listas suspensas para Prioridade, Status e Categoria",
    "4. As fórmulas de Dias Decorridos e Dias Restantes são automáticas",
    "5. O Dashboard atualiza automaticamente conforme você adiciona demandas",
    "6. Cores no Status: Verde=Concluída, Amarelo=Em Andamento, Vermelho=Bloqueada",
    "7. Use a aba Relatórios para visualizar análises específicas",
]

for i, instrucao in enumerate(instrucoes, 24):
    ws_config[f'A{i}'] = instrucao
    ws_config.merge_cells(f'A{i}:C{i}')

# Salvar arquivo
filename = "Sistema_Gerenciamento_Demandas.xlsx"
wb.save(filename)
print(f"✅ Sistema criado com sucesso: {filename}")
print("\n📊 Recursos implementados:")
print("   ✓ Controle de Demandas com validação de dados")
print("   ✓ Fórmulas automáticas para cálculos de prazos")
print("   ✓ Dashboard com KPIs e gráficos dinâmicos")
print("   ✓ Formatação condicional por status")
print("   ✓ Relatórios e análises")
print("   ✓ Configurações e instruções")
print("\n🎯 Abra o arquivo no Excel para começar a usar!")
