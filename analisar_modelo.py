"""
Analisador do arquivo modelo_luma.xlsx
"""
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# Carregar o arquivo
wb = load_workbook('modelo_luma.xlsx', data_only=False)

print("="*80)
print("ANÁLISE DO ARQUIVO: modelo_luma.xlsx")
print("="*80)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n{'='*80}")
    print(f"ABA: {sheet_name}")
    print(f"{'='*80}")
    
    # Dimensões
    print(f"\nDimensões utilizadas: {ws.dimensions}")
    
    # Células mescladas
    if ws.merged_cells:
        print(f"\nCélulas mescladas: {len(ws.merged_cells.ranges)} áreas")
        for merged_range in list(ws.merged_cells.ranges)[:10]:  # Mostrar primeiras 10
            print(f"  - {merged_range}")
    
    # Análise de conteúdo (primeiras 30 linhas)
    print(f"\nConteúdo (primeiras 30 linhas):")
    print("-"*80)
    
    max_row = min(ws.max_row, 30)
    max_col = min(ws.max_column, 15)
    
    for row in range(1, max_row + 1):
        row_data = []
        has_content = False
        
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            value = cell.value
            
            if value is not None and value != "":
                has_content = True
                # Verificar se tem fórmula
                if isinstance(value, str) and value.startswith('='):
                    row_data.append(f"[FÓRMULA: {value[:50]}...]" if len(value) > 50 else f"[FÓRMULA: {value}]")
                else:
                    row_data.append(str(value)[:40])
            else:
                row_data.append("")
        
        if has_content:
            print(f"Linha {row:2d}: {' | '.join(row_data)}")
    
    # Formatações especiais
    print(f"\n--- Análise de Formatação ---")
    
    # Verificar cabeçalhos (primeira linha)
    if ws.max_row >= 1:
        print("\nPrimeira linha (possível cabeçalho):")
        for col in range(1, min(ws.max_column + 1, 20)):
            cell = ws.cell(row=1, column=col)
            if cell.value:
                col_letter = get_column_letter(col)
                print(f"  {col_letter}: {cell.value}")
                
                # Informações de estilo
                if cell.fill and cell.fill.start_color:
                    print(f"      Cor de fundo: {cell.fill.start_color.rgb if hasattr(cell.fill.start_color, 'rgb') else 'N/A'}")
                if cell.font:
                    print(f"      Fonte: {'Negrito' if cell.font.bold else 'Normal'}, Tamanho: {cell.font.size}")
    
    # Validações de dados
    if hasattr(ws, 'data_validations') and ws.data_validations.dataValidation:
        print(f"\nValidações de dados encontradas: {len(ws.data_validations.dataValidation)}")
        for dv in ws.data_validations.dataValidation[:5]:
            print(f"  - Tipo: {dv.type}, Fórmula: {dv.formula1}")
    
    # Largura das colunas
    print(f"\nLargura das colunas:")
    for col in range(1, min(ws.max_column + 1, 20)):
        col_letter = get_column_letter(col)
        if col_letter in ws.column_dimensions:
            width = ws.column_dimensions[col_letter].width
            if width:
                print(f"  Coluna {col_letter}: {width}")
    
    print("\n")

wb.close()
print("="*80)
print("ANÁLISE CONCLUÍDA")
print("="*80)
